#!/usr/bin/env python3
"""Parse the ARKA/AIM "TB & Detail" workbook into per-company GL detail CSVs.

The workbook holds, per company, a trial balance sheet (``TB AIM`` / ``TB ARKA``)
and two detail sheets (``BS *`` / ``PL *``). Each detail sheet is a stack of
per-account blocks: a header row, the transaction rows, then three control rows
(``Total`` / ``TB`` / ``diff``). Retained Earnings has no detail block -- it is
the balancing figure, and this script takes it straight from the TB.

Usage::

    python3 tools/parse_arkaaim_begbal_detail.py --check     # reconcile only
    python3 tools/parse_arkaaim_begbal_detail.py --write     # also write the CSVs

Output CSVs (source of truth for scripts/tenants/arkaaim/load_begbal_detail.py):
    addons/_tenants/custom_arka_aim_opening_balance/data/opening_detail_aim.csv
    addons/_tenants/custom_arka_aim_opening_balance/data/opening_detail_arka.csv
"""

import argparse
import csv
import datetime
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
WORKBOOK = REPO / "imports/arka_aim/TB_Detail_AIM_ARKA_052026_040826.xlsx"
OUT_DIR = REPO / "addons/_tenants/custom_arka_aim_opening_balance/data"

# The ARKA detail sheets spell three accounts with codes that do not exist in the
# chart of accounts; the TB (and the accounts created by the opening-balance
# module) use the codes on the right. Amounts match exactly.
ARKA_CODE_FIX = {
    "1103019870": "1103019290",  # BCA - IDR-268.262.6268 - Main Bank
    "1103019900": "1103019300",  # BCA - IDR-268.222.9595 - Main Bank
    "1105020003": "1105020007",  # Time Deposit BRI
}

# Accounts whose detail block is missing get a single line straight from the TB.
# In practice that is only Retained Earnings, which the workbook leaves as the
# balancing figure -- but the rule is derived, not hardcoded, so a future export
# that drops another block is handled the same way.

# The workbook's own per-account control rows tolerate a few rupiah of rounding
# (e.g. account 2103300001 shows "Selisih -4"). Anything larger is a data error.
ROUNDING_TOLERANCE = 100.0

COMPANIES = OrderedDict(
    [
        ("AIM", {"tb": "TB AIM", "detail": ["BS AIM", "PL AIM"], "codefix": {}}),
        ("ARKA", {"tb": "TB ARKA", "detail": ["BS ARKA", "PL ARKA"], "codefix": ARKA_CODE_FIX}),
    ]
)

ACCOUNT_RE = re.compile(r"^\d{10}$")


def _num(value):
    """Excel cell -> float. Blanks, dashes and stray strings become 0.0."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", " ")
    if text in ("", "-", "–", "—"):
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "").replace(" ", "")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative else number


def _text(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date(value):
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    text = _text(value)
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def read_tb(worksheet):
    """Return {code: (name, signed_balance)} from a TB sheet; debit positive."""
    balances = OrderedDict()
    for row in worksheet.iter_rows(values_only=True):
        code = _text(row[0] if row else None)
        if not ACCOUNT_RE.match(code):
            continue
        name = _text(row[1]) if len(row) > 1 else ""
        debit = _num(row[2]) if len(row) > 2 else 0.0
        credit = _num(row[3]) if len(row) > 3 else 0.0
        balances[code] = (name, debit - credit)
    return balances


def read_detail(worksheet, codefix):
    """Return the transaction rows of one detail sheet.

    Columns are fixed by the export: Document No, Transaction Date, Transaction
    Period, Account Number, Account Description, Debit, Credit, Amount, Notes,
    Company BP, Journal No. Control rows (Total / TB / diff) carry no account
    number, so filtering on a 10-digit account code drops them.
    """
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        cells = list(row) + [None] * (11 - len(row))
        code = _text(cells[3])
        if not ACCOUNT_RE.match(code):
            continue
        # The signed Amount column is what the sheet's own block totals add up,
        # and it wins where it disagrees with Debit/Credit (row "Accrue Gaji
        # 05-2026" carries the monthly figure in Debit but the 5-month total in
        # Amount, and the block total follows Amount).
        amount = _num(cells[7])
        if not amount:
            amount = _num(cells[5]) - _num(cells[6])
        debit, credit = (amount, 0.0) if amount > 0 else (0.0, -amount)
        rows.append(
            {
                "code": codefix.get(code, code),
                "code_source": code,
                "txn_date": _date(cells[1]),
                "doc_no": _text(cells[0]),
                "journal_no": _text(cells[10]),
                "partner": _text(cells[9]),
                "notes": _text(cells[8]),
                "debit": round(debit, 2),
                "credit": round(credit, 2),
            }
        )
    return rows


def _line(code, notes, balance, doc_no):
    """Build an extra CSV line carrying a signed balance (debit positive)."""
    return {
        "code": code,
        "code_source": code,
        "txn_date": "2026-05-31",
        "doc_no": doc_no,
        "journal_no": "",
        "partner": "",
        "notes": notes,
        "debit": round(balance, 2) if balance > 0 else 0.0,
        "credit": round(-balance, 2) if balance < 0 else 0.0,
    }


def reconcile(company, tb, rows):
    """Force every account to its TB balance; return (ok, extra_lines, report).

    Accounts whose detail already ties to the TB are left alone. Accounts the
    workbook rounds by a few rupiah get an explicit adjustment line. Accounts
    with no detail block at all (Retained Earnings) get one line from the TB.
    """
    report, ok = [], True
    detail_by_code = OrderedDict()
    for row in rows:
        detail_by_code.setdefault(row["code"], 0.0)
        detail_by_code[row["code"]] += row["debit"] - row["credit"]

    extra_lines = []
    for code, (name, tb_balance) in tb.items():
        got = detail_by_code.get(code)
        if got is None:
            report.append(f"  no detail block for {code} {name} -> single TB line {tb_balance:,.2f}")
            extra_lines.append(_line(code, f"Saldo awal {name} (dari TB, tanpa blok detail)", tb_balance, "TB"))
            continue
        diff = round(tb_balance - got, 2)
        if not diff:
            continue
        if abs(diff) > ROUNDING_TOLERANCE:
            ok = False
            report.append(f"  DIFF {code} {name}: detail {got:,.2f} vs TB {tb_balance:,.2f} (diff {diff:,.2f})")
            continue
        report.append(f"  rounding {code} {name}: detail {got:,.2f} vs TB {tb_balance:,.2f} -> adjust {diff:,.2f}")
        extra_lines.append(_line(code, f"Pembulatan ke TB {name}", diff, "PEMBULATAN"))

    for code in detail_by_code:
        if code not in tb:
            ok = False
            report.append(f"  UNKNOWN account {code} present in detail but not in TB")

    total = sum(row["debit"] - row["credit"] for row in rows + extra_lines)
    tb_debit = sum(balance for _name, balance in tb.values() if balance > 0)
    report.append(f"  TB total debit          : {tb_debit:,.2f}")
    report.append(f"  loaded total (D-C)      : {total:,.2f}")
    report.append(f"  rows                    : {len(rows)} detail + {len(extra_lines)} adjustment/TB")
    if abs(total) > 0.005:
        ok = False
        report.append("  MOVE DOES NOT BALANCE -- do not post")
    return ok, extra_lines, report, total


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true", help="reconcile only (default)")
    parser.add_argument("--write", action="store_true", help="write the CSVs")
    parser.add_argument("--workbook", default=str(WORKBOOK))
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.workbook, data_only=True)
    all_ok = True
    for company, spec in COMPANIES.items():
        tb = read_tb(workbook[spec["tb"]])
        rows = []
        for sheet in spec["detail"]:
            rows += read_detail(workbook[sheet], spec["codefix"])

        remapped = {r["code_source"] for r in rows if r["code_source"] != r["code"]}
        ok, extra_lines, report, total = reconcile(company, tb, rows)
        all_ok &= ok

        print(f"== {company} ==")
        if remapped:
            fixes = ", ".join(f"{src}->{ARKA_CODE_FIX[src]}" for src in sorted(remapped))
            print(f"  account codes remapped  : {fixes}")
        print("\n".join(report))
        print(f"  RESULT: {'OK' if ok else 'FAILED'}")

        if args.write:
            if not ok:
                print(f"  refusing to write {company} CSV while checks fail")
                continue
            out = OUT_DIR / f"opening_detail_{company.lower()}.csv"
            fields = ["code", "txn_date", "doc_no", "journal_no", "partner", "notes", "debit", "credit"]
            with out.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows + extra_lines)
            print(f"  wrote {out.relative_to(REPO)} ({len(rows) + len(extra_lines)} rows)")
        print()

    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
