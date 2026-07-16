# -*- coding: utf-8 -*-
"""
Generator skenario end-to-end Arka (jasa drone show) x Aim (rental drone).
Output: docs/projects/arka-aim/Skenario-Arka-Aim-Drone-Show-Rental.xlsx

Skenario: Customer (Event Organizer) menyewa jasa drone show ke ARKA.
ARKA menyewa unit drone + kru teknis ke AIM. Termasuk survey lokasi, BAST,
AR (piutang) ARKA, AP (utang) ARKA, AR AIM, serta jurnal akuntansi 2 sisi.
Dipetakan ke objek/modul Odoo (Platform).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Palet & style
# ----------------------------------------------------------------------------
C_TITLE   = "1F3864"   # biru tua
C_HEAD    = "2E5496"   # biru header
C_HEAD2   = "548235"   # hijau (akuntansi)
C_ARKA    = "C55A11"   # oranye Arka
C_AIM     = "2E75B6"   # biru Aim
C_CUST    = "7030A0"   # ungu customer
C_BAND    = "DDEBF7"   # baris belang terang
C_BAND2   = "FFFFFF"
C_NOTE    = "FFF2CC"   # kuning catatan
C_TOTAL   = "FCE4D6"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WHITE = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
HDR   = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
BODY  = Font(name="Calibri", size=10, color="000000")
BODYB = Font(name="Calibri", size=10, color="000000", bold=True)
SMALL = Font(name="Calibri", size=9, color="595959")

TOP = Alignment(vertical="top", wrap_text=True)
TOPL= Alignment(vertical="top", horizontal="left", wrap_text=True)
CEN = Alignment(vertical="center", horizontal="center", wrap_text=True)
RIGHT = Alignment(vertical="top", horizontal="right", wrap_text=True)

def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)

def style_header_row(ws, row, ncols, color=C_HEAD):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(color)
        cell.font = HDR
        cell.alignment = CEN
        cell.border = BORDER

def write_table(ws, start_row, headers, rows, widths, header_color=C_HEAD,
                band=True, money_cols=None, center_cols=None):
    money_cols = money_cols or []
    center_cols = center_cols or []
    ncols = len(headers)
    # header
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=j, value=h)
        cell.fill = fill(header_color)
        cell.font = HDR
        cell.alignment = CEN
        cell.border = BORDER
    # widths
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    # body
    r = start_row + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = BODY
            cell.border = BORDER
            if j in money_cols:
                cell.alignment = RIGHT
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
            elif j in center_cols:
                cell.alignment = CEN
            else:
                cell.alignment = TOP
            if band:
                cell.fill = fill(C_BAND if i % 2 == 0 else C_BAND2)
        r += 1
    return r  # next free row

def sheet_title(ws, text, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name="Calibri", size=15, color="FFFFFF", bold=True)
    c.fill = fill(C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
    s.fill = fill(C_HEAD)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 18

wb = Workbook()

# ============================================================================
# 00 - RINGKASAN
# ============================================================================
ws = wb.active
ws.title = "00-Ringkasan"
ws.sheet_view.showGridLines = False
sheet_title(ws, "SKENARIO BISNIS: ARKA x AIM — DRONE SHOW & RENTAL DRONE",
            "End-to-end: Mastering -> Survey -> Sales -> Rental -> BAST -> AR/AP -> Akuntansi | Platform Odoo", 6)

r = 4
ws.cell(row=r, column=1, value="1. PARA PIHAK").font = Font(size=12, bold=True, color=C_TITLE)
r += 1
parties = [
    ["Kode", "Pihak", "Peran", "Posisi dalam Transaksi"],
    ["ARKA", "PT Aero Reksa Kreasi Angkasa", "Penyedia jasa DRONE SHOW (event)", "Penjual jasa ke Customer (AR) + Penyewa drone dari Aim (AP)"],
    ["AIM", "PT Aero Inovasi Media", "Penyedia RENTAL unit drone + kru teknis", "Penjual jasa sewa ke Arka (AR Aim)"],
    ["CUST", "PT Gemilang Event Organizer", "Pembeli jasa drone show", "Pemberi kerja / pembeli jasa ke Arka"],
]
r = write_table(ws, r, parties[0], parties[1:], [10, 26, 40, 50],
                header_color=C_HEAD, center_cols=[1])
r += 1

ws.cell(row=r, column=1, value="2. ALUR RINGKAS (RANTAI NILAI)").font = Font(size=12, bold=True, color=C_TITLE)
r += 1
flow = [
    ["#", "Fase", "Ringkasan", "Hasil / Dokumen Kunci"],
    [1, "Mastering", "Setup company, CoA, pajak, jurnal, produk, kontak (Arka & Aim)", "Master data siap"],
    [2, "Inquiry & Survey", "Customer minta drone show; Arka survey lokasi (cek area terbang, GPS, perizinan)", "Berita Acara Survey Lokasi"],
    [3, "Quotation & SO Arka", "Arka kirim penawaran -> Customer setuju -> Sales Order", "SO ARKA/2026/xxxx"],
    [4, "Procurement Rental", "Arka butuh unit drone -> PO/RFQ ke Aim -> Aim buat Rental Order", "PO ARKA + Rental Order AIM"],
    [5, "Mobilisasi & Eksekusi", "Aim kirim unit (DO sewa) -> Arka jalankan drone show di lokasi", "Delivery Order sewa, Laporan Show"],
    [6, "BAST", "BAST Aim->Arka (serah unit & pengembalian) + BAST Arka->Customer (serah hasil jasa)", "2 dokumen BAST bernomor"],
    [7, "AR & AP", "Aim tagih Arka (AP Arka) ; Arka tagih Customer (AR Arka)", "Faktur AIM, Faktur ARKA, e-Faktur PPN"],
    [8, "Pembayaran & Pajak", "Customer bayar Arka (potong PPh23) ; Arka bayar Aim (potong PPh23) ; setor PPN/PPh", "Bukti bayar, bukti potong PPh23"],
    [9, "Akuntansi", "Posting jurnal otomatis 2 sisi; rekonsiliasi; laba kotor Arka", "Jurnal, Laporan L/R"],
]
r = write_table(ws, r, flow[0], flow[1:], [5, 22, 60, 38], center_cols=[1])
r += 1

ws.cell(row=r, column=1, value="3. ASUMSI ANGKA & PAJAK").font = Font(size=12, bold=True, color=C_TITLE)
r += 1
asum = [
    ["Item", "Nilai", "Keterangan"],
    ["DPP Jasa Drone Show (Arka -> Customer)", 150_000_000, "Harga jasa sebelum pajak"],
    ["DPP Sewa Drone (Aim -> Arka)", 60_000_000, "Harga sewa unit+kru sebelum pajak"],
    ["Tarif PPN", "11%", "PPN Keluaran/Masukan (sesuaikan bila 12%)"],
    ["Tarif PPh 23 - Jasa", "2%", "Atas jasa drone show (dipotong Customer)"],
    ["Tarif PPh 23 - Sewa", "2%", "Atas sewa selain tanah/bangunan (dipotong Arka)"],
    ["Termin Pembayaran", "DP 30% + Pelunasan 70%", "Atau Net 14 hari setelah BAST (opsional)"],
    ["Margin Kotor Arka", 90_000_000, "150jt jasa - 60jt sewa (sebelum biaya lain)"],
]
r = write_table(ws, r, asum[0], asum[1:], [42, 28, 55], money_cols=[2])
# format the money rows that are numbers
r += 1

ws.cell(row=r, column=1, value="4. CATATAN PENTING").font = Font(size=12, bold=True, color=C_TITLE)
r += 1
notes = [
    "• Skenario ini multi-company di satu Platform Odoo (Arka & Aim 2 company). Bisa juga 2 instance terpisah — kolom 'Company' menandai bukunya.",
    "• BAST memakai sequence nyata (ref: commit feat/bast). 2 BAST: (a) BAST sewa unit Aim->Arka, (b) BAST hasil jasa Arka->Customer.",
    "• Survey lokasi = prasyarat sebelum Quotation difinalkan (kelayakan area terbang & perizinan penerbangan/AirNav bila perlu).",
    "• PPh 23 = pajak dipotong PEMBELI. Customer potong PPh23 jasa Arka; Arka potong PPh23 sewa Aim. Penjual terima bukti potong = kredit pajak.",
    "• Semua nomor dokumen contoh (ARKA/2026/...) — pada Odoo digenerate oleh ir.sequence per jurnal/modul. BAST pakai 1 sequence global BAST/YYYY/NNNNN, dibedakan field 'kind'.",
    "• KEPUTUSAN PRODUK SEWA: unit drone = BARANG (Goods, serial-tracked), BUKAN jasa. Pendapatan tetap diakui sbg JASA SEWA via Income Account (4-10200). Alasan & detail lengkap di sheet 03.",
    "• Kolom 'Navigasi Odoo (Klik / Menu)' di tiap sheet langkah = jalur menu/klik di Odoo (mis. Sales > Orders > Quotations > New). Penanda [DI LUAR ODOO] = langkah dikerjakan di luar sistem (lapangan/pihak ketiga), hasilnya cukup dicatat/di-upload ke Odoo.",
]
for n in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws.cell(row=r, column=1, value=n)
    cell.font = BODY
    cell.alignment = TOPL
    cell.fill = fill(C_NOTE)
    cell.border = BORDER
    ws.row_dimensions[r].height = 30
    r += 1

# ============================================================================
# 01 - PERAN & PIC
# ============================================================================
ws = wb.create_sheet("01-Peran & PIC")
ws.sheet_view.showGridLines = False
sheet_title(ws, "DAFTAR PERAN & PIC (PERSON IN CHARGE)",
            "Kode PIC dipakai sebagai referensi pada seluruh sheet langkah", 6)
pic = [
    ["Kode PIC", "Perusahaan", "Jabatan / Peran", "Tanggung Jawab Utama", "Akses Odoo (Grup)"],
    ["ARK-DIR", "ARKA", "Direktur / Owner", "Approve quotation besar, approve PO sewa, tanda tangan BAST", "Sales Manager, Approver"],
    ["ARK-ADM", "ARKA", "Admin / Master Data", "Input master produk, kontak, pajak, jurnal", "Settings (terbatas), Contacts"],
    ["ARK-SLS", "ARKA", "Sales / Account Executive", "Handle inquiry customer, buat quotation & SO, follow up", "Sales User"],
    ["ARK-SVY", "ARKA", "Tim Survey Lokasi", "Survey lokasi, isi Berita Acara Survey, foto & koordinat GPS", "Project/Field User"],
    ["ARK-OPS", "ARKA", "Operasional / Drone Pilot", "Eksekusi drone show, koordinasi mobilisasi, laporan show", "Inventory/Project User"],
    ["ARK-PUR", "ARKA", "Purchasing", "Buat RFQ/PO sewa ke Aim, negosiasi harga sewa", "Purchase User"],
    ["ARK-FIN", "ARKA", "Finance & Accounting", "Buat faktur AR, validasi tagihan AP, bayar, potong/setor pajak, jurnal", "Accounting / Billing"],
    ["AIM-ADM", "AIM", "Admin / Master Data", "Input master unit drone, produk sewa, kontak Arka", "Settings (terbatas)"],
    ["AIM-RNT", "AIM", "Rental Officer / Sales", "Terima order sewa, buat Rental Order, jadwal unit", "Sales/Rental User"],
    ["AIM-OPS", "AIM", "Operasional / Teknisi", "Cek & kirim unit (DO sewa), dampingi teknis, terima pengembalian", "Inventory User"],
    ["AIM-FIN", "AIM", "Finance & Accounting", "Buat faktur sewa ke Arka, terima bukti potong PPh23, jurnal", "Accounting / Billing"],
    ["CUS-PIC", "CUSTOMER", "PIC Event Organizer", "Sumber kebutuhan, dampingi survey, tanda tangan BAST, bayar", "(eksternal)"],
]
write_table(ws, 4, pic[0], pic[1:], [12, 14, 26, 52, 24], center_cols=[1, 2])

# ============================================================================
# 02 - MASTER DATA (MASTERING)
# ============================================================================
ws = wb.create_sheet("02-Master Data")
ws.sheet_view.showGridLines = False
sheet_title(ws, "FASE 1 — MASTERING (SETUP MASTER DATA)",
            "Dilakukan SEKALI di awal sebelum transaksi | Objek Odoo & PIC dicantumkan", 9)
md = [
    ["No", "Company", "Aktivitas Mastering", "Detail / Field Kunci", "Objek/Modul Odoo", "PIC", "Output", "Catatan", "Navigasi Odoo (Klik / Menu)"],
    ["M-01", "ARKA", "Buat Company Arka", "Nama: PT Aero Reksa Kreasi Angkasa; NPWP; alamat; mata uang IDR", "res.company / Settings", "ARK-ADM", "Company aktif", "Multi-company", "Settings > Users & Companies > Companies > New"],
    ["M-02", "AIM", "Buat Company Aim", "Nama: PT Aero Inovasi Media; NPWP; alamat; IDR", "res.company", "AIM-ADM", "Company aktif", "Multi-company", "Settings > Users & Companies > Companies > New"],
    ["M-03", "ARKA", "Set Chart of Account (CoA)", "Import COA-Import-ARKA.xlsx (546 akun: 544 bersama + 2 bank ARKA)", "account.account", "ARK-FIN", "CoA siap", "COA asli user; lihat sheet 10", "Accounting > Configuration > Chart of Accounts > Import records"],
    ["M-04", "AIM", "Set Chart of Account (CoA)", "Import COA-Import-AIM.xlsx (546 akun: 544 bersama + 2 bank AIM)", "account.account", "AIM-FIN", "CoA siap", "Identik ARKA kecuali akun bank", "Accounting > Configuration > Chart of Accounts > Import records"],
    ["M-05", "ARKA", "Set Pajak (Tax)", "PPN Keluaran 11%, PPN Masukan 11%, PPh23 Jasa 2%, PPh23 Sewa 2%", "account.tax", "ARK-FIN", "Pajak siap", "Map ke akun pajak", "Accounting > Configuration > Taxes > New"],
    ["M-06", "AIM", "Set Pajak (Tax)", "PPN Keluaran 11%, PPh23 Sewa 2% (sbg objek potong)", "account.tax", "AIM-FIN", "Pajak siap", "", "Accounting > Configuration > Taxes > New"],
    ["M-07", "ARKA", "Set Jurnal", "Sales (AR), Purchase (AP), Bank, Cash, Miscellaneous", "account.journal", "ARK-FIN", "Jurnal siap", "Sequence per jurnal", "Accounting > Configuration > Journals > New"],
    ["M-08", "AIM", "Set Jurnal", "Sales (AR sewa), Bank, Miscellaneous", "account.journal", "AIM-FIN", "Jurnal siap", "", "Accounting > Configuration > Journals > New"],
    ["M-09", "ARKA", "Set Termin Pembayaran", "DP 30%/Pelunasan 70%; Net 14 hari", "account.payment.term", "ARK-FIN", "Termin siap", "", "Accounting > Configuration > Payment Terms > New"],
    ["M-10", "ARKA", "Buat Produk JASA Drone Show", "Tipe: Service; harga 150jt/paket; akun pendapatan 5199000000; pajak PPN+PPh23", "product.template", "ARK-ADM", "Produk jasa siap", "Lihat sheet 03", "Sales > Products > Products > New (Product Type = Service)"],
    ["M-11", "AIM", "Buat Produk SEWA Drone (GOODS)", "Tipe: Goods (consu+is_storable=True); tracking=Serial; is_rentable=True; is_rental_asset=True; income 5122000000 (Sewa)", "product.template (custom_rental)", "AIM-ADM", "Produk sewa siap", "Lihat sheet 03", "Sales > Products > Products > New (Type=Goods; Track=Serial; tab Rental: Can be Rented)"],
    ["M-11b", "AIM", "Buat BoM phantom komponen drone", "Kit: body + 2 baterai + charger + controller -> di-explode jadi baris komponen di BAST", "mrp.bom + custom_rental_bom_explosion", "AIM-ADM", "BoM phantom siap", "Detail komponen utk handover", "Inventory/Manufacturing > Products > Bills of Materials > New (BoM Type = Kit/Phantom)"],
    ["M-12", "AIM", "Goods Receipt 200 unit -> Convert to Assets", "Terima 200 drone (serial); wizard auto-buat 200 rental.asset (+ fixed asset utk depresiasi)", "stock.picking + custom_asset_from_receipt", "AIM-OPS", "200 rental.asset", "auto_create_rental_asset per serial", "Inventory > Operations > Receipts > Validate > tombol 'Convert to Assets'"],
    ["M-12b", "AIM", "Set tarif sewa berjenjang", "rental_pricing_ids: per hari / minggu / bulan / event (kru sudah termasuk)", "custom.rental.pricing", "AIM-FIN", "Tarif sewa siap", "1 baris (kru dilebur)", "Rental > Configuration > Pricing Tiers > New (atau tab Rental di produk)"],
    ["M-13", "ARKA", "Daftar Kontak Customer", "PT Gemilang EO: NPWP, alamat tagih, PIC, term", "res.partner (customer)", "ARK-SLS", "Kontak customer", "", "Contacts > New (centang Customer)"],
    ["M-14", "ARKA", "Daftar Kontak Vendor Aim", "PT Aero Inovasi Media sbg Vendor (untuk AP)", "res.partner (vendor)", "ARK-PUR", "Kontak vendor", "", "Contacts > New (centang Vendor)"],
    ["M-15", "AIM", "Daftar Kontak Customer Arka", "PT Aero Reksa Kreasi Angkasa sbg Customer (untuk AR Aim)", "res.partner (customer)", "AIM-ADM", "Kontak customer", "", "Contacts > New (centang Customer)"],
    ["M-16", "ARKA", "Template Dokumen", "Template Quotation, BAST, Berita Acara Survey (custom report)", "ir.actions.report / custom_bast", "ARK-ADM", "Template siap", "Header kop perusahaan", "Settings > Technical > Reporting > Reports  (perlu mode Developer)"],
]
write_table(ws, 4, md[0], md[1:], [7, 10, 30, 44, 28, 9, 18, 18, 46], center_cols=[1, 2, 6])

# ============================================================================
# 03 - REGISTRASI PRODUK
# ============================================================================
ws = wb.create_sheet("03-Registrasi Produk")
ws.sheet_view.showGridLines = False
sheet_title(ws, "FASE 1b — REGISTRASI PRODUK (DETAIL FIELD)",
            "Konfigurasi field produk jasa (Arka) & produk sewa (Aim) di Odoo", 5)

r = 4
ws.cell(row=r, column=1, value="KEPUTUSAN: SEWA = BARANG (GOODS) ATAU JASA (SERVICE)?").font = Font(size=12, bold=True, color=C_TITLE)
r += 1
decision_notes = [
    "JAWABAN: Unit drone di-setup sebagai BARANG (Goods: consu + is_storable + tracking Serial), BUKAN Service. Tipe produk mengatur LOGISTIK: stok, serial, pickup/return (stock.picking), cek ketersediaan (anti double-book), kondisi/kerusakan/hilang saat return, deposit, late fee. Drone itu fisik & dikembalikan -> wajib Goods agar mesin custom_rental jalan.",
    "TAPI pendapatan tetap diakui sebagai JASA SEWA — lewat Income Account (4-10200 Pendapatan Sewa Drone) yang di-map ke produk, BUKAN lewat tipe produk. Jadi: tipe produk = logistik; income account = klasifikasi pendapatan. Di faktur, pajak (PPN + objek PPh23 sewa), dan L/R semuanya terbaca sebagai jasa sewa walau produknya Goods.",
    "Kalau dipaksa Service: kehilangan serial/asset booking, pickup/return stok, cek ketersediaan, tracking kerusakan saat return, deposit, late fee, dan BAST detail komponen. Constraint repo (custom_asset_from_receipt) bahkan MENOLAK rental asset tanpa serial/lot. Service hanya cocok utk kru teknis — di skenario ini kru dilebur ke harga sewa.",
]
for n in decision_notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    cell = ws.cell(row=r, column=1, value=n)
    cell.font = BODY
    cell.alignment = TOPL
    cell.fill = fill(C_NOTE)
    cell.border = BORDER
    ws.row_dimensions[r].height = 58
    r += 1
r += 1
ws.cell(row=r, column=1, value="PRODUK A — JASA DRONE SHOW (Buku ARKA)").font = Font(size=12, bold=True, color=C_ARKA)
r += 1
pa = [
    ["Field Odoo", "Nilai", "Keterangan"],
    ["Nama Produk", "Jasa Drone Show 200 Unit - Paket Event", "Tampil di quotation/invoice"],
    ["Internal Reference", "SVC-DRONESHOW-200", "Kode internal"],
    ["Product Type", "Service", "Bukan stockable"],
    ["Sales Price", "150.000.000", "Per paket (bisa per menit/per unit)"],
    ["Unit of Measure", "Paket / Event", "UoM"],
    ["Customer Taxes", "PPN Keluaran 11% ; PPh23 Jasa 2%", "Pajak saat dijual"],
    ["Income Account", "5199000000 Gross Sales-Others (Jasa Drone Show)", "Akun pendapatan (COA asli; lihat sheet 10)"],
    ["Invoicing Policy", "Ordered / Delivered (setelah BAST)", "Tagih setelah BAST"],
    ["Sales Description", "Paket drone show 200 unit, durasi 12 menit, animasi custom", "Deskripsi"],
    ["PIC Registrasi", "ARK-ADM (input) -> ARK-FIN (validasi akun & pajak)", ""],
]
r = write_table(ws, r, pa[0], pa[1:], [26, 50, 46], header_color=C_ARKA)
r += 1

ws.cell(row=r, column=1, value="PRODUK B — SEWA UNIT DRONE + KRU (Buku AIM)").font = Font(size=12, bold=True, color=C_AIM)
r += 1
pb = [
    ["Field Odoo", "Nilai", "Keterangan"],
    ["Nama Produk", "Sewa Unit Drone Show + Kru (paket 200 unit)", "Kru teknis sudah termasuk di harga"],
    ["Internal Reference", "RNT-DRONE-UNIT", "Kode internal"],
    ["Product Type", "Goods (consu, is_storable=True) — BUKAN Service", "Tipe = logistik; lihat callout di atas"],
    ["Tracking", "By Serial Number (1 drone = 1 serial -> rental.asset)", "Wajib serial/lot (constraint custom_asset_from_receipt)"],
    ["is_rentable", "True (Can be Rented)", "Aktifkan mode rental"],
    ["is_rental_asset", "True", "Tandai aset rental utk konversi dari goods receipt"],
    ["auto_create_rental_asset", "True", "Auto-buat rental.asset per serial saat receipt"],
    ["rental_pricing_ids", "Tier: per hari / minggu / bulan / event", "Tabel tarif sewa berjenjang"],
    ["Harga Sewa (paket)", "60.000.000 / event (SUDAH termasuk kru teknis)", "Kru dilebur ke tarif sewa (1 baris)"],
    ["Unit of Measure", "Unit / Paket", "UoM"],
    ["Customer Taxes", "PPN Keluaran 11% (objek PPh23 Sewa 2% dipotong Arka)", "Pajak saat dijual"],
    ["Income Account", "5122000000 Gross Sales-Rental Asset", "KUNCI: ini yang membuat pendapatan diakui sbg JASA SEWA (COA asli; lihat sheet 10)"],
    ["Bill of Materials", "BoM phantom: body + 2 baterai + charger + controller", "Di-explode custom_rental_bom_explosion -> baris komponen di BAST"],
    ["Invoicing Policy", "Setelah BAST pickup / saat return (custom_rental_invoicing)", "Tagih sewa + late fee + damages"],
    ["PIC Registrasi", "AIM-ADM (input) -> AIM-FIN (validasi akun & pajak)", ""],
]
write_table(ws, r, pb[0], pb[1:], [26, 50, 46], header_color=C_AIM)

# ============================================================================
# 04 - PRA-SALES & SURVEY
# ============================================================================
ws = wb.create_sheet("04-PraSales & Survey")
ws.sheet_view.showGridLines = False
sheet_title(ws, "FASE 2 — INQUIRY & SURVEY LOKASI",
            "Sebelum quotation final; menentukan kelayakan area terbang & perizinan", 9)
sv = [
    ["No", "Aktivitas", "Detail", "Company", "PIC", "Objek/Dokumen Odoo", "Output", "Pra-syarat", "Navigasi Odoo (Klik / Menu)"],
    ["S-01", "Customer kirim inquiry", "PT Gemilang minta drone show utk grand opening; kirim brief (lokasi, tanggal, durasi)", "CUST->ARKA", "CUS-PIC / ARK-SLS", "crm.lead (Lead/Opportunity)", "Lead tercatat", "-", "CRM > Sales > My Pipeline > New  |  [DI LUAR ODOO] inquiry datang dari customer"],
    ["S-02", "Kualifikasi kebutuhan", "Sales catat: 200 unit, 12 menit, outdoor, tanggal show", "ARKA", "ARK-SLS", "crm.lead", "Opportunity", "S-01", "CRM > buka Opportunity > isi detail kebutuhan"],
    ["S-03", "Jadwalkan survey lokasi", "Buat agenda survey ke lokasi event", "ARKA", "ARK-SLS -> ARK-SVY", "calendar.event / project.task", "Jadwal survey", "S-02", "CRM > Opportunity > Activities > Schedule Activity  (atau Project > Tasks > New)"],
    ["S-04", "Survey lokasi (on-site)", "Cek luas area terbang, obstacle, sumber listrik, line of sight, koordinat GPS, ketinggian izin", "ARKA(+CUST)", "ARK-SVY", "project.task / custom survey", "Data lapangan", "S-03", "[DI LUAR ODOO] survey fisik di lapangan; hasil dicatat ke Project task / Attachments"],
    ["S-05", "Cek perizinan terbang", "Verifikasi zona (dekat bandara?), kebutuhan izin AirNav/otoritas, NOTAM bila perlu", "ARKA", "ARK-SVY", "checklist/attachment", "Status izin", "S-04", "[DI LUAR ODOO] verifikasi zona & izin AirNav; lampirkan dokumen ke task"],
    ["S-06", "Isi Berita Acara Survey", "Form: tanggal, lokasi, koordinat, foto, kelayakan (LAYAK/TIDAK), rekomendasi titik", "ARKA", "ARK-SVY", "custom report (BA Survey)", "Berita Acara Survey Lokasi (nomor)", "S-04, S-05", "Project > buka Task > isi field hasil survey  (atau custom report BA Survey)"],
    ["S-07", "TTD Berita Acara Survey", "Ditandatangani ARK-SVY & CUS-PIC", "ARKA+CUST", "ARK-SVY / CUS-PIC", "attachment (signed PDF)", "BA Survey final", "S-06", "[DI LUAR ODOO] tanda tangan / e-sign; upload PDF ke Attachments"],
    ["S-08", "Hasil survey -> dasar penawaran", "Kelayakan & titik koordinat jadi basis paket & harga; tentukan jumlah unit (200)", "ARKA", "ARK-SLS", "crm.lead -> quotation", "Input quotation", "S-07", "CRM > buka Opportunity > tombol New Quotation"],
]
write_table(ws, 4, sv[0], sv[1:], [7, 24, 46, 12, 16, 26, 22, 11, 46], center_cols=[1, 4])

r = ws.max_row + 2
ws.cell(row=r, column=1, value="ISI FORM BERITA ACARA SURVEY LOKASI (field minimum)").font = Font(size=11, bold=True, color=C_TITLE)
r += 1
baf = [
    ["Field", "Contoh Isi"],
    ["No. Berita Acara", "BAS/ARKA/2026/06/001"],
    ["Tanggal Survey", "10 Juni 2026"],
    ["Nama Lokasi / Venue", "Lapangan Utama Gemilang City, Jakarta"],
    ["Koordinat GPS (titik take-off)", "-6.21462, 106.84513"],
    ["Luas Area Terbang", "120 m x 80 m (clear zone)"],
    ["Obstacle / Hambatan", "Tiang lampu sisi timur (24 m) — aman, di luar flight box"],
    ["Sumber Listrik / Genset", "Tersedia genset 20 kVA"],
    ["Status Perizinan Terbang", "Di luar KKOP bandara — tidak perlu izin AirNav khusus"],
    ["Kelayakan", "LAYAK (dengan flight box 100x60 m, ketinggian max 120 m)"],
    ["Rekomendasi", "Show jam 19:30; pasang 2 ground station; backup 10 unit"],
    ["Surveyor (TTD)", "ARK-SVY"],
    ["Wakil Customer (TTD)", "CUS-PIC"],
]
write_table(ws, r, baf[0], baf[1:], [34, 64], header_color=C_HEAD2)

# ============================================================================
# 05 - SALES & RENTAL (TRANSAKSI INTI)
# ============================================================================
ws = wb.create_sheet("05-Sales & Rental")
ws.sheet_view.showGridLines = False
sheet_title(ws, "FASE 3-4 — SALES ORDER (ARKA) & RENTAL ORDER (AIM)",
            "Penjualan jasa Arka ke Customer + Pengadaan sewa Arka ke Aim", 9)
sr = [
    ["No", "Aktivitas", "Detail", "Company", "PIC", "Objek Odoo", "Dokumen", "Pra-syarat", "Navigasi Odoo (Klik / Menu)"],
    ["T-01", "Buat Quotation Arka", "Produk: Jasa Drone Show 200; harga 150jt; pajak PPN+PPh23; termin DP30/70", "ARKA", "ARK-SLS", "sale.order (draft)", "QUO/ARKA/2026/0007", "S-08", "Sales > Orders > Quotations > New"],
    ["T-02", "Approval quotation", "Direktur approve (nilai > limit)", "ARKA", "ARK-DIR", "sale.order (approval)", "Quotation approved", "T-01", "Sales > buka Quotation > tombol Approve  (custom_approval_engine)"],
    ["T-03", "Kirim quotation ke Customer", "Email PDF quotation + lampiran BA Survey", "ARKA->CUST", "ARK-SLS", "sale.order (sent)", "Quotation terkirim", "T-02", "Sales > buka Quotation > tombol Send by Email"],
    ["T-04", "Customer setuju & PO", "Customer terbitkan PO; tanda setuju", "CUST->ARKA", "CUS-PIC", "PO Customer (attach)", "PO Customer", "T-03", "[DI LUAR ODOO] customer terbitkan PO; upload ke Attachments di SO"],
    ["T-05", "Konfirmasi SO Arka", "Quotation -> Sales Order terkonfirmasi", "ARKA", "ARK-SLS", "sale.order (confirmed)", "SO/ARKA/2026/0007", "T-04", "Sales > buka Quotation > tombol Confirm"],
    ["T-06", "Tagih DP 30% (opsional)", "Buat invoice DP 30% x 150jt = 45jt + PPN", "ARKA", "ARK-FIN", "account.move (down payment)", "Inv DP ARKA", "T-05", "Sales > buka SO > Create Invoice > Down payment"],
    ["T-07", "Identifikasi kebutuhan sewa", "Show butuh 200 unit drone + kru -> tidak punya sendiri -> sewa ke Aim", "ARKA", "ARK-OPS", "internal note / req", "Kebutuhan sewa", "T-05", "[INTERNAL] catat di SO note / Activity (belum ada transaksi)"],
    ["T-08", "Buat RFQ/PO sewa ke Aim", "PO: Sewa Unit Drone 200 + Kru; harga 60jt; tanggal event", "ARKA->AIM", "ARK-PUR", "purchase.order", "PO/ARKA/2026/0015", "T-07", "Purchase > Orders > Requests for Quotation > New"],
    ["T-09", "Approve PO sewa", "Direktur approve PO sewa", "ARKA", "ARK-DIR", "purchase.order (approval)", "PO approved", "T-08", "Purchase > buka RFQ > tombol Confirm Order"],
    ["T-10", "Aim terima order & buat Rental Order", "Aim catat order sewa dari Arka -> Rental Order", "AIM", "AIM-RNT", "rental.order (custom_rental)", "RO/AIM/2026/0031", "T-08", "Rental > Orders > New   (buku AIM)"],
    ["T-11", "Aim cek ketersediaan unit", "Reservasi 200 unit + 10 cadangan; jadwal kru teknis", "AIM", "AIM-OPS", "rental.asset / rental schedule", "Unit ter-reserve", "T-10", "Rental > Schedule  /  Rental > Assets  (cek availability)"],
    ["T-12", "Konfirmasi Rental Order Aim", "RO dikonfirmasi; jadwal pickup/kirim ditetapkan", "AIM", "AIM-RNT", "rental.order (confirmed)", "RO confirmed", "T-11", "Rental > buka Order > tombol Confirm"],
]
write_table(ws, 4, sr[0], sr[1:], [7, 26, 44, 12, 9, 24, 22, 11, 46], center_cols=[1, 4])

# ============================================================================
# 06 - EKSEKUSI & BAST
# ============================================================================
ws = wb.create_sheet("06-Eksekusi & BAST")
ws.sheet_view.showGridLines = False
sheet_title(ws, "FASE 5-6 — MOBILISASI, EKSEKUSI & BAST",
            "Pengiriman unit (Aim), eksekusi show (Arka), 2 dokumen BAST", 9)
ex = [
    ["No", "Aktivitas", "Detail", "Company", "PIC", "Objek/Dokumen Odoo", "Output", "Pra-syarat", "Navigasi Odoo (Klik / Menu)"],
    ["E-01", "Aim siapkan & kirim unit", "Aim keluarkan 210 unit (200+10) + ground station; Delivery Order sewa", "AIM->ARKA", "AIM-OPS", "stock.picking (DO sewa)", "DO/AIM/2026/0031", "T-12", "Rental > buka Order > Confirm (auto buat Delivery) ; Inventory > Deliveries > Validate"],
    ["E-02", "Serah terima unit (BAST #1)", "Arka terima unit dari Aim; cek per serial; BAST merinci komponen per unit (BOM explosion); tanda tangan", "AIM+ARKA", "AIM-OPS / ARK-OPS", "custom_bast (pickup)", "BAST/2026/00001 (pickup)", "E-01", "Rental > buka Order > tombol 'Generate BAST Pickup'"],
    ["E-03", "Mobilisasi ke lokasi", "Tim Arka + unit dibawa ke venue; setup ground station 2 titik", "ARKA", "ARK-OPS", "project.task", "Setup selesai", "E-02", "[DI LUAR ODOO] angkut unit ke venue (opsional catat di Project task)"],
    ["E-04", "Uji terbang / rehearsal", "Test flight, kalibrasi GPS, cek animasi sesuai koordinat survey", "ARKA(+AIM)", "ARK-OPS / AIM-OPS", "checklist", "Lolos uji", "E-03", "[DI LUAR ODOO] rehearsal/uji terbang fisik (catat checklist di task)"],
    ["E-05", "Eksekusi drone show", "Show 200 unit, durasi 12 menit, malam hari event", "ARKA", "ARK-OPS", "project.task / log", "Show sukses + Laporan Show", "E-04", "[DI LUAR ODOO] drone show; catat Laporan Show di Project task"],
    ["E-06", "BAST hasil jasa (BAST #2)", "Arka serahkan hasil jasa drone show ke Customer; TTD penerimaan", "ARKA+CUST", "ARK-OPS / CUS-PIC", "custom_bast (delivery/SO)", "BAST/2026/00002 (delivery)", "E-05", "Sales > buka SO > tombol 'Generate BAST'"],
    ["E-07", "Pengembalian unit ke Aim", "Bongkar; kembalikan 210 unit; cek kerusakan; update kondisi", "ARKA->AIM", "ARK-OPS / AIM-OPS", "stock.picking (return)", "Unit kembali", "E-05", "Rental > buka Order > 'Mark Returned' (auto buat Return) ; Inventory > Receipts > Validate"],
    ["E-08", "BAST pengembalian (BAST #3)", "Aim konfirmasi unit kembali lengkap/kondisi; cek damage/hilang; selesai sewa", "AIM+ARKA", "AIM-OPS / ARK-OPS", "custom_bast (return)", "BAST/2026/00003 (return)", "E-07", "Rental > buka Order > tombol 'Generate BAST Return'"],
    ["E-09", "BAST jadi trigger penagihan", "BAST #1 -> Aim boleh tagih Arka; BAST #2 -> Arka boleh tagih Customer", "ARKA/AIM", "ARK-FIN / AIM-FIN", "sale.order -> invoice", "Siap buat faktur", "E-06, E-08", "(otomatis) BAST jadi dasar penagihan -> lanjut ke sheet 07"],
]
write_table(ws, 4, ex[0], ex[1:], [7, 26, 44, 12, 16, 24, 22, 11, 46], center_cols=[1, 4])

r = ws.max_row + 2
ws.cell(row=r, column=1, value="ISI DOKUMEN BAST (kedua BAST memakai field serupa)").font = Font(size=11, bold=True, color=C_TITLE)
r += 1
bast = [
    ["Field BAST", "BAST #1 (Aero Inovasi Media -> Aero Reksa, Pickup Sewa)", "BAST #2 (Aero Reksa -> Customer, Hasil Jasa)"],
    ["No. BAST (sequence global)", "BAST/2026/00001  (kind=pickup)", "BAST/2026/00002  (kind=delivery)"],
    ["Tanggal", "18 Juni 2026 (sebelum show)", "18 Juni 2026 (malam, setelah show)"],
    ["Pihak Pertama (penyerah)", "PT Aero Inovasi Media (AIM-OPS)", "PT Aero Reksa Kreasi Angkasa (ARK-OPS)"],
    ["Pihak Kedua (penerima)", "PT Aero Reksa Kreasi Angkasa (ARK-OPS)", "PT Gemilang EO (CUS-PIC)"],
    ["Objek Serah Terima", "210 unit drone (per serial) + 2 ground station", "Jasa Drone Show 200 unit, 12 menit (selesai)"],
    ["Rincian Baris BAST", "Auto dari BOM explosion: tiap serial unit dirinci komponen (body, 2 baterai, charger, controller) + kolom kondisi/qty per komponen", "1 baris jasa drone show (qty 1 paket)"],
    ["Referensi (reference)", "rental.order RO/AIM/2026/0031 ; PO/ARKA/2026/0015", "sale.order SO/ARKA/2026/0007 ; PO Customer"],
    ["Kondisi / Hasil", "Semua unit baik & lengkap", "Show terlaksana sesuai spesifikasi"],
    ["Catatan", "10 unit cadangan (loan_qty); 0 hilang. BAST pengembalian = BAST/2026/00003 (kind=return) saat unit balik", "Animasi sesuai brief; cuaca cerah"],
    ["Sequence (Odoo)", "ir.sequence custom.bast.document — GLOBAL: BAST/YYYY/NNNNN; jenis dibedakan field 'kind' (pickup/return/delivery), bukan nomor", "Sama; kind=delivery utk BAST jasa dari SO"],
]
write_table(ws, r, bast[0], bast[1:], [26, 42, 42], header_color=C_HEAD2)

# ============================================================================
# 07 - AR & AP
# ============================================================================
ws = wb.create_sheet("07-AR & AP")
ws.sheet_view.showGridLines = False
sheet_title(ws, "FASE 7-8 — PENAGIHAN (AR/AP) & PEMBAYARAN + PAJAK",
            "AP Arka (Aim tagih Arka) | AR Arka (Arka tagih Customer) | AR Aim", 9)
ar = [
    ["No", "Aktivitas", "Detail", "Company", "PIC", "Objek Odoo", "Nilai (Rp)", "Dokumen", "Navigasi Odoo (Klik / Menu)"],
    ["F-01", "Aim buat faktur sewa ke Arka", "Dari RO; DPP 60jt + PPN 6,6jt = 66,6jt", "AIM (AR)", "AIM-FIN", "account.move (out_invoice)", 66_600_000, "INV/AIM/2026/0031", "Rental > buka Order > tombol 'Create Invoice'  (buku AIM)"],
    ["F-02", "Arka terima & validasi tagihan", "Catat sbg Vendor Bill; cek vs PO & BAST #1", "ARKA (AP)", "ARK-FIN", "account.move (in_invoice)", 66_600_000, "BILL/ARKA/2026/0015", "Accounting > Vendors > Bills > New  (buku ARKA)"],
    ["F-03", "Arka potong PPh23 sewa 2%", "PPh23 = 2% x 60jt = 1,2jt; net bayar 65,4jt", "ARKA", "ARK-FIN", "tax (PPh23) + bukti potong", 1_200_000, "Bupot PPh23 -> Aim", "di Vendor Bill: tambah pajak PPh23 (custom_pph_witholding)  |  [DI LUAR ODOO] bupot via Coretax/DJP"],
    ["F-04", "Arka bayar Aim", "Transfer 65,4jt (66,6jt - 1,2jt PPh23)", "ARKA->AIM", "ARK-FIN", "account.payment", 65_400_000, "Bukti transfer", "Accounting > buka Bill > tombol Register Payment (net 65,4jt)"],
    ["F-05", "Aim terima pembayaran & bupot", "Rekon piutang; catat PPh23 dipotong sbg kredit pajak", "AIM", "AIM-FIN", "account.payment + reconcile", 65_400_000, "Pelunasan AR Aim", "Accounting > buka Invoice > Register Payment / Reconcile  (buku AIM)"],
    ["F-06", "Arka buat faktur jasa ke Customer", "Dari SO (dikurangi DP bila ada); DPP 150jt + PPN 16,5jt", "ARKA (AR)", "ARK-FIN", "account.move (out_invoice)", 166_500_000, "INV/ARKA/2026/0007", "Accounting > Customers > Invoices > New  (atau dari SO: Create Invoice)"],
    ["F-07", "Terbitkan e-Faktur PPN", "Upload e-Faktur (PPN Keluaran 16,5jt)", "ARKA", "ARK-FIN", "l10n_id e-faktur / export", 16_500_000, "e-Faktur", "Accounting > Customers > e-Faktur / Coretax Export (custom_coretax)"],
    ["F-08", "Customer potong PPh23 jasa 2%", "PPh23 = 2% x 150jt = 3jt; net diterima 163,5jt", "CUST->ARKA", "CUS-PIC", "bukti potong (diterima Arka)", 3_000_000, "Bupot PPh23 -> Arka", "[DI LUAR ODOO] customer yang memotong; Arka input bupot sbg kredit pajak"],
    ["F-09", "Customer bayar Arka", "Transfer 163,5jt (166,5jt - 3jt PPh23)", "CUST->ARKA", "CUS-PIC / ARK-FIN", "account.payment", 163_500_000, "Pelunasan AR Arka", "Accounting > buka Invoice > Register Payment (net 163,5jt)"],
    ["F-10", "Arka setor PPN & PPh23", "Setor PPN (keluaran-masukan) & PPh23 sewa terpotong ke negara", "ARKA", "ARK-FIN", "tax payment / SPT", 0, "SSP/Bukti setor", "Accounting > Reporting > Tax Report > Closing  |  [DI LUAR ODOO] setor via bank/DJP"],
    ["F-11", "Rekonsiliasi & tutup", "Rekon bank, piutang, utang; arsip BAST/faktur/bupot", "ARKA/AIM", "ARK-FIN / AIM-FIN", "reconciliation", 0, "Closing", "Accounting > Dashboard > Reconcile / Bank"],
]
write_table(ws, 4, ar[0], ar[1:], [7, 26, 40, 11, 12, 24, 15, 20, 46],
            center_cols=[1, 4], money_cols=[7])

r = ws.max_row + 2
ws.cell(row=r, column=1, value="RINCIAN ANGKA TAGIHAN (RINGKAS)").font = Font(size=11, bold=True, color=C_TITLE)
r += 1
calc = [
    ["Komponen", "Aim -> Arka (Sewa)", "Arka -> Customer (Jasa)"],
    ["DPP (Dasar Pengenaan Pajak)", 60_000_000, 150_000_000],
    ["PPN 11% (Keluaran penjual)", 6_600_000, 16_500_000],
    ["Nilai Faktur (DPP+PPN)", 66_600_000, 166_500_000],
    ["PPh 23 (2%) dipotong pembeli", 1_200_000, 3_000_000],
    ["Dibayar tunai (Faktur - PPh23)", 65_400_000, 163_500_000],
]
rr = write_table(ws, r, calc[0], calc[1:], [34, 26, 28], header_color=C_HEAD2, money_cols=[2, 3])

# ============================================================================
# 08 - JURNAL AKUNTANSI
# ============================================================================
ws = wb.create_sheet("08-Jurnal Akuntansi")
ws.sheet_view.showGridLines = False
sheet_title(ws, "FASE 9 — JURNAL AKUNTANSI (2 SISI)",
            "Posting otomatis Odoo dari faktur & pembayaran | Buku ARKA & AIM | kode akun = COA asli (lihat sheet 10)", 6)

def journal_block(ws, start_row, title, color, rows):
    ws.cell(row=start_row, column=1, value=title).font = Font(size=11, bold=True, color=color)
    headers = ["Akun", "Nama Akun", "Debit", "Kredit"]
    rr = write_table(ws, start_row + 1, headers, rows, [13, 52, 17, 17],
                     header_color=color, money_cols=[3, 4], band=False)
    # total
    tot_d = sum(x[2] for x in rows if isinstance(x[2], (int, float)))
    tot_k = sum(x[3] for x in rows if isinstance(x[3], (int, float)))
    for j, val in enumerate(["", "TOTAL", tot_d, tot_k], start=1):
        cell = ws.cell(row=rr, column=j, value=val)
        cell.font = BODYB
        cell.fill = fill(C_TOTAL)
        cell.border = BORDER
        if j in (3, 4):
            cell.alignment = RIGHT
            cell.number_format = '#,##0'
        else:
            cell.alignment = TOP
    return rr + 2

r = 4
r = journal_block(ws, r, "JURNAL 1 (ARKA) — Vendor Bill Sewa dari Aim [F-02/F-03]", C_ARKA, [
    ["7214004000", "Equipment Rental (Beban Sewa Drone)", 60_000_000, 0],
    ["1117200001", "VAT In (PPN Masukan)", 6_600_000, 0],
    ["2104100005", "Income tax payable art.23 (Utang PPh23)", 0, 1_200_000],
    ["2103100001", "Trade Payables-Third parties (AP-Aim)", 0, 65_400_000],
])
r = journal_block(ws, r, "JURNAL 2 (ARKA) — Bayar Aim [F-04]", C_ARKA, [
    ["2103100001", "Trade Payables-Third parties (AP-Aim)", 65_400_000, 0],
    ["1103019270", "BCA 268.695.5777 (Bank ARKA)", 0, 65_400_000],
])
r = journal_block(ws, r, "JURNAL 3 (ARKA) — Faktur Jasa ke Customer [F-06]", C_ARKA, [
    ["1106000001", "Trade Receivables-Third Parties (AR)", 163_500_000, 0],
    ["1117100002", "Prepaid Income tax art.23 (Uang Muka PPh23)", 3_000_000, 0],
    ["5199000000", "Gross Sales-Others (Pendapatan Jasa Drone Show)", 0, 150_000_000],
    ["2104300001", "VAT Out (PPN Keluaran)", 0, 16_500_000],
])
r = journal_block(ws, r, "JURNAL 4 (ARKA) — Terima Bayar Customer [F-09]", C_ARKA, [
    ["1103019270", "BCA 268.695.5777 (Bank ARKA)", 163_500_000, 0],
    ["1106000001", "Trade Receivables-Third Parties (AR)", 0, 163_500_000],
])
r = journal_block(ws, r, "JURNAL 5 (ARKA) — Setor PPh 23 ke Negara [F-10]", C_ARKA, [
    ["2104100005", "Income tax payable art.23 (Utang PPh23)", 1_200_000, 0],
    ["1103019270", "BCA 268.695.5777 (Bank ARKA)", 0, 1_200_000],
])

r = journal_block(ws, r, "JURNAL 6 (AIM) — Faktur Sewa ke Arka [F-01]", C_AIM, [
    ["1106000001", "Trade Receivables-Third Parties (AR-Arka)", 65_400_000, 0],
    ["1117100002", "Prepaid Income tax art.23 (Uang Muka PPh23)", 1_200_000, 0],
    ["5122000000", "Gross Sales-Rental Asset (Pendapatan Sewa)", 0, 60_000_000],
    ["2104300001", "VAT Out (PPN Keluaran)", 0, 6_600_000],
])
r = journal_block(ws, r, "JURNAL 7 (AIM) — Terima Bayar dari Arka [F-05]", C_AIM, [
    ["1103019300", "BCA 268.222.9595 (Bank AIM)", 65_400_000, 0],
    ["1106000001", "Trade Receivables-Third Parties (AR-Arka)", 0, 65_400_000],
])

# ringkasan L/R Arka
ws.cell(row=r, column=1, value="RINGKASAN LABA KOTOR ARKA (dari transaksi ini)").font = Font(size=11, bold=True, color=C_TITLE)
r += 1
lr = [
    ["Komponen", "Nilai (Rp)"],
    ["Pendapatan Jasa Drone Show", 150_000_000],
    ["(-) Beban Sewa Drone (ke Aim)", -60_000_000],
    ["= Laba Kotor", 90_000_000],
    ["Catatan", "Belum termasuk beban mobilisasi, kru, listrik, dll."],
]
write_table(ws, r, lr[0], lr[1:], [40, 22], header_color=C_HEAD2, money_cols=[2])

# ============================================================================
# 09 - ALUR STEP-BY-STEP LENGKAP
# ============================================================================
ws = wb.create_sheet("09-Alur Lengkap")
ws.sheet_view.showGridLines = False
sheet_title(ws, "ALUR STEP-BY-STEP LENGKAP (MASTER) — MASTERING -> AKUNTANSI",
            "Satu tabel kronologis lengkap dengan PIC, company, objek Odoo, dokumen & navigasi klik", 8)
allsteps = [
    ["#", "Fase", "Langkah", "PIC", "Company", "Objek/Modul Odoo", "Dokumen/Output", "Navigasi Odoo (Klik / Menu)"],
    # mastering
    [1, "Mastering", "Setup company Arka & Aim", "ARK-ADM/AIM-ADM", "ARKA/AIM", "res.company", "2 company", "Settings > Users & Companies > Companies > New"],
    [2, "Mastering", "Setup CoA, pajak (PPN/PPh23), jurnal, termin", "ARK-FIN/AIM-FIN", "ARKA/AIM", "account.account/tax/journal", "Akun & pajak siap", "Accounting > Configuration > Chart of Accounts / Taxes / Journals / Payment Terms"],
    [3, "Mastering", "Registrasi produk: jasa (Arka) & sewa drone GOODS+serial (Aim) + BoM phantom + receipt->rental.asset", "ARK-ADM/AIM-ADM", "ARKA/AIM", "product.template/mrp.bom/custom_asset_from_receipt", "2 produk + 200 rental.asset", "Sales > Products > New ; Inventory > Bills of Materials ; Inventory > Receipts > Convert to Assets"],
    [4, "Mastering", "Registrasi kontak: Customer, Vendor Aim, Customer Arka", "ARK-SLS/ARK-PUR/AIM-ADM", "ARKA/AIM", "res.partner", "Kontak siap", "Contacts > New"],
    # survey
    [5, "Survey", "Customer inquiry drone show", "CUS-PIC/ARK-SLS", "CUST->ARKA", "crm.lead", "Lead/Opportunity", "CRM > Sales > My Pipeline > New"],
    [6, "Survey", "Survey lokasi on-site (GPS, area, izin)", "ARK-SVY", "ARKA", "project.task", "Data lapangan", "[DI LUAR ODOO] survey lapangan (catat di Project task)"],
    [7, "Survey", "Terbitkan & TTD Berita Acara Survey Lokasi", "ARK-SVY/CUS-PIC", "ARKA+CUST", "custom report", "BAS/ARKA/2026/06/001", "Project > Task / custom report  |  [DI LUAR ODOO] TTD"],
    # sales
    [8, "Sales", "Buat & approve quotation Arka (150jt)", "ARK-SLS/ARK-DIR", "ARKA", "sale.order", "QUO/ARKA/2026/0007", "Sales > Orders > Quotations > New > Approve"],
    [9, "Sales", "Customer setuju + terbit PO; konfirmasi SO", "CUS-PIC/ARK-SLS", "CUST->ARKA", "sale.order", "SO/ARKA/2026/0007", "Sales > buka Quotation > Confirm  |  [DI LUAR ODOO] PO customer"],
    [10, "Sales", "(Opsional) Tagih DP 30% = 45jt+PPN", "ARK-FIN", "ARKA", "account.move", "Inv DP", "Sales > SO > Create Invoice > Down payment"],
    # rental
    [11, "Rental", "Arka buat & approve PO sewa ke Aim (60jt)", "ARK-PUR/ARK-DIR", "ARKA->AIM", "purchase.order", "PO/ARKA/2026/0015", "Purchase > Requests for Quotation > New > Confirm Order"],
    [12, "Rental", "Aim buat Rental Order & reservasi 210 unit (rental.asset per serial)", "AIM-RNT/AIM-OPS", "AIM", "rental.order (custom_rental)", "RO/AIM/2026/0031", "Rental > Orders > New > Confirm  (buku AIM)"],
    # exec
    [13, "Eksekusi", "Aim kirim unit (DO sewa)", "AIM-OPS", "AIM->ARKA", "stock.picking", "DO/AIM/2026/0031", "Rental > Order > Confirm ; Inventory > Deliveries > Validate"],
    [14, "Eksekusi", "BAST #1 (pickup): serah unit Aim->Arka, rinci komponen via BOM explosion", "AIM-OPS/ARK-OPS", "AIM+ARKA", "custom_bast (pickup)", "BAST/2026/00001", "Rental > Order > tombol 'Generate BAST Pickup'"],
    [15, "Eksekusi", "Mobilisasi, uji terbang, eksekusi drone show", "ARK-OPS", "ARKA", "project.task", "Laporan Show", "[DI LUAR ODOO] drone show (Laporan Show di task)"],
    [16, "Eksekusi", "BAST #2 (delivery): serah hasil jasa Arka->Customer", "ARK-OPS/CUS-PIC", "ARKA+CUST", "custom_bast (delivery)", "BAST/2026/00002", "Sales > SO > tombol 'Generate BAST'"],
    [17, "Eksekusi", "Kembalikan unit; BAST #3 (return) cek damage/hilang", "ARK-OPS/AIM-OPS", "ARKA->AIM", "stock.picking/custom_bast (return)", "BAST/2026/00003", "Rental > Order > 'Mark Returned' + 'Generate BAST Return'"],
    # AP
    [18, "AP", "Aim faktur sewa ke Arka (66,6jt)", "AIM-FIN", "AIM(AR)", "account.move (out_invoice)", "INV/AIM/2026/0031", "Rental > Order > tombol 'Create Invoice'  (buku AIM)"],
    [19, "AP", "Arka catat Vendor Bill; potong PPh23 1,2jt", "ARK-FIN", "ARKA(AP)", "account.move (in_invoice)", "BILL + Bupot", "Accounting > Vendors > Bills > New  (buku ARKA)"],
    [20, "AP", "Arka bayar Aim 65,4jt", "ARK-FIN", "ARKA->AIM", "account.payment", "Bukti transfer", "Accounting > buka Bill > Register Payment"],
    # AR
    [21, "AR", "Arka faktur jasa ke Customer (166,5jt) + e-Faktur", "ARK-FIN", "ARKA(AR)", "account.move (out_invoice)", "INV/ARKA/2026/0007", "Accounting > Customers > Invoices > New  (buku ARKA)"],
    [22, "AR", "Customer potong PPh23 3jt; bayar 163,5jt", "CUS-PIC/ARK-FIN", "CUST->ARKA", "account.payment", "Pelunasan + Bupot", "Accounting > Invoice > Register Payment  |  [DI LUAR ODOO] PPh23 dipotong customer"],
    # accounting
    [23, "Akuntansi", "Arka setor PPN & PPh23 ke negara", "ARK-FIN", "ARKA", "tax payment", "SSP/SPT", "Accounting > Reporting > Tax Report  |  [DI LUAR ODOO] setor DJP"],
    [24, "Akuntansi", "Posting jurnal 2 sisi; rekonsiliasi bank/AR/AP", "ARK-FIN/AIM-FIN", "ARKA/AIM", "account.move", "Jurnal final", "(otomatis saat Validate) ; Accounting > Reconcile"],
    [25, "Akuntansi", "Laba kotor Arka = 90jt; arsip dokumen", "ARK-FIN", "ARKA", "report (P&L)", "Laporan L/R", "Accounting > Reporting > Profit and Loss"],
]
write_table(ws, 4, allsteps[0], allsteps[1:], [5, 12, 40, 18, 12, 26, 22, 50],
            center_cols=[1, 2, 5])

# ============================================================================
# 10 - COA MAPPING (penyelarasan ke COA asli user)
# ============================================================================
ws = wb.create_sheet("10-COA Mapping")
ws.sheet_view.showGridLines = False
sheet_title(ws, "PENYELARASAN COA — KODE AKUN ASLI (DARI TEMPLATE USER)",
            "COA ARKA & AIM IDENTIK (544 akun) kecuali 4 akun bank | Import: COA-Import-ARKA.xlsx & COA-Import-AIM.xlsx", 5)

r = 4
ws.cell(row=r, column=1, value="A. AKUN BANK (satu-satunya perbedaan ARKA vs AIM)").font = Font(size=12, bold=True, color=C_TITLE)
r += 1
banks = [
    ["Perusahaan", "Kode", "Nama Akun", "account_type"],
    ["ARKA", "1103019270", "BCA - IDR-268.695.5777 - Main Bank", "asset_cash"],
    ["ARKA", "1103019280", "BCA - IDR-268.150.7878 - Main Bank", "asset_cash"],
    ["AIM", "1103019300", "BCA - IDR-268.222.9595 - Main Bank", "asset_cash"],
    ["AIM", "1103019290", "BCA - IDR-268.262.6268 - Main Bank", "asset_cash"],
]
r = write_table(ws, r, banks[0], banks[1:], [12, 14, 46, 16], header_color=C_HEAD2, center_cols=[1, 2])
r += 1

ws.cell(row=r, column=1, value="B. PEMETAAN AKUN SKENARIO -> KODE COA ASLI").font = Font(size=12, bold=True, color=C_TITLE)
r += 1
mapping = [
    ["Akun di Skenario (lama)", "Kode COA Asli", "Nama Akun (COA)", "account_type", "Dipakai"],
    ["Bank ARKA", "1103019270", "BCA 268.695.5777 - Main Bank", "asset_cash", "ARKA"],
    ["Bank AIM", "1103019300", "BCA 268.222.9595 - Main Bank", "asset_cash", "AIM"],
    ["Kas (IDR)", "1102000001", "Cash on hand - IDR", "asset_cash", "Both"],
    ["Piutang Usaha (AR)", "1106000001", "Trade Receivables - Third Parties", "asset_receivable", "Both"],
    ["Utang Usaha (AP)", "2103100001", "Trade Payables - Third parties", "liability_current*", "Both"],
    ["PPN Masukan", "1117200001", "VAT In", "asset_current", "Both"],
    ["PPN Keluaran", "2104300001", "VAT Out", "liability_current", "Both"],
    ["PPN Payable (net)", "2104300009", "VAT Payable", "liability_current", "Both"],
    ["Utang PPh 23", "2104100005", "Income tax payable art. 23", "liability_current", "Both"],
    ["Uang Muka PPh 23 (kredit)", "1117100002", "Prepaid Income tax art. 23", "asset_current", "Both"],
    ["Pendapatan Sewa (AIM)", "5122000000", "Gross Sales-Rental Asset", "income", "AIM"],
    ["Pendapatan Jasa (ARKA)", "5199000000", "Gross Sales-Others", "income", "ARKA"],
    ["Beban Sewa Drone (ARKA)", "7214004000", "Equipment Rental", "expense", "ARKA"],
    ["COGS Rental (AIM)", "6122000000", "COGS-Rental Asset", "expense_direct_cost*", "AIM"],
]
r = write_table(ws, r, mapping[0], mapping[1:], [28, 16, 40, 22, 10], header_color=C_HEAD, center_cols=[2, 5])
r += 1

notes10 = [
    "* Trade Payables (2103100001) di template bertipe Current Liabilities. Agar bisa dipakai sebagai Account Payable di kontak vendor, Odoo butuh tipe 'liability_payable' — pertimbangkan ubah. Lihat sheet REVIEW di file import.",
    "* COGS-Rental Asset (6122000000) di template bertipe 'Income'; fungsinya COGS -> saran ubah ke 'expense_direct_cost'. Lihat sheet REVIEW.",
    "Belum ada akun 'pendapatan jasa' khusus; jasa drone show ARKA dipetakan ke 5199000000 Gross Sales-Others. Opsi: buat sub-akun 51990000xx 'Drone Show Service'.",
    "File import siap upload: COA-Import-ARKA.xlsx & COA-Import-AIM.xlsx (masing-masing 546 akun) — sheet 'account.account', header = nama field Odoo (id/code/name/account_type/reconcile).",
]
for n in notes10:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    cell = ws.cell(row=r, column=1, value=n)
    cell.font = BODY
    cell.alignment = TOPL
    cell.fill = fill(C_NOTE)
    cell.border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1

# freeze header on the big sheets
for name in ["02-Master Data", "04-PraSales & Survey", "05-Sales & Rental",
             "06-Eksekusi & BAST", "07-AR & AP", "09-Alur Lengkap", "01-Peran & PIC"]:
    wb[name].freeze_panes = "A5"

import os
outdir = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "docs", "projects", "arka-aim",
)
os.makedirs(outdir, exist_ok=True)
outpath = os.path.join(outdir, "Skenario-Arka-Aim-Drone-Show-Rental.xlsx")
try:
    wb.save(outpath)
except PermissionError:
    alt = os.path.join(outdir, "Skenario-Arka-Aim-Drone-Show-Rental-NEW.xlsx")
    wb.save(alt)
    print("LOCKED (file terbuka di Excel) -> disimpan ke:", alt)
    outpath = alt
else:
    print("SAVED:", outpath)
print("Sheets:", wb.sheetnames)
