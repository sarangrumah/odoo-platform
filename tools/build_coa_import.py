# -*- coding: utf-8 -*-
"""
Generate Odoo-importable COA files per company from the user's master template.

Source : docs/skenario-arka-aim/ARKA-AIM-Master-Data-Template.xlsx  (sheet 06_COA)
Output : docs/skenario-arka-aim/COA-Import-ARKA.xlsx
         docs/skenario-arka-aim/COA-Import-AIM.xlsx

COA identik untuk kedua PT KECUALI 4 akun bank (2 ARKA, 2 AIM) -> kolom L template
menandai kepemilikan per-company. account_type template campur (kode teknis + label
human) -> dipetakan ke selection value Odoo. File siap di-import via
Accounting > Configuration > Chart of Accounts > (gear) Import records.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "docs/skenario-arka-aim/ARKA-AIM-Master-Data-Template.xlsx"
OUTDIR = "docs/skenario-arka-aim"
ARKA = "PT Aero Reksa Kreasi Angkasa"
AIM = "PT Aero Inovasi Media"

# account_type label (template) -> Odoo selection value
TYPE_MAP = {
    "asset_cash": "asset_cash",
    "asset_receivable": "asset_receivable",
    "asset_current": "asset_current",
    "asset_prepayments": "asset_prepayments",
    "asset_non_current": "asset_non_current",
    "asset_fixed": "asset_fixed",
    "Current Assets": "asset_current",
    "Non-current Assets": "asset_non_current",
    "Fixed Assets": "asset_fixed",
    "Current Liabilities": "liability_current",
    "Non-current Liabilities": "liability_non_current",
    "Equity": "equity",
    "Current Year Earnings": "equity_unaffected",
    "Income": "income",
    "Other Income": "income_other",
    "Cost of Revenue": "expense_direct_cost",
    "Expenses": "expense",
    "Depreciation": "expense_depreciation",
}
# 6 baris account_type kosong (akun konsolidasi) -> resolve berdasarkan prefix kode
NONE_BY_PREFIX = {"75": "income_other", "83": "equity", "88": "equity", "89": "equity"}


def map_type(raw, code):
    if raw in (None, ""):
        return NONE_BY_PREFIX.get(str(code)[:2], "equity")
    return TYPE_MAP.get(str(raw).strip(), None)


def applies_to(rowL):
    s = str(rowL or "")
    return (ARKA in s, AIM in s)


def load_accounts():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["06_COA"]
    accts = []
    for r in ws.iter_rows(min_row=7, values_only=True):
        if r[0] in (None, ""):
            continue
        code = str(r[1]).strip()
        name = str(r[2]).strip()
        raw_type = r[3]
        atype = map_type(raw_type, code)
        reconcile = bool(r[6]) if r[6] not in (None, "") else False
        a_ok, m_ok = applies_to(r[11] if len(r) > 11 else None)
        accts.append(
            {
                "code": code,
                "name": name,
                "type": atype,
                "raw_type": raw_type,
                "reconcile": reconcile,
                "arka": a_ok,
                "aim": m_ok,
            }
        )
    return accts


# styling
HEAD = PatternFill("solid", fgColor="1F3864")
HEADF = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND = PatternFill("solid", fgColor="DDEBF7")
BANKFILL = PatternFill("solid", fgColor="FCE4D6")
WRAP = Alignment(vertical="top", wrap_text=True)


def write_company_file(accts, company, out_name):
    rows = [a for a in accts if (a["arka"] if company == ARKA else a["aim"])]
    rows.sort(key=lambda a: a["code"])
    wb = Workbook()

    # ---- Sheet 1: import data (Odoo field names as headers) ----
    ws = wb.active
    ws.title = "account.account"
    headers = ["id", "code", "name", "account_type", "reconcile"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEAD
        c.font = HEADF
        c.border = BORDER
    widths = [22, 14, 56, 22, 11]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, a in enumerate(rows):
        rr = i + 2
        is_bank = a["type"] == "asset_cash" and "Bank" in a["name"]
        vals = [f"coa_{a['code']}", a["code"], a["name"], a["type"], "TRUE" if a["reconcile"] else "FALSE"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=j, value=v)
            c.border = BORDER
            c.font = Font(size=10)
            c.alignment = WRAP
            if is_bank:
                c.fill = BANKFILL
            elif i % 2 == 0:
                c.fill = BAND
    ws.freeze_panes = "A2"

    # ---- Sheet 2: READ_ME ----
    rm = wb.create_sheet("READ_ME")
    rm.column_dimensions["A"].width = 120
    lines = [
        (f"FILE IMPORT COA — {company}", True),
        ("", False),
        (f"Jumlah akun: {len(rows)}  (544 akun bersama + 2 akun bank milik {company.split()[-1]}).", False),
        ("COA ARKA & AIM IDENTIK kecuali 4 akun bank; file ini hanya memuat bank milik perusahaan ini.", False),
        ("", False),
        ("CARA IMPORT di Odoo:", True),
        ("1. Login ke database pada company yang sesuai (ARKA atau AIM).", False),
        ("2. Buka Accounting > Configuration > Chart of Accounts.", False),
        ("3. Klik ikon gear/Favorites > Import records (atau Action > Import).", False),
        ("4. Upload file ini; sheet 'account.account' terbaca otomatis (header = nama field).", False),
        (
            "5. Mapping: id->External ID, code->Code, name->Name, account_type->Type, reconcile->Allow Reconciliation.",
            False,
        ),
        ("6. Test, lalu Import.", False),
        ("", False),
        ("CATATAN account_type:", True),
        (
            "- Label human di template (Expenses, Current Liabilities, dst) sudah dipetakan ke selection value Odoo.",
            False,
        ),
        ("- Akun bank perusahaan ini disorot oranye di sheet data.", False),
        ("", False),
        ("PERLU REVIEW (tipe template mungkin perlu disesuaikan utk fungsi Odoo) — lihat sheet REVIEW:", True),
    ]
    for i, (t, bold) in enumerate(lines, 1):
        c = rm.cell(row=i, column=1, value=t)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 10)
        c.alignment = WRAP

    # ---- Sheet 3: REVIEW (flagged accounts) ----
    rv = wb.create_sheet("REVIEW")
    rvh = ["Kode", "Nama", "Tipe (template)", "Tipe Odoo (dipakai)", "Catatan / Saran"]
    for j, h in enumerate(rvh, 1):
        c = rv.cell(row=1, column=j, value=h)
        c.fill = HEAD
        c.font = HEADF
        c.border = BORDER
    for j, w in enumerate([14, 50, 20, 22, 60], 1):
        rv.column_dimensions[get_column_letter(j)].width = w
    flags = []
    for a in rows:
        note = None
        if a["name"].startswith("Trade Payables") or a["code"].startswith("21031"):
            note = "AP partner di Odoo butuh tipe 'liability_payable'. Pertimbangkan ubah dari liability_current."
        elif str(a["raw_type"]) == "Income" and a["code"].startswith("6"):
            note = "Ini akun COGS tapi template bertipe 'Income'. Saran: ubah ke 'expense_direct_cost'."
        elif a["raw_type"] in (None, ""):
            note = "account_type kosong di template (akun konsolidasi). Di-assign otomatis; mohon verifikasi."
        if note:
            flags.append((a["code"], a["name"], str(a["raw_type"]), a["type"], note))
    for i, f in enumerate(flags, 2):
        for j, v in enumerate(f, 1):
            c = rv.cell(row=i, column=j, value=v)
            c.border = BORDER
            c.font = Font(size=10)
            c.alignment = WRAP

    path = os.path.join(OUTDIR, out_name)
    wb.save(path)
    return path, len(rows), len(flags)


def main():
    accts = load_accounts()
    unmapped = [a for a in accts if a["type"] is None]
    if unmapped:
        print("WARNING unmapped types:", [(a["code"], a["raw_type"]) for a in unmapped])
    p1, n1, f1 = write_company_file(accts, ARKA, "COA-Import-ARKA.xlsx")
    p2, n2, f2 = write_company_file(accts, AIM, "COA-Import-AIM.xlsx")
    print(f"SAVED: {p1}  ({n1} akun, {f1} flagged)")
    print(f"SAVED: {p2}  ({n2} akun, {f2} flagged)")
    # sanity
    arka_banks = [a["code"] for a in accts if a["arka"] and not a["aim"]]
    aim_banks = [a["code"] for a in accts if a["aim"] and not a["arka"]]
    shared = [a for a in accts if a["arka"] and a["aim"]]
    print("shared:", len(shared), "| ARKA-only:", arka_banks, "| AIM-only:", aim_banks)


if __name__ == "__main__":
    main()
