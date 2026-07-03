# -*- coding: utf-8 -*-
"""Generate Finance Portal Odoo — Project Estimation XLSX.

Sheets:
  1. Ringkasan        – key metrics: total 462 / non-PMO 408 / build 202
  2. Semua Fase (408) – 7-phase breakdown, all roles → total 462 / non-PMO 408
  3. Build Detail (202) – feature-level Build phase → BA+Dev+QA = 202

Usage : python docs/finance-portal/build_estimation_xlsx.py
Output: docs/finance-portal/Project-Estimation-Finance-Portal-Odoo.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "Project-Estimation-Finance-Portal-Odoo.xlsx")

# ── Palette & helpers ─────────────────────────────────────────────────────────
NAVY, PURPLE, LILA, LILA2, WHITE, ALT = (
    "1A2332", "714B67", "E8E0E7", "C9B8C6", "FFFFFF", "F5F2F4"
)
_T = Side(style="thin",   color="CCCCCC")
_M = Side(style="medium", color="1A2332")
BORD = Border(left=_T, right=_T, top=_T, bottom=_T)

def fill(h):       return PatternFill("solid", fgColor=h)
def fnt(bold=False, color=NAVY, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def w(ws, r, c, v, bg=WHITE, bold=False, color=NAVY, sz=10,
      h="left", wrap=True, italic=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill  = fill(bg)
    cell.font  = fnt(bold=bold, color=color, size=sz, italic=italic)
    cell.alignment = aln(h=h, wrap=wrap)
    cell.border = BORD
    return cell

def hdr(ws, r, c, v, bg=NAVY):
    return w(ws, r, c, v, bg=bg, bold=True, color=WHITE, h="center")

def sub(ws, r, c, v):
    h = "center" if isinstance(v, (int, float)) else "left"
    return w(ws, r, c, v, bg=LILA, bold=True, h=h)

def tot(ws, r, c, v):
    h = "center" if isinstance(v, (int, float)) else "left"
    return w(ws, r, c, v, bg=LILA2, bold=True, h=h)

def merge_title(ws, r, c1, c2, v, bg=NAVY, sz=14, bold=True, color=WHITE, h="center",
               italic=False):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=v)
    cell.fill = fill(bg); cell.font = fnt(bold=bold, color=color, size=sz, italic=italic)
    cell.alignment = aln(h=h); cell.border = BORD
    return cell

def merge_w(ws, r, c1, c2, v, bg=WHITE, bold=False, color=NAVY, sz=10, h="left",
            italic=False):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=v)
    cell.fill = fill(bg); cell.font = fnt(bold=bold, color=color, size=sz, italic=italic)
    cell.alignment = aln(h=h); cell.border = BORD
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Ringkasan
# ─────────────────────────────────────────────────────────────────────────────
def build_ringkasan(wb):
    ws = wb.active
    ws.title = "Ringkasan"
    ws.sheet_view.showGridLines = False

    for ltr, wd in zip("ABCDEFGHI", [3, 32, 14, 14, 14, 14, 14, 14, 20]):
        ws.column_dimensions[ltr].width = wd
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20

    # ── title ──
    merge_title(ws, 1, 2, 9, "Finance Portal Odoo — Estimasi Mandays", sz=16)
    merge_title(ws, 2, 2, 9,
        "Scope: Odoo-only  ·  SAP & HC/HRIS = Out of Scope  ·  2026-07-01",
        bg=PURPLE, sz=10, bold=False, italic=True)
    ws.row_dimensions[2].height = 18

    # ── 3 metric cards: rows 4-6 ──
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 52
    ws.row_dimensions[6].height = 20
    cards = [
        (2, 3, "Total Proyek\n(incl. PMO + Kontingensi 15%)", 462,
         "Semua 7 fase · semua role"),
        (4, 5, "Total Non-PMO\n(dengan Kontingensi 15%)", 408,
         "BA + Dev + QA · semua 7 fase"),
        (6, 7, "Build Phase Non-PMO\n(tanpa Kontingensi)", 202,
         "BA + Dev + QA · Fase Build saja"),
    ]
    for c1, c2, label, val, note in cards:
        merge_title(ws, 4, c1, c2, label, bg=PURPLE, sz=10, bold=True)
        merge_w(ws, 5, c1, c2, val,   bg=LILA,  bold=True,  sz=32, h="center")
        ws.cell(row=5, column=c1).font = fnt(bold=True, color=NAVY, size=32)
        merge_w(ws, 6, c1, c2, note,  bg=ALT,   italic=True, sz=9, h="center")
        ws.cell(row=6, column=c1).font = fnt(italic=True, color=NAVY, size=9)

    # ── section header row 8 ──
    merge_title(ws, 8, 2, 9, "Rekap per Fase — Semua Role", bg=PURPLE, sz=11)
    ws.row_dimensions[8].height = 22

    # ── table header row 9 ──
    for col, lbl in enumerate(["No", "Fase", "PMO", "BA", "Dev", "QA", "Total", "Keterangan"], 2):
        hdr(ws, 9, col, lbl)
    ws.row_dimensions[9].height = 18

    phases = [
        ("1", "Requirement & Analysis",         8,  22,   3,  4,  37,
         "FSD draft, kontrak integrasi final"),
        ("2", "Design (FSD/TSD)",                3,   8,  10,  6,  27,
         "TSD, data model sign-off, mockup"),
        ("3", "Build (lihat sheet Build Detail)",18,  35, 123, 44, 220,
         "Sprint 2-mingguan → breakdown di sheet 'Build Detail (202)'"),
        ("4", "SIT",                             5,   6,  12, 16,  39,
         "E2E pass — gated: konektor SAP/Kafka/HRIS ready"),
        ("5", "UAT",                             5,  12,  12,  8,  37,
         "UAT sign-off user"),
        ("6", "Cutover & Go-Live",               4,   5,   8,  3,  20,
         "Parallel run, cutover day"),
        ("7", "Hypercare",                       4,   5,   9,  3,  21,
         "Stabilisasi 3 minggu, handover support"),
    ]
    for i, (no, fase, pmo, ba, dev, qa, ttl, note) in enumerate(phases):
        r = 10 + i
        bg = ALT if i % 2 else WHITE
        for col, v, num in [
            (2, no, False), (3, fase, False), (4, pmo, True),
            (5, ba, True),  (6, dev, True),   (7, qa, True),
            (8, ttl, True), (9, note, False),
        ]:
            w(ws, r, col, v, bg=bg, h="center" if num else "left")
        ws.row_dimensions[r].height = 22

    r = 17
    sub(ws, r, 2, ""); sub(ws, r, 3, "Subtotal (tanpa kontingensi)")
    for col, v in [(4,47),(5,93),(6,177),(7,84),(8,401)]: sub(ws, r, col, v)
    w(ws, r, 9, "", bg=LILA)

    r = 18
    for col, v in [(2,""),(3,"+ Kontingensi 15%"),(4,7),(5,14),(6,27),(7,13),(8,61)]:
        w(ws, r, col, v, bg=LILA, bold=True,
          h="center" if isinstance(v, int) else "left")
    w(ws, r, 9, "", bg=LILA)

    r = 19
    tot(ws, r, 2, ""); tot(ws, r, 3, "TOTAL (dengan kontingensi)")
    for col, v in [(4,54),(5,107),(6,204),(7,97),(8,462)]: tot(ws, r, col, v)
    w(ws, r, 9, "Grand total termasuk PMO", bg=LILA2, bold=True)

    r = 20
    tot(ws, r, 2, ""); tot(ws, r, 3, "Total Non-PMO  =  462 − 54")
    for col, v in [(4,"—"),(5,107),(6,204),(7,97),(8,408)]: tot(ws, r, col, v)
    w(ws, r, 9, "Angka ini setara spreadsheet klien (tanpa PMO)", bg=LILA2,
      italic=True, sz=9)

    r = 22
    merge_w(ws, r, 2, 9,
        "Keterangan: PMO (54 md) dikeluarkan karena spreadsheet klien hanya mencantumkan kolom IT BA & IT Dev. "
        "Angka 202 = Fase Build saja (BA 35 + Dev 123 + QA 44). "
        "Angka 408 = semua 7 fase + kontingensi 15%, tanpa PMO.",
        bg=ALT, italic=True, sz=9)
    ws.row_dimensions[r].height = 32


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Semua Fase (408)
# ─────────────────────────────────────────────────────────────────────────────
def build_semua_fase(wb):
    ws = wb.create_sheet("Semua Fase (408)")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    for ltr, wd in zip("ABCDEFGHI", [4, 38, 10, 10, 10, 10, 12, 12, 44]):
        ws.column_dimensions[ltr].width = wd
    ws.row_dimensions[1].height = 30

    merge_title(ws, 1, 2, 9,
        "Estimasi Mandays — Semua Fase  (Total Non-PMO = 408 md  ·  incl. PMO = 462 md)", sz=13)

    for col, lbl in enumerate(["No","Fase / Aktivitas","PMO","BA","Dev","QA","Total","% dari total","Keterangan"], 1):
        hdr(ws, 2, col, lbl)
    ws.row_dimensions[2].height = 20

    phases = [
        ("1","Requirement & Analysis",      8, 22,  3,  4,  37,
         "FSD draft, stakeholder interview, kontrak integrasi final"),
        ("2","Design (FSD/TSD)",             3,  8, 10,  6,  27,
         "TSD & data model sign-off, UI mockup, JSON payload spec"),
        ("3","Build / Development",         18, 35,123, 44, 220,
         "Sprint 2-mingguan, feature complete → detail di sheet 'Build Detail (202)'"),
        ("4","SIT",                          5,  6, 12, 16,  39,
         "Gated: konektor SAP/Kafka/HRIS ready; E2E pass"),
        ("5","UAT",                          5, 12, 12,  8,  37,
         "UAT sign-off dari user"),
        ("6","Cutover & Go-Live",            4,  5,  8,  3,  20,
         "Data migration, parallel run, cutover day"),
        ("7","Hypercare",                    4,  5,  9,  3,  21,
         "Stabilisasi 3 minggu, handover ke support"),
    ]
    for i, (no, fase, pmo, ba, dev, qa, ttl, note) in enumerate(phases):
        r = 3 + i
        bg = ALT if i % 2 else WHITE
        pct = f"{ttl/401*100:.1f}%"
        for col, v, num in [
            (1, no, False), (2, fase, False), (3, pmo, True),
            (4, ba, True),  (5, dev, True),   (6, qa, True),
            (7, ttl, True), (8, pct, True),   (9, note, False),
        ]:
            w(ws, r, col, v, bg=bg, h="center" if num else "left")
        ws.row_dimensions[r].height = 24

    r = 10
    sub(ws, r, 1, ""); sub(ws, r, 2, "Subtotal (tanpa kontingensi)")
    for col, v in [(3,47),(4,93),(5,177),(6,84),(7,401),(8,"100%")]: sub(ws, r, col, v)
    w(ws, r, 9, "", bg=LILA)

    r = 11
    for col, v in [(1,""),(2,"+ Kontingensi 15%"),(3,7),(4,14),(5,27),(6,13),(7,61),(8,"15%")]:
        w(ws, r, col, v, bg=LILA, bold=True,
          h="center" if isinstance(v, (int, float)) or "%" in str(v) else "left")
    w(ws, r, 9, "", bg=LILA)

    r = 12
    tot(ws, r, 1, ""); tot(ws, r, 2, "TOTAL (dengan kontingensi)")
    for col, v in [(3,54),(4,107),(5,204),(6,97),(7,462),(8,"115%")]: tot(ws, r, col, v)
    w(ws, r, 9, "Grand total — termasuk PMO", bg=LILA2, bold=True)

    r = 13
    tot(ws, r, 1, ""); tot(ws, r, 2, "Total Non-PMO  =  462 − 54 PMO")
    for col, v in [(3,"—"),(4,107),(5,204),(6,97),(7,408),(8,"")]: tot(ws, r, col, v)
    w(ws, r, 9, "Setara spreadsheet klien — tanpa PMO", bg=LILA2, italic=True, sz=9)

    # note box
    r = 15
    merge_w(ws, r, 1, 9,
        "Catatan: Angka di tabel ini adalah TOTAL PROYEK (semua 7 fase). "
        "Sheet 'Build Detail (202)' berisi breakdown per-fitur untuk Fase 3 saja. "
        "Fase 1-2-4-5-6-7 tidak di-breakdown per-fitur karena aktivitasnya berupa kegiatan (workshop, review, testing), bukan deliverable fitur.",
        bg=ALT, italic=True, sz=9, h="left")
    ws.row_dimensions[r].height = 40


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Build Detail (202)
# ─────────────────────────────────────────────────────────────────────────────

# ── master data ──
WORKSTREAMS = {
    "A": ("Foundation & SSO",               2,  12,  3),
    "B": ("Master Data & Sync",             6,  14,  4),
    "C": ("Cash Advance + Realization",     4,  14,  5),
    "D": ("Reimbursement & Expenses",       3,   9,  4),
    "E": ("Vendor Invoice + Vendor Portal", 4,  18,  6),
    "F": ("Perjalanan Dinas",               3,   7,  3),
    "G": ("Budget Control",                 3,   8,  3),
    "H": ("Integration Odoo-side",          4,  16,  6),
    "I": ("Reporting + Dashboard",          3,  12,  4),
    "J": ("Configuration",                  2,   4,  2),
    "K": ("Non-Functional",                 1,   9,  4),
}

# (no, ws, modul, fitur, sub_fitur, komponen, interfacing, complexity, ba, dev, qa)
FEATURE_ROWS = [
    # ── A. Foundation & SSO ─────────────────────────────────────────────────
    ("A.1","A","Foundation","Multi-company base scaffold",
     "Manifest, __init__, install check, base model mixin",
     "custom_finance_portal","Internal","Low",0,2,0.5),
    ("A.2","A","Foundation","Keycloak SSO sisi Odoo",
     "auth_oauth config, role mapping, tenant isolation, redirect URI",
     "custom_finance_portal_sso","SSO/Keycloak","High",1,6,1.5),
    ("A.3","A","Foundation","Security groups & access rules",
     "Groups per role (Submitter / Tax / Finance / Admin), record rules",
     "custom_finance_portal","Internal","Low",0.5,2,0.5),
    ("A.4","A","Foundation","CI/CD & static validation",
     "py_compile, XML/ACL lint, install smoke test script",
     "DevOps","Internal","Low",0,1,0),
    ("A.5","A","Foundation","Vendor Portal SSO login",
     "Vendor auth flow via auth_oauth, session management sisi Odoo",
     "custom_finance_portal","SSO/Keycloak","Low",0.5,1,0.5),

    # ── B. Master Data & Sync ────────────────────────────────────────────────
    ("B.1","B","Master Data","Submission Type",
     "CRUD + seed data (CA, Expense, PD, Invoice)",
     "custom_finance_portal","Internal","Low",0.5,0.5,0),
    ("B.2","B","Master Data","Invoice Routine Type",
     "CRUD (Routine / Non-Routine)",
     "custom_finance_portal","Internal","Low",0.5,0.5,0),
    ("B.3","B","Master Data","Invoice Type",
     "CRUD (PO / Non-PO Non-Trade)",
     "custom_finance_portal","Internal","Low",0.5,0.5,0),
    ("B.4","B","Master Data","Item Category",
     "Odoo model + upsert idempoten dari SAP webhook",
     "custom_finance_portal_sap","SAP-API","Low",0.5,1.5,0.5),
    ("B.5","B","Master Data","Item of Submission",
     "Odoo model + upsert idempoten dari SAP webhook",
     "custom_finance_portal_sap","SAP-API","Low",0.5,1.5,0.5),
    ("B.6","B","Master Data","Supplier / Vendor Master",
     "Upsert res.partner dari SAP webhook, dedup by vendor code",
     "custom_finance_portal_sap","SAP-API","Low",0.5,1.5,0.5),
    ("B.7","B","Master Data","COA (Chart of Accounts)",
     "Upsert account.account (filter GL relevan) dari SAP",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,1),
    ("B.8","B","Master Data","Cost Budget per Divisi",
     "Model budget + upsert dari SAP",
     "custom_finance_budget","SAP-API","Low",0.5,2,0.5),
    ("B.9","B","Master Data","Approval Matrix",
     "Upsert approval matrix + mapping approver Odoo user",
     "custom_finance_portal_sap","SAP-API","Low",0.5,1.5,0.5),
    ("B.10","B","Master Data","Finance Approval Matrix",
     "CRUD local — mapping PIC Finance per vertical",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("B.11","B","Master Data","Vertical / Business Plant",
     "Upsert dari SAP, linked ke res.company",
     "custom_finance_portal_sap","SAP-API","Low",0.5,0.5,0),
    ("B.12","B","Master Data","User Master (sync SAP/HRIS)",
     "Upsert res.users + cron job scheduler",
     "custom_finance_portal_sap","SAP-API/HRIS","Low",0,0.5,0),
    ("B.13","B","Master Data","User Role (sync SAP)",
     "Upsert res.groups mapping dari role SAP",
     "custom_finance_portal_sap","SAP-API","Low",0.5,0.5,0.5),

    # ── C. Cash Advance + Realization ────────────────────────────────────────
    ("C.1","C","Cash Advance","Request CA – List & Filter",
     "List/kanban view, search panel, filter status/date/CA number",
     "custom_finance_portal","Internal","Low",0.5,1.5,0.5),
    ("C.2","C","Cash Advance","Request CA – Create/Edit form",
     "Form: CA No, PR/PO, requester, NIK, company, division, amount, tgl payment, metode, bank, rekening, penerima, approver, note",
     "custom_finance_portal","Internal","High",1,3,1),
    ("C.3","C","Cash Advance","Validasi PR (> Rp 1jt wajib PR)",
     "Constraint + lookup nomor PR dari SAP (real-time)",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,0.5),
    ("C.4","C","Cash Advance","Approval workflow Tax→Finance",
     "Approval engine 2 level: Tax Review → Finance Review",
     "custom_finance_portal","Internal","Low",0.5,1.5,0.5),
    ("C.5","C","Cash Advance","Push CA ke SAP",
     "Adapter: POST ke SAP endpoint, async via queue_job",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,0.5),
    ("C.6","C","CA Realization","Realization CA – List & Filter",
     "List view, filter by CA/status",
     "custom_finance_portal","Internal","Low",0,0.5,0.5),
    ("C.7","C","CA Realization","Realization CA – Create/Edit form",
     "Link ke CA asal, tabel detail realisasi, sisa budget, tgl realisasi",
     "custom_finance_portal","Internal","Medium",1,2,1),
    ("C.8","C","CA Realization","Approval Realization Tax→Finance",
     "Reuse approval engine (beda state machine)",
     "custom_finance_portal","Internal","Low",0,0.5,0.5),
    ("C.9","C","CA Realization","Terima status payment dari SAP",
     "Webhook handler: update state CA + notif user",
     "custom_finance_portal_sap","SAP-API","Low",0,1,0),

    # ── D. Reimbursement & Expenses ──────────────────────────────────────────
    ("D.1","D","Reimbursement","List & Filter",
     "List view, filter by status/type/date",
     "custom_finance_portal","Internal","Low",0,1.5,0.5),
    ("D.2","D","Reimbursement","Create/Edit form",
     "Form field expense/reimbursement + attachment",
     "custom_finance_portal","Internal","High",1,3,1),
    ("D.3","D","Reimbursement","Validasi PR (> Rp 1jt)",
     "Reuse constraint + SAP PR lookup dari C.3",
     "custom_finance_portal_sap","SAP-API","Low",0.5,0.5,0.5),
    ("D.4","D","Reimbursement","Approval workflow Tax→Finance",
     "Reuse approval engine",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("D.5","D","Reimbursement","Push expense ke SAP",
     "Adapter: POST ke SAP endpoint, async via queue_job",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,0.5),
    ("D.6","D","Reimbursement","Terima status payment dari SAP",
     "Webhook handler (reuse, beda state machine)",
     "custom_finance_portal_sap","SAP-API","Low",0.5,1,1),

    # ── E. Vendor Invoice + Vendor Portal ────────────────────────────────────
    ("E.1","E","Invoice Non-PO","List & Filter",
     "List view + filter status/date/vendor",
     "custom_finance_portal","Internal","Low",0,1.5,0.5),
    ("E.2","E","Invoice Non-PO","Create/Edit form Non-PO",
     "Form: nomor inv, GL Account, COA, attachment, approval",
     "custom_finance_portal","Internal","High",1,3,1),
    ("E.3","E","Invoice Non-PO","Approval workflow Tax→Finance",
     "Reuse approval engine",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("E.4","E","Invoice Non-PO","Push MIRO posting ke SAP",
     "Adapter: POST invoice posting ke SAP, async queue_job",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,1),
    ("E.5","E","Invoice PO","List & Filter PO-linked",
     "List view + lookup PO/GR dari SAP",
     "custom_finance_portal","SAP-API","Low",0,1.5,0.5),
    ("E.6","E","Invoice PO","Create form PO-linked",
     "Auto-populate dari PO/GR SAP, validasi qty/nilai GR",
     "custom_finance_portal_sap","SAP-API","High",1,3,1),
    ("E.7","E","Invoice PO","Approval workflow Tax→Finance",
     "Reuse approval engine",
     "custom_finance_portal","Internal","Low",0,1,0.5),
    ("E.8","E","Invoice PO","Push MIRO posting ke SAP (PO-linked)",
     "Adapter (reuse E.4, beda payload PO-linked)",
     "custom_finance_portal_sap","SAP-API","Low",0.5,1.5,0.5),
    ("E.9","E","Vendor Portal","Login Vendor (SSO) + landing",
     "Vendor auth flow via auth_oauth, landing page submit invoice",
     "custom_finance_portal","SSO/Keycloak","Low",0.5,2,0.5),
    ("E.10","E","Vendor Portal","Submit invoice + tracking status",
     "Vendor-facing form + polling status dari SAP",
     "custom_finance_portal_sap","SAP-API","Low",0,1.5,0.5),

    # ── F. Perjalanan Dinas ───────────────────────────────────────────────────
    ("F.1","F","PD Request","List & Filter",
     "List view + filter by date/status",
     "custom_finance_portal","HRIS-API","Low",0,1,0.5),
    ("F.2","F","PD Request","Create/Edit form",
     "Form field sesuai kontrak HRIS: tujuan, tanggal, anggaran, tipe PD",
     "custom_finance_portal","HRIS-API","Medium",1,2,0.5),
    ("F.3","F","PD Request","Approval workflow Tax→Finance",
     "Reuse approval engine",
     "custom_finance_portal","Internal","Low",0.5,0.5,0),
    ("F.4","F","PD Realization","List & Filter",
     "List view",
     "custom_finance_portal","Internal","Low",0,1,0.5),
    ("F.5","F","PD Realization","Create/Edit form Realization",
     "Settlement: link ke PD asal, actual expense, lampiran bukti",
     "custom_finance_portal","Internal","Low",1,1.5,1),
    ("F.6","F","PD Realization","Approval Realization",
     "Reuse approval engine",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),

    # ── G. Budget Control ────────────────────────────────────────────────────
    ("G.1","G","Budget Control","Budget limit check per divisi",
     "Computed field sisa budget pada form dokumen",
     "custom_finance_budget","Internal","Low",0.5,2,0.5),
    ("G.2","G","Budget Control","Validasi & lookup PR (≥ Rp 1jt)",
     "SAP PR lookup adapter + onchange constraint",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,0.5),
    ("G.3","G","Budget Control","Lookup & validasi PO/GR",
     "SAP PO/GR adapter + populate fields",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,0.5),
    ("G.4","G","Budget Control","Remaining budget display (computed)",
     "Computed field: budget − committed − realisasi",
     "custom_finance_budget","Internal","Low",0.5,1,0),
    ("G.5","G","Budget Control","Alert + blocking rule over-budget",
     "Warning popup + hard-block state constraint",
     "custom_finance_budget","Internal","Low",1,1,1.5),

    # ── H. Integration Odoo-side ─────────────────────────────────────────────
    ("H.1","H","Integration","Adapter base (REST caller + retry)",
     "Base class HTTP REST, timeout, retry 3×, error mapping",
     "custom_finance_portal_sap","SAP-API","High",1,3,1),
    ("H.2","H","Integration","Push job async via queue_job",
     "CA, expense, MIRO push job + priority queue",
     "custom_finance_portal_sap","Internal","Low",0.5,3,1),
    ("H.3","H","Integration","Webhook receiver + HMAC validator",
     "HTTP controller endpoint, signature validation, dispatch handler",
     "custom_finance_portal_sap","SAP-API","Low",0.5,3,1),
    ("H.4","H","Integration","Sync log menu",
     "Tree/form view log (sukses/gagal/pending) per dokumen + retry button",
     "custom_finance_portal_sap","Internal","Low",0.5,3,1),
    ("H.5","H","Integration","Master data upsert cron",
     "ir.cron jobs untuk upsert harian semua master dari SAP",
     "custom_finance_portal_sap","SAP-API","Low",0.5,2,1),
    ("H.6","H","Integration","Dead-letter retry + alert email",
     "queue_job failure handler, email notif ke admin",
     "custom_finance_portal_sap","Internal","Low",1,2,1),

    # ── I. Reporting + Dashboard ─────────────────────────────────────────────
    ("I.1","I","Dashboard","Cash Advance Dashboard",
     "Stat blocks (New/On Process/Reject/Pending) + filter date + search CA#",
     "OWL/ir.ui.view","Internal","Low",0.5,2,0.5),
    ("I.2","I","Dashboard","Reimbursement Dashboard",
     "Stat blocks + list mini",
     "OWL/ir.ui.view","Internal","Low",0,1,0),
    ("I.3","I","Dashboard","Invoice Vendor Dashboard",
     "Stat blocks (Non-PO & PO) + filter",
     "OWL/ir.ui.view","Internal","Low",0,1,0),
    ("I.4","I","Dashboard","Perjalanan Dinas Dashboard",
     "Stat blocks + list mini",
     "OWL/ir.ui.view","Internal","Low",0,1,0),
    ("I.5","I","Reporting","Login log",
     "Wrapper ir.logging + list view + filter user/date",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("I.6","I","Reporting","Transaction audit log",
     "Audit trail model (create/update/approve/reject) + wizard export",
     "custom_finance_portal","Internal","Low",0.5,2,1),
    ("I.7","I","Reporting","Cash Advance report (Excel/PDF)",
     "QWeb report template + xlsx export",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("I.8","I","Reporting","Reimbursement report",
     "QWeb report template + xlsx export",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("I.9","I","Reporting","Invoice Vendor report",
     "QWeb report template + xlsx export",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("I.10","I","Reporting","Sync log report / export",
     "List view + CSV export",
     "custom_finance_portal_sap","Internal","Low",0,1,0.5),

    # ── J. Configuration ─────────────────────────────────────────────────────
    ("J.1","J","Configuration","Finance Portal settings panel",
     "ir.config_parameter + settings view (base URL SAP, secret token, flags)",
     "custom_finance_portal","Internal","Low",1,2,0.5),
    ("J.2","J","Configuration","Limitation for Submission master",
     "Model + CRUD view (limit per submission type/divisi)",
     "custom_finance_portal","Internal","Low",0.5,1,0.5),
    ("J.3","J","Configuration","Sync config (endpoint, topic, token)",
     "Settings form per integrasi + test-connection button",
     "custom_finance_portal_sap","Internal","Low",0.5,1,1),

    # ── K. Non-Functional ────────────────────────────────────────────────────
    ("K.1","K","Non-Functional","PDP classification audit",
     "Data category fields, access log, custom_pdp_classification reuse",
     "custom_pdp_classification","Internal","Low",0.5,2,1),
    ("K.2","K","Non-Functional","Security hardening",
     "SQL injection, XSS, IDOR review + perbaikan sisi Odoo",
     "All modules","Internal","Low",0,2,1),
    ("K.3","K","Non-Functional","Performance (N+1 / prefetch)",
     "Profiling query, prefetch_ids, pagination",
     "All modules","Internal","Low",0,2,1),
    ("K.4","K","Non-Functional","Resilience (circuit breaker + timeout)",
     "Adapter timeout, fallback state, exponential backoff",
     "custom_finance_portal_sap","Internal","Low",0.5,3,1),
]


def build_detail(wb):
    ws = wb.create_sheet("Build Detail (202)")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    COL_W = [5, 5, 16, 22, 38, 24, 13, 11, 6, 6, 6, 7]
    for i, wd in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.row_dimensions[1].height = 28

    merge_title(ws, 1, 1, 12,
        "Build Phase Detail — Finance Portal Odoo  "
        "(BA 35 + Dev 123 + QA 44 = 202 mandays, tanpa PMO & kontingensi)", sz=12)

    HDR = ["No","WS","Modul","Fitur / Menu","Sub-Fitur / Task Odoo",
           "Komponen Odoo","Interfacing","Complexity","BA","Dev","QA","Total"]
    for col, lbl in enumerate(HDR, 1):
        hdr(ws, 2, col, lbl)
    ws.row_dimensions[2].height = 22

    cur_ws_code = None
    r = 3
    alt = False

    for row_data in FEATURE_ROWS:
        no, ws_code, modul, fitur, sub_fitur, komp, intf, cplx, ba, dev, qa = row_data
        total = round(ba + dev + qa, 2)

        # ── workstream section header ──
        if ws_code != cur_ws_code:
            cur_ws_code = ws_code
            ws_name, ws_ba, ws_dev, ws_qa = WORKSTREAMS[ws_code]
            ws_tot = ws_ba + ws_dev + ws_qa

            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            cell = ws.cell(row=r, column=1,
                           value=f"  {ws_code}.  {ws_name}")
            cell.fill = fill(PURPLE); cell.font = fnt(bold=True, color=WHITE, size=10)
            cell.alignment = aln(h="left"); cell.border = BORD
            for col, v in enumerate([ws_ba, ws_dev, ws_qa, ws_tot], 9):
                w(ws, r, col, v, bg=PURPLE, bold=True, color=WHITE, h="center")
            ws.row_dimensions[r].height = 18
            r += 1
            alt = False

        # ── data row ──
        bg = ALT if alt else WHITE
        for col, v, num in [
            (1, no, False),  (2, ws_code, False), (3, modul, False),
            (4, fitur, False),(5, sub_fitur, False),(6, komp, False),
            (7, intf, False), (8, cplx, False),
            (9, ba, True),   (10, dev, True),      (11, qa, True), (12, total, True),
        ]:
            cell = w(ws, r, col, v, bg=bg, h="center" if num else "left")
            if num and isinstance(v, float) and v != int(v):
                cell.number_format = "0.0"
        ws.row_dimensions[r].height = 28
        r += 1
        alt = not alt

    # ── grand total ──
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell = ws.cell(row=r, column=1,
                   value="GRAND TOTAL FASE BUILD  (tanpa PMO, tanpa kontingensi)")
    cell.fill = fill(LILA2); cell.font = fnt(bold=True, color=NAVY, size=11)
    cell.alignment = aln(h="center"); cell.border = BORD
    for col, v in [(9, 35), (10, 123), (11, 44), (12, 202)]:
        tot(ws, r, col, v)
    ws.row_dimensions[r].height = 24

    # ── complexity legend (below grand total) ──
    r += 2
    merge_title(ws, r, 1, 12, "Rumus Kompleksitas", bg=PURPLE, sz=10)
    r += 1
    for col, lbl in enumerate(["Complexity","BA","Dev","QA","Total/baris"], 1):
        hdr(ws, r, col, lbl)
    r += 1
    for cplx, ba, dev, qa in [("Low",0.5,1.5,0.5),("Medium",1,3,1),("High",2,5,2)]:
        for col, v in enumerate([cplx, ba, dev, qa, ba+dev+qa], 1):
            w(ws, r, col, v, bg=ALT if cplx=="Medium" else WHITE,
              h="center" if isinstance(v,(int,float)) else "left")
        ws.row_dimensions[r].height = 18
        r += 1


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    build_ringkasan(wb)
    build_semua_fase(wb)
    build_detail(wb)
    wb.save(OUT)
    print(f"Saved: {OUT}")

if __name__ == "__main__":
    main()
