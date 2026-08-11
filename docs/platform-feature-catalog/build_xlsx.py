#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Render ``catalog.json`` into the feature-matrix workbook.

Six sheets: Ringkasan, Matriks Fitur, Peta Brand, Domain x Grup, Kesenjangan,
Sumber & Metode. Styling follows ``docs/projects/finance-portal/build_estimation_xlsx.py``
so the two client deliverables look like they came from the same place.

No charts: openpyxl charts survive a LibreOffice or Google Sheets round-trip
badly, and the reader will re-pivot anyway.

Usage : python3 build_xlsx.py [--out PATH]
"""

from __future__ import annotations

import argparse
import collections
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT = os.path.join(HERE, "dist", "Matriks_Fitur_Platform_Odoo_Erajaya.xlsx")

NAVY, PURPLE, LILA, LILA2, WHITE, ALT = "1A2332", "714B67", "E8E0E7", "C9B8C6", "FFFFFF", "F5F2F4"
GREEN, AMBER, RED, GREY = "D8EAD8", "FBEBCC", "F6D6D6", "EDEDED"

_T = Side(style="thin", color="CCCCCC")
BORD = Border(left=_T, right=_T, top=_T, bottom=_T)


def fill(hexcode):
    return PatternFill("solid", fgColor=hexcode)


def fnt(bold=False, color=NAVY, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def aln(h="left", wrap=True, rotate=0):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap, text_rotation=rotate)


def w(ws, r, c, v, bg=WHITE, bold=False, color=NAVY, sz=10, h="left", wrap=True, italic=False, rotate=0):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = fill(bg)
    cell.font = fnt(bold=bold, color=color, size=sz, italic=italic)
    cell.alignment = aln(h=h, wrap=wrap, rotate=rotate)
    cell.border = BORD
    return cell


def hdr(ws, r, c, v, bg=NAVY, rotate=0, h="center"):
    return w(ws, r, c, v, bg=bg, bold=True, color=WHITE, h=h, rotate=rotate)


def title(ws, r, c1, c2, v, bg=NAVY, sz=14, color=WHITE):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=v)
    cell.fill = fill(bg)
    cell.font = fnt(bold=True, color=color, size=sz)
    cell.alignment = aln(h="center")
    cell.border = BORD
    return cell


def widths(ws, spec: dict):
    for col, width in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# --- vocabulary -------------------------------------------------------------

SCOPE_LABEL = {
    "general": "Umum",
    "tenant": "Khusus Brand",
    "platform": "Platform",
    "vendor": "Pihak Ketiga",
}
MATURITY_LABEL = {
    "production": "Produksi",
    "beta": "Beta",
    "scaffold": "Kerangka",
    "vendor": "Vendor",
    "disabled": "Nonaktif",
}
MATURITY_BG = {
    "production": GREEN, "beta": AMBER, "scaffold": GREY,
    "vendor": GREY, "disabled": RED,
}
CONFIDENCE_LABEL = {"high": "Tinggi", "medium": "Sedang", "low": "Rendah"}
CONFIDENCE_BG = {"high": GREEN, "medium": AMBER, "low": RED}
KNOWLEDGE_LABEL = {
    "override": "Override (ditulis tangan)",
    "reviewed": "Ada — reviewed",
    "draft": "Ada — draft",
    "missing": "Tidak ada",
}
SEVERITY_LABEL = {"high": "Tinggi", "medium": "Sedang", "low": "Rendah"}
SEVERITY_BG = {"high": RED, "medium": AMBER, "low": GREEN}


def scope_label(mod) -> str:
    """Three honest statements, not two.

    A generic engine that already ships one brand's profile is neither "umum"
    nor "khusus brand" — calling it either misleads a reader deciding whether a
    second tenant can use it as-is.
    """
    if mod["scope"] == "general" and mod["tenants"]:
        return "Umum, dikonfigurasi"
    return SCOPE_LABEL.get(mod["scope"], mod["scope"])


# --- sheets -----------------------------------------------------------------


def sheet_ringkasan(wb, cat, brand_by_id):
    ws = wb.create_sheet("Ringkasan")
    widths(ws, {1: 42, 2: 12, 3: 4, 4: 42, 5: 12})
    meta, counts = cat["meta"], cat["meta"]["counts"]

    title(ws, 1, 1, 5, "Katalog Fitur Platform Odoo — Erajaya Group", sz=15)
    title(ws, 2, 1, 5, "Matriks Fitur · Ringkasan", bg=PURPLE, sz=11)

    r = 4
    w(ws, r, 1, "Total modul custom", bg=LILA, bold=True)
    w(ws, r, 2, counts["modules_total"], bg=LILA, bold=True, h="center")
    w(ws, r, 4, "Basis data tenant terdaftar", bg=LILA, bold=True)
    w(ws, r, 5, len(cat["tenants"]), bg=LILA, bold=True, h="center")

    r += 2
    w(ws, r, 1, "Modul per grup teknis", bg=PURPLE, bold=True, color=WHITE)
    w(ws, r, 2, "Jumlah", bg=PURPLE, bold=True, color=WHITE, h="center")
    w(ws, r, 4, "Modul per domain fungsional", bg=PURPLE, bold=True, color=WHITE)
    w(ws, r, 5, "Jumlah", bg=PURPLE, bold=True, color=WHITE, h="center")

    groups = sorted(counts["by_group"].items())
    domains = [(d["title_id"], d["module_count"]) for d in cat["domains"]]
    start = r + 1
    for i in range(max(len(groups), len(domains))):
        rr = start + i
        bg = ALT if i % 2 else WHITE
        if i < len(groups):
            w(ws, rr, 1, f"addons/{groups[i][0]}/", bg=bg)
            w(ws, rr, 2, groups[i][1], bg=bg, h="center")
        if i < len(domains):
            w(ws, rr, 4, domains[i][0], bg=bg)
            w(ws, rr, 5, domains[i][1], bg=bg, h="center")
    r = start + max(len(groups), len(domains))
    w(ws, r, 1, "Total", bg=LILA2, bold=True)
    w(ws, r, 2, sum(v for _k, v in groups), bg=LILA2, bold=True, h="center")
    w(ws, r, 4, "Total", bg=LILA2, bold=True)
    w(ws, r, 5, sum(v for _k, v in domains), bg=LILA2, bold=True, h="center")

    r += 2
    w(ws, r, 1, "Cakupan", bg=PURPLE, bold=True, color=WHITE)
    w(ws, r, 2, "Jumlah", bg=PURPLE, bold=True, color=WHITE, h="center")
    w(ws, r, 4, "Keyakinan informasi", bg=PURPLE, bold=True, color=WHITE)
    w(ws, r, 5, "Jumlah", bg=PURPLE, bold=True, color=WHITE, h="center")

    scope_tally = collections.Counter(scope_label(m) for m in cat["modules"])
    conf = [(CONFIDENCE_LABEL[k], counts["by_confidence"].get(k, 0)) for k in ("high", "medium", "low")]
    scopes = sorted(scope_tally.items(), key=lambda kv: -kv[1])
    start = r + 1
    for i in range(max(len(scopes), len(conf))):
        rr = start + i
        bg = ALT if i % 2 else WHITE
        if i < len(scopes):
            w(ws, rr, 1, scopes[i][0], bg=bg)
            w(ws, rr, 2, scopes[i][1], bg=bg, h="center")
        if i < len(conf):
            w(ws, rr, 4, conf[i][0], bg=bg)
            w(ws, rr, 5, conf[i][1], bg=bg, h="center")
    r = start + max(len(scopes), len(conf)) + 1

    w(ws, r, 1, "Dokumen pengetahuan modul", bg=PURPLE, bold=True, color=WHITE)
    w(ws, r, 2, "Jumlah", bg=PURPLE, bold=True, color=WHITE, h="center")
    for i, (label, key) in enumerate([
        ("Override ditulis tangan", "knowledge_override"),
        ("MODULE_KNOWLEDGE.md — reviewed", "knowledge_reviewed"),
        ("MODULE_KNOWLEDGE.md — draft", None),
        ("Belum ada (diringkas dari manifest)", "knowledge_missing"),
    ]):
        rr = r + 1 + i
        val = (counts["knowledge_present"] - counts["knowledge_reviewed"] - counts["knowledge_override"]
               if key is None else counts[key])
        w(ws, rr, 1, label, bg=ALT if i % 2 else WHITE)
        w(ws, rr, 2, val, bg=ALT if i % 2 else WHITE, h="center")
    r += 6

    w(ws, r, 1, "Dihasilkan otomatis dari repositori", bg=LILA, bold=True)
    w(ws, r, 2, meta["generated_at"][:10], bg=LILA, h="center")
    w(ws, r, 4, "Commit", bg=LILA, bold=True)
    w(ws, r, 5, f"{meta['git_commit']}{'+' if meta['git_dirty'] else ''}", bg=LILA, h="center")
    ws.freeze_panes = "A4"
    return ws


MATRIX_COLS = [
    ("No", 5), ("Modul (teknis)", 34), ("Nama Fitur", 34), ("Domain", 26),
    ("Domain Sekunder", 24), ("Grup Teknis", 14), ("Cakupan", 18),
    ("Brand Terkait", 22), ("Ringkasan", 62), ("Kematangan", 12),
    ("Keyakinan Info", 13), ("Dok. Modul", 22), ("Model", 7), ("Endpoint", 9),
    ("Test", 7), ("Versi", 14), ("Dependensi", 40), ("Tag Kapabilitas", 34),
    ("Path", 46),
]


def sheet_matrix(wb, cat, domain_title, brand_by_id):
    ws = wb.create_sheet("Matriks Fitur")
    widths(ws, {i + 1: wdt for i, (_h, wdt) in enumerate(MATRIX_COLS)})
    for i, (head, _wd) in enumerate(MATRIX_COLS):
        hdr(ws, 1, i + 1, head)

    for i, mod in enumerate(sorted(cat["modules"], key=lambda m: (m["domain"], m["name"]))):
        r = i + 2
        bg = ALT if i % 2 else WHITE
        src, man, k = mod["source"], mod["manifest"], mod["knowledge"]
        row = [
            i + 1,
            mod["name"],
            mod["id"]["nama"],
            domain_title[mod["domain"]],
            ", ".join(domain_title[d] for d in mod["domain_secondary"]),
            mod["group"],
            scope_label(mod),
            ", ".join(brand_by_id[t] for t in mod["tenants"]),
            mod["id"]["ringkasan"],
            MATURITY_LABEL.get(mod["maturity"], mod["maturity"]),
            CONFIDENCE_LABEL[mod["confidence"]],
            KNOWLEDGE_LABEL.get(k["status"], k["status"]),
            len(src["models_own"]),
            len(src["routes"]),
            "Ya" if src["has_tests"] else "Tidak",
            man["version"],
            ", ".join(man["depends"]),
            ", ".join(man["capability_tags"]),
            mod["relpath"],
        ]
        for c, val in enumerate(row, start=1):
            centered = c in (1, 13, 14, 15)
            w(ws, r, c, val, bg=bg, h="center" if centered else "left")
        # Colour the two judgement columns so a reader scanning the sheet sees
        # where the catalog is confident and where it is not.
        ws.cell(row=r, column=10).fill = fill(MATURITY_BG.get(mod["maturity"], GREY))
        ws.cell(row=r, column=11).fill = fill(CONFIDENCE_BG[mod["confidence"]])

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MATRIX_COLS))}{len(cat['modules']) + 1}"
    return ws


def sheet_brand_map(wb, cat):
    ws = wb.create_sheet("Peta Brand")
    tenants = cat["tenants"]
    widths(ws, {1: 34, 2: 30, **{i + 3: 9 for i in range(len(tenants))}})
    ws.row_dimensions[1].height = 96

    hdr(ws, 1, 1, "Modul (teknis)", h="left")
    hdr(ws, 1, 2, "Domain")
    for i, t in enumerate(tenants):
        hdr(ws, 1, i + 3, t["brand"], bg=PURPLE if t["erajaya"] else NAVY, rotate=90)

    domain_title = {d["id"]: d["title_id"] for d in cat["domains"]}
    for i, mod in enumerate(sorted(cat["modules"], key=lambda m: (m["domain"], m["name"]))):
        r = i + 2
        bg = ALT if i % 2 else WHITE
        w(ws, r, 1, mod["name"], bg=bg)
        w(ws, r, 2, domain_title[mod["domain"]], bg=bg)
        for j, t in enumerate(tenants):
            mark = ""
            if t["id"] in mod["tenants"]:
                mark = "●" if mod["scope"] == "tenant" else "○"
            w(ws, r, j + 3, mark, bg=bg, h="center")

    r = len(cat["modules"]) + 3
    w(ws, r, 1, "● khusus brand (addons/_tenants/)   ○ modul umum yang sudah "
                "membawa data atau konfigurasi brand tersebut", bg=LILA, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(tenants) + 2)
    ws.freeze_panes = "C2"
    return ws


def sheet_pivot(wb, cat):
    ws = wb.create_sheet("Domain x Grup")
    groups = sorted(cat["meta"]["counts"]["by_group"])
    widths(ws, {1: 40, **{i + 2: 15 for i in range(len(groups) + 1)}})

    hdr(ws, 1, 1, "Domain", h="left")
    for i, g in enumerate(groups):
        hdr(ws, 1, i + 2, g)
    hdr(ws, 1, len(groups) + 2, "Total", bg=PURPLE)

    grid = collections.Counter((m["domain"], m["group"]) for m in cat["modules"])
    for i, d in enumerate(cat["domains"]):
        r = i + 2
        bg = ALT if i % 2 else WHITE
        w(ws, r, 1, d["title_id"], bg=bg)
        for j, g in enumerate(groups):
            n = grid.get((d["id"], g), 0)
            w(ws, r, j + 2, n or "", bg=bg, h="center")
        w(ws, r, len(groups) + 2, d["module_count"], bg=LILA, bold=True, h="center")

    r = len(cat["domains"]) + 2
    w(ws, r, 1, "Total", bg=LILA2, bold=True)
    grand = 0
    for j, g in enumerate(groups):
        n = cat["meta"]["counts"]["by_group"][g]
        grand += n
        w(ws, r, j + 2, n, bg=LILA2, bold=True, h="center")
    w(ws, r, len(groups) + 2, grand, bg=PURPLE, bold=True, color=WHITE, h="center")

    assert grand == cat["meta"]["counts"]["modules_total"], (
        f"pivot grand total {grand} != {cat['meta']['counts']['modules_total']}"
    )
    ws.freeze_panes = "B2"
    return ws


GAP_COLS = [
    ("ID", 15), ("Area", 20), ("Kesenjangan", 40), ("Kondisi Saat Ini (NOW)", 62),
    ("Sasaran (TARGET)", 52), ("Dampak Bisnis", 52), ("Prioritas", 11),
    ("Upaya", 8), ("Horizon (bln)", 12), ("Rujukan", 52),
]


def sheet_gaps(wb, cat):
    ws = wb.create_sheet("Kesenjangan")
    widths(ws, {i + 1: wd for i, (_h, wd) in enumerate(GAP_COLS)})
    title(ws, 1, 1, len(GAP_COLS), "Analisis Kesenjangan — Kondisi Saat Ini vs Sasaran", sz=13)
    for i, (head, _wd) in enumerate(GAP_COLS):
        hdr(ws, 2, i + 1, head)

    for i, g in enumerate(cat["gaps"]):
        r = i + 3
        bg = ALT if i % 2 else WHITE
        row = [
            g["id"], g["area"], g["title_id"], g["now_id"], g["target_id"],
            g.get("impact_id", ""), SEVERITY_LABEL.get(g["severity"], g["severity"]),
            g.get("effort", ""), g.get("horizon", ""), "\n".join(g.get("evidence", [])),
        ]
        for c, val in enumerate(row, start=1):
            w(ws, r, c, val, bg=bg, h="center" if c in (7, 8, 9) else "left")
        ws.cell(row=r, column=7).fill = fill(SEVERITY_BG.get(g["severity"], GREY))
        ws.row_dimensions[r].height = 78

    ws.freeze_panes = "C3"
    return ws


def sheet_method(wb, cat):
    ws = wb.create_sheet("Sumber & Metode")
    widths(ws, {1: 26, 2: 100})
    meta, counts = cat["meta"], cat["meta"]["counts"]
    title(ws, 1, 1, 2, "Sumber data dan cara kolom ini dihasilkan", sz=13)

    rows = [
        ("Cara dihasilkan",
         "Seluruh angka di workbook ini dipindai langsung dari repositori oleh "
         "docs/platform-feature-catalog/build_catalog_json.py, bukan diketik. "
         "Menambah modul ke addons/ membuat proses build gagal sampai modul itu "
         "diklasifikasikan — tidak ada kategori penampung."),
        ("Waktu pemindaian", meta["generated_at"]),
        ("Commit", f"{meta['git_commit']} (branch {meta['git_branch']})"
                   + (" — ada perubahan belum ter-commit" if meta["git_dirty"] else "")),
        ("Seri Odoo", meta["odoo_series"]),
        ("Modul (teknis)", "Nama direktori di addons/. Penelusuran tanpa batas kedalaman: "
                           "verticals/_template/custom_vertical_example berada satu tingkat "
                           "lebih dalam, dan pencarian berbatas kedalaman menghasilkan 157, "
                           "bukan 158."),
        ("Nama Fitur & Ringkasan", "Ditulis tangan dalam Bahasa Indonesia. Entri yang belum "
                                   "diterjemahkan tampil dengan awalan [EN] agar terlihat, "
                                   "bukan tercampur diam-diam."),
        ("Domain", "Klasifikasi fungsional buatan manusia di taxonomy.py, satu domain primer "
                   "per modul. Domain sekunder hanya rujukan silang dan tidak dihitung."),
        ("Cakupan", "Umum = tersedia untuk tenant mana pun. Umum, dikonfigurasi = mesin generik "
                    "yang sudah membawa data satu brand. Khusus Brand = ada di addons/_tenants/ "
                    "dan tidak bisa dipakai ulang apa adanya. Platform = lapisan kendali. "
                    "Pihak Ketiga = komponen OCA."),
        ("Kematangan", "Diturunkan: ada suite test → Produksi; ada model/route/XML tapi tanpa "
                       "test → Beta; kosong → Kerangka. Sebelas modul dikoreksi manual karena "
                       "berjalan di produksi tanpa test."),
        ("Keyakinan Info", "Tinggi = override ditulis tangan atau berkas berstatus reviewed. "
                           "Sedang = MODULE_KNOWLEDGE.md berstatus draft hasil generator. "
                           "Rendah = tidak ada dokumen, atau audit menemukan model yang "
                           "disebut tetapi tidak ada di kode."),
        ("Arti status draft", f"{counts['knowledge_present'] - counts['knowledge_reviewed'] - counts['knowledge_override']} "
                              "berkas pengetahuan modul dihasilkan generator dan belum diperiksa "
                              "manusia. Isinya dipakai di katalog ini, tetapi selalu disertai "
                              "kolom Keyakinan Info — bukan disembunyikan dan bukan pula "
                              "diperlakukan setara dengan yang sudah diperiksa."),
        ("Gerbang akurasi", "build_catalog_json.py --audit membandingkan setiap klaim di berkas "
                            "pengetahuan dengan kode. Model yang disebut tetapi tidak ada di "
                            "mana pun menurunkan modul ke Keyakinan Rendah dan daftar modelnya "
                            "dibuang dari lampiran. Hasil lengkap: catalog-audit.md."),
        ("Temuan audit", ", ".join(f"{k}: {v}" for k, v in sorted(counts.get("audit_flags", {}).items()))
                         or "tidak ada"),
        ("Model / Endpoint / Test", "Dihitung dengan analisis statis atas models/, controllers/ "
                                    "dan tests/. Modul yang seluruhnya berupa controller (mis. "
                                    "custom_wms_hht) sah bernilai 0 model."),
        ("Kesenjangan", "Ditulis tangan di gaps.yaml. Setiap baris menyertakan rujukan berkas "
                        "yang bisa dibuka untuk memeriksa klaimnya."),
    ]
    for i, (label, text) in enumerate(rows):
        r = i + 3
        bg = ALT if i % 2 else WHITE
        w(ws, r, 1, label, bg=LILA, bold=True)
        w(ws, r, 2, text, bg=bg)
        ws.row_dimensions[r].height = max(15, 13 * (len(text) // 95 + 1))
    return ws


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--catalog", default=os.path.join(HERE, "catalog.json"))
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    with open(args.catalog, "r", encoding="utf-8") as fh:
        cat = json.load(fh)

    domain_title = {d["id"]: d["title_id"] for d in cat["domains"]}
    brand_by_id = {t["id"]: t["brand"] for t in cat["tenants"]}

    wb = Workbook()
    wb.remove(wb.active)
    sheet_ringkasan(wb, cat, brand_by_id)
    sheet_matrix(wb, cat, domain_title, brand_by_id)
    sheet_brand_map(wb, cat)
    sheet_pivot(wb, cat)
    sheet_gaps(wb, cat)
    sheet_method(wb, cat)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)
    print(f"{len(cat['modules'])} modules, {len(cat['gaps'])} gaps → {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
