# -*- coding: utf-8 -*-
"""Bank import template — declares how to parse a bank's CSV statement.

Derived from arkaaim's ``era.bank.import.template`` but simplified to the
column-index model requested in the spec (1-based, stored as integers).
Header-name resolution is still supported when ``has_header`` is True
and the index resolves to a header cell that exists.
"""

from __future__ import annotations

import base64
import csv
import io
import logging
import re
from datetime import date as date_cls
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from odoo import fields, models

_logger = logging.getLogger(__name__)


class BankImportTemplate(models.Model):
    _name = "custom.bank.import.template"
    _description = "Bank Import Template"
    _order = "sequence, name"

    name = fields.Char(required=True, index=True)
    sequence = fields.Integer(default=10)
    code = fields.Char(required=True, index=True, help="Stable identifier, e.g. 'bca_csv'.")
    bank_id = fields.Many2one("res.bank", string="Bank")
    company_id = fields.Many2one(
        "res.company",
        string="Company",
        default=lambda s: s.env.company,
        required=True,
    )
    active = fields.Boolean(default=True)

    encoding = fields.Selection(
        [("utf-8", "UTF-8"), ("latin-1", "Latin-1")],
        default="utf-8",
        required=True,
    )
    delimiter = fields.Char(default=",", size=1, required=True)
    has_header = fields.Boolean(default=True, help="Skip first row of file (column headers).")
    date_format = fields.Char(
        default="%d/%m/%Y",
        required=True,
        help="Python strptime format. BCA: %d/%m/%Y, Mandiri: %d-%m-%Y.",
    )

    # 1-based column indices. -1 means "not used".
    date_column_index = fields.Integer(default=1, required=True)
    ref_column_index = fields.Integer(default=-1)
    partner_column_index = fields.Integer(default=-1)
    amount_credit_column_index = fields.Integer(default=-1)
    amount_debit_column_index = fields.Integer(default=-1)
    balance_column_index = fields.Integer(default=-1)
    signed_amount_column_index = fields.Integer(
        default=-1,
        help="If set, this column holds a signed amount and overrides amount_credit/amount_debit.",
    )

    sample_file = fields.Binary(string="Sample File", attachment=True)
    sample_filename = fields.Char()

    decimal_separator = fields.Char(default=".", size=1, required=True)
    thousand_separator = fields.Char(default=",", size=1)

    _sql_constraints = [
        ("code_uniq", "unique(code, company_id)", "Template code must be unique per company."),
    ]

    # ------------------------------------------------------------------
    # Parsing
    # ------------------------------------------------------------------

    def _parse_amount(self, raw: Any) -> Decimal:
        if raw is None:
            return Decimal("0")
        s = str(raw).strip()
        if not s:
            return Decimal("0")
        if self.thousand_separator:
            s = s.replace(self.thousand_separator, "")
        if self.decimal_separator and self.decimal_separator != ".":
            s = s.replace(self.decimal_separator, ".")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return Decimal(s)
        except InvalidOperation:
            return Decimal("0")

    def _parse_date(self, raw: Any):
        if not raw:
            return False
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date_cls):
            return raw
        s = str(raw).strip()
        try:
            return datetime.strptime(s, self.date_format).date()
        except ValueError:
            return False

    @staticmethod
    def _safe_cell(row, idx_1based: int) -> Optional[str]:
        if idx_1based is None or idx_1based <= 0:
            return None
        i = idx_1based - 1
        if 0 <= i < len(row):
            return row[i]
        return None

    def _read_csv(self, file_bytes: bytes) -> list[list[str]]:
        text = file_bytes.decode(self.encoding or "utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text), delimiter=self.delimiter or ",")
        return list(reader)

    @staticmethod
    def _read_xls(file_bytes: bytes) -> list[list[Any]]:
        """Legacy BIFF .xls → list of rows. Date cells become datetime objects."""
        import xlrd

        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheet = wb.sheet_by_index(0)
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            row: list[Any] = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, wb.datemode)
                row.append(value)
            rows.append(row)
        return rows

    @staticmethod
    def _read_xlsx(file_bytes: bytes) -> list[list[Any]]:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = wb.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def _read_rows(self, file_bytes: bytes) -> list[list[Any]]:
        """Sniff the container format (OLE2 .xls / zip .xlsx / plain CSV) by
        magic bytes — bank portals lie about extensions, so the filename is
        deliberately not consulted."""
        if file_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            return self._read_xls(file_bytes)
        if file_bytes[:2] == b"PK":
            return self._read_xlsx(file_bytes)
        return self._read_csv(file_bytes)

    def parse_csv(self, file_b64: str) -> dict:
        """Parse a base64 statement file (CSV, .xls or .xlsx). Returns dict with keys:
        - lines: list of {date, ref, partner_hint, amount(Decimal), balance}
        - errors: list of (row_number, error_string)
        - total_rows: int
        """
        self.ensure_one()
        file_bytes = base64.b64decode(file_b64)
        rows = self._read_rows(file_bytes)
        if self._is_bca_corp(rows):
            return self._parse_bca_corp(rows)
        if self._is_brisim(rows):
            return self._parse_brisim(rows)
        if self._is_trx_inquiry(rows):
            return self._parse_trx_inquiry(rows)
        if self._is_acc_statement(rows):
            return self._parse_acc_statement(rows)
        if self.has_header and rows:
            rows = rows[1:]
        lines: list[dict] = []
        errors: list[tuple[int, str]] = []
        for n, row in enumerate(rows, start=2 if self.has_header else 1):
            if not row or all((c is None or str(c).strip() == "") for c in row):
                continue
            raw_date = self._safe_cell(row, self.date_column_index)
            d = self._parse_date(raw_date)
            if not d:
                errors.append((n, f"Bad/missing date: {raw_date!r}"))
                continue
            ref = self._safe_cell(row, self.ref_column_index) or ""
            partner_hint = self._safe_cell(row, self.partner_column_index) or ""
            balance_raw = self._safe_cell(row, self.balance_column_index)
            balance = self._parse_amount(balance_raw) if balance_raw else None
            if self.signed_amount_column_index and self.signed_amount_column_index > 0:
                amount = self._parse_amount(self._safe_cell(row, self.signed_amount_column_index))
            else:
                credit = self._parse_amount(self._safe_cell(row, self.amount_credit_column_index))
                debit = self._parse_amount(self._safe_cell(row, self.amount_debit_column_index))
                amount = credit - debit
            if amount == Decimal("0"):
                continue
            lines.append(
                {
                    "date": d,
                    "ref": str(ref).strip(),
                    "partner_hint": str(partner_hint).strip(),
                    "amount": amount,
                    "balance": balance,
                }
            )
        return {
            "lines": lines,
            "errors": errors,
            "total_rows": len(rows),
        }

    # ------------------------------------------------------------------
    # BCA corporate ("KlikBCA Bisnis" / CorpAcctTrxn) CSV
    # ------------------------------------------------------------------
    # This export cannot be described by the generic column-index model:
    #  * a metadata preamble precedes the column header;
    #  * dates are either ``DD/MM`` with NO year (year lives in the "Periode :"
    #    line) or full ``DD/MM/YYYY`` — both variants exist in the wild;
    #  * amount + sign are fused in one "Jumlah" cell, e.g. ``30,000.00 DB`` /
    #    ``10,930,941.82 CR`` (CR = money in / positive, DB = money out / negative);
    #  * ``Saldo Awal / Mutasi / Saldo Akhir`` footer rows trail the data.
    # It is auto-detected (independent of template column config) and handled here.

    _BCA_CORP_HEADER = ["Tanggal Transaksi", "Keterangan", "Cabang", "Jumlah", "Saldo"]
    _BCA_CORP_FOOTER = ("Saldo Awal", "Mutasi Debet", "Mutasi Kredit", "Saldo Akhir")

    @staticmethod
    def _bca_corp_amount(raw: Any) -> Decimal:
        # BCA corporate amounts are always ``1,234,567.89`` (comma thousands, dot
        # decimal) regardless of the template's configured separators, because this
        # handler is auto-detected and may run under any selected template.
        s = str(raw or "").strip().replace(",", "")
        if not s:
            return Decimal("0")
        try:
            return Decimal(s)
        except InvalidOperation:
            return Decimal("0")

    def _is_bca_corp(self, rows) -> bool:
        if not rows:
            return False
        first = (self._safe_cell(rows[0], 1) or "").strip()
        if first.startswith("Informasi Rekening"):
            return True
        for row in rows:
            if [(str(c).strip()) for c in row[:5]] == self._BCA_CORP_HEADER:
                return True
        return False

    def _parse_bca_corp(self, rows) -> dict:
        lines: list[dict] = []
        errors: list[tuple[int, str]] = []

        # Resolve the statement's year(s) from the "Periode :" preamble and find the
        # column-header row; data starts on the row after it.
        from_month = from_year = to_year = None
        header_idx = None
        period_re = re.compile(r"Periode\s*:\s*(\d{2})/(\d{2})/(\d{4})\s*-\s*\d{2}/\d{2}/(\d{4})")
        for i, row in enumerate(rows):
            joined = " ".join(str(c) for c in row)
            m = period_re.search(joined)
            if m:
                from_month = int(m.group(2))
                from_year = int(m.group(3))
                to_year = int(m.group(4))
            if [(str(c).strip()) for c in row[:5]] == self._BCA_CORP_HEADER:
                header_idx = i
                break

        start = (header_idx + 1) if header_idx is not None else 0
        date_re = re.compile(r"^(\d{2})/(\d{2})(?:/(\d{4}))?$")

        for n, row in enumerate(rows[start:], start=start + 1):
            if not row or all(str(c).strip() == "" for c in row):
                continue
            c0 = (self._safe_cell(row, 1) or "").strip()
            if c0.startswith(self._BCA_CORP_FOOTER):
                break
            dm = date_re.match(c0)
            if not dm:
                errors.append((n, f"Bad/missing date: {c0!r}"))
                continue
            day, month = int(dm.group(1)), int(dm.group(2))
            if dm.group(3):
                year = int(dm.group(3))
            else:
                # Year rollover: a period spanning Dec->Jan uses the later year for
                # the months that wrapped past the start month. Inert for a
                # single-month file.
                if from_year is None:
                    errors.append((n, "Missing 'Periode' line; cannot resolve year"))
                    continue
                year = to_year if (to_year != from_year and month < from_month) else from_year
            try:
                d = date_cls(year, month, day)
            except ValueError as e:
                errors.append((n, f"Invalid date {c0!r}: {e}"))
                continue

            raw_amount = (self._safe_cell(row, 4) or "").strip()
            am = re.match(r"^(.*?)\s*(CR|DB)$", raw_amount, re.IGNORECASE)
            if am:
                num, sign = am.group(1), (1 if am.group(2).upper() == "CR" else -1)
            else:
                num, sign = raw_amount, 1
            amount = self._bca_corp_amount(num) * sign
            if amount == Decimal("0"):
                continue

            keterangan = re.sub(r"\s+", " ", str(self._safe_cell(row, 2) or "")).strip()
            balance_raw = self._safe_cell(row, 5)
            balance = self._bca_corp_amount(balance_raw) if balance_raw else None

            lines.append(
                {
                    "date": d,
                    "ref": keterangan,
                    "partner_hint": "",
                    "amount": amount,
                    "balance": balance,
                }
            )

        return {
            "lines": lines,
            "errors": errors,
            "total_rows": max(0, len(rows) - start),
        }

    # ------------------------------------------------------------------
    # BRI "BRISIM" internet-banking XLSX (journal IBRI)
    # ------------------------------------------------------------------
    # Header row: Tanggal | Uraian | Teller | Debet | Kredit | Saldo — but the
    # export is column-shifted garbage: only col0 (timestamp of THIS row) and
    # col1 (Uraian) are trustworthy; col2+ hold fragments of the NEXT row. The
    # real numbers live INSIDE the Uraian text as a "debet kredit saldo" triple
    # (US separators), e.g.
    #   "OffUs 1 260701 001999632292 LEVIS KELAPA BRIMTXDT 0.00 4,323,118.00
    #    299,942,072.00 AMT:4.368.570,00MDR:59.796,00"
    # A single Uraian cell can even embed a SECOND full transaction (seen with
    # "DEBET BY CEK ..."), each introduced by its own dd/mm/yy hh:mm:ss stamp.

    _BRISIM_HEADER = ["Tanggal", "Uraian", "Teller", "Debet", "Kredit", "Saldo"]
    _BRISIM_DATE_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{2})\s+\d{2}:\d{2}:\d{2}$")
    _BRISIM_TRIPLE_RE = re.compile(
        r"(?P<desc>.*?)\s"
        r"(?P<debet>\d[\d,]*\.\d{2})\s+"
        r"(?P<kredit>\d[\d,]*\.\d{2})\s+"
        r"(?P<saldo>\d[\d,]*\.\d{2})"
    )
    _BRISIM_EMBEDDED_RE = re.compile(r"\s(\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}:\d{2})\s")

    def _is_brisim(self, rows) -> bool:
        for row in rows[:5]:
            if [str(c or "").strip() for c in row[:6]] == self._BRISIM_HEADER:
                return True
        return False

    def _parse_brisim(self, rows) -> dict:
        lines: list[dict] = []
        errors: list[tuple[int, str]] = []
        header_idx = next(
            i for i, row in enumerate(rows) if [str(c or "").strip() for c in row[:6]] == self._BRISIM_HEADER
        )
        data_rows = rows[header_idx + 1 :]

        def add_segment(n, d, segment):
            m = self._BRISIM_TRIPLE_RE.search(segment)
            if not m:
                errors.append((n, f"No debet/kredit/saldo triple in Uraian: {segment[:80]!r}"))
                return
            debet = self._bca_corp_amount(m.group("debet"))
            kredit = self._bca_corp_amount(m.group("kredit"))
            amount = kredit - debet
            if amount == Decimal("0"):
                return
            desc = re.sub(r"\s+", " ", m.group("desc")).strip()
            lines.append(
                {
                    "date": d,
                    "ref": desc,
                    "partner_hint": "",
                    "amount": amount,
                    "balance": self._bca_corp_amount(m.group("saldo")),
                }
            )

        for n, row in enumerate(data_rows, start=header_idx + 2):
            c0 = str(self._safe_cell(row, 1) or "").strip()
            uraian = str(self._safe_cell(row, 2) or "").strip()
            if not c0 and not uraian:
                continue
            dm = self._BRISIM_DATE_RE.match(c0)
            if not dm:
                errors.append((n, f"Bad/missing date: {c0!r}"))
                continue
            day, month, yy = int(dm.group(1)), int(dm.group(2)), int(dm.group(3))
            try:
                d = date_cls(2000 + yy, month, day)
            except ValueError as e:
                errors.append((n, f"Invalid date {c0!r}: {e}"))
                continue
            # Split embedded extra transactions (each has its own timestamp).
            parts = self._BRISIM_EMBEDDED_RE.split(uraian)
            add_segment(n, d, parts[0])
            for stamp, segment in zip(parts[1::2], parts[2::2]):
                sd, sm, sy = int(stamp[:2]), int(stamp[3:5]), int(stamp[6:8])
                try:
                    seg_date = date_cls(2000 + sy, sm, sd)
                except ValueError:
                    seg_date = d
                add_segment(n, seg_date, segment)

        return {"lines": lines, "errors": errors, "total_rows": len(data_rows)}

    # ------------------------------------------------------------------
    # BNI "TRANSACTION INQUIRY" XLS (BNIDirect corporate download)
    # ------------------------------------------------------------------
    # Sparse grid with a metadata block, then a header row containing
    # 'Post Date' and 'Db/Cr'. Columns are located by header text (their
    # physical indexes vary with the merged-cell layout). Post Date cells are
    # native Excel datetimes; Amount is unsigned with sign in Db/Cr (C/D).

    def _find_header_row(self, rows, required: tuple[str, ...]):
        """Return (row_index, {header_label: col_index}) for the first row
        containing every label in ``required``. Labels are matched on the
        stripped cell text."""
        for i, row in enumerate(rows):
            labels = {str(c or "").strip(): j for j, c in enumerate(row) if str(c or "").strip()}
            if all(r in labels for r in required):
                return i, labels
        return None, {}

    def _is_trx_inquiry(self, rows) -> bool:
        idx, _labels = self._find_header_row(rows, ("Post Date", "Db/Cr", "Amount"))
        return idx is not None

    def _parse_trx_inquiry(self, rows) -> dict:
        lines: list[dict] = []
        errors: list[tuple[int, str]] = []
        header_idx, cols = self._find_header_row(rows, ("Post Date", "Db/Cr", "Amount"))
        data_rows = rows[header_idx + 1 :]
        for n, row in enumerate(data_rows, start=header_idx + 2):
            raw_date = self._safe_cell(row, cols["Post Date"] + 1)
            if raw_date is None or str(raw_date).strip() == "":
                continue
            if not isinstance(raw_date, (datetime, date_cls)):
                errors.append((n, f"Bad/missing date: {raw_date!r}"))
                continue
            d = raw_date.date() if isinstance(raw_date, datetime) else raw_date
            raw_amount = self._safe_cell(row, cols["Amount"] + 1)
            dbcr = str(self._safe_cell(row, cols["Db/Cr"] + 1) or "").strip().upper()
            sign = -1 if dbcr == "D" else 1
            try:
                amount = Decimal(str(raw_amount)) * sign
            except InvalidOperation:
                errors.append((n, f"Bad amount: {raw_amount!r}"))
                continue
            if amount == Decimal("0"):
                continue
            desc = str(self._safe_cell(row, cols.get("Description", -1) + 1) or "")
            journal_no = str(self._safe_cell(row, cols.get("Journal No.", -1) + 1) or "")
            raw_balance = self._safe_cell(row, cols.get("Balance", -1) + 1)
            try:
                balance = Decimal(str(raw_balance)) if raw_balance not in (None, "") else None
            except InvalidOperation:
                balance = None
            lines.append(
                {
                    "date": d,
                    "ref": re.sub(r"\s+", " ", desc).strip() or journal_no,
                    "partner_hint": journal_no,
                    "amount": amount,
                    "balance": balance,
                }
            )
        return {"lines": lines, "errors": errors, "total_rows": len(data_rows)}

    # ------------------------------------------------------------------
    # Mandiri "Acc_Statement" XLS (corporate account statement)
    # ------------------------------------------------------------------
    # Metadata block (Account No / Period / Currency / Branch / Opening
    # Balance), then a header row with 'Posting Date', 'Remark', 'Reference
    # No', 'Debit', 'Credit', 'Balance'. Every value is a string; dates are
    # 'dd/mm/YYYY HH:MM:SS'; amounts use US separators. A 'No of Debit'
    # summary block trails the data.

    _ACC_STMT_DATE_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})\s+\d{2}:\d{2}:\d{2}$")

    def _is_acc_statement(self, rows) -> bool:
        idx, _labels = self._find_header_row(rows, ("Posting Date", "Remark", "Debit", "Credit"))
        return idx is not None

    def _parse_acc_statement(self, rows) -> dict:
        lines: list[dict] = []
        errors: list[tuple[int, str]] = []
        header_idx, cols = self._find_header_row(rows, ("Posting Date", "Remark", "Debit", "Credit"))
        data_rows = rows[header_idx + 1 :]
        for n, row in enumerate(data_rows, start=header_idx + 2):
            c_date = str(self._safe_cell(row, cols["Posting Date"] + 1) or "").strip()
            joined = " ".join(str(c or "").strip() for c in row)
            if not c_date:
                if re.search(r"No of (Debit|Credit)|Closing Balance", joined):
                    break
                continue
            dm = self._ACC_STMT_DATE_RE.match(c_date)
            if not dm:
                errors.append((n, f"Bad/missing date: {c_date!r}"))
                continue
            try:
                d = date_cls(int(dm.group(3)), int(dm.group(2)), int(dm.group(1)))
            except ValueError as e:
                errors.append((n, f"Invalid date {c_date!r}: {e}"))
                continue
            debit = self._bca_corp_amount(self._safe_cell(row, cols["Debit"] + 1))
            credit = self._bca_corp_amount(self._safe_cell(row, cols["Credit"] + 1))
            amount = credit - debit
            if amount == Decimal("0"):
                continue
            remark = re.sub(r"\s+", " ", str(self._safe_cell(row, cols["Remark"] + 1) or "")).strip()
            ref_no = str(self._safe_cell(row, cols.get("Reference No", -1) + 1) or "").strip()
            raw_balance = self._safe_cell(row, cols.get("Balance", -1) + 1)
            balance = self._bca_corp_amount(raw_balance) if raw_balance else None
            lines.append(
                {
                    "date": d,
                    "ref": remark or (ref_no if ref_no != "-" else ""),
                    "partner_hint": "",
                    "amount": amount,
                    "balance": balance,
                }
            )
        return {"lines": lines, "errors": errors, "total_rows": len(data_rows)}
