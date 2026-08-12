#!/usr/bin/env python3
"""Turn the returned mapping workbook back into the CSV ``assign_roles.py`` reads.

    python3 scripts/security/xlsx_to_assign_csv.py pemetaan_role_<db>.xlsx

Writes ``<name>.csv`` beside it. Plain Python — no Odoo, no database.

Only rows where somebody filled in a role are carried over; the rest are counted
and reported, so it is obvious how much of the sheet came back answered.

Multiple roles (or units) in one cell are separated by ``|``. The workbook's
dropdowns should make an unknown code impossible, but a cell can still be typed
over, so ``assign_roles.py`` validates every code again before writing anything.
"""

import csv
import os
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

COLUMNS = {"login": 1, "name": 2, "current": 3, "roles": 4, "units": 5, "note": 6}


def main(argv):
    if len(argv) != 2:
        sys.exit(__doc__.strip())
    src = argv[1]
    dst = os.path.splitext(src)[0] + ".csv"

    book = openpyxl.load_workbook(src, data_only=True)
    if "Pemetaan User" not in book.sheetnames:
        sys.exit("sheet 'Pemetaan User' not found — is this the right workbook?")
    sheet = book["Pemetaan User"]

    filled = blank = 0
    with open(dst, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["login", "role_codes", "ou_codes"])
        for row in sheet.iter_rows(min_row=2, values_only=True):
            login = (row[COLUMNS["login"] - 1] or "").strip()
            roles = (row[COLUMNS["roles"] - 1] or "").strip()
            units = (row[COLUMNS["units"] - 1] or "").strip()
            if not login:
                continue
            if not roles:
                blank += 1
                continue
            writer.writerow([login, roles, units])
            filled += 1

    print("%d row(s) with a role → %s" % (filled, dst))
    if blank:
        print("%d user(s) left blank — they keep exactly the rights they have now." % blank)
    print("Next: dry-run it.")
    print("  docker exec -i -e CSV=%s odoo19-platform-odoo-mgmt \\" % dst)
    print("      odoo shell -d <db> --no-http < scripts/security/assign_roles.py")


if __name__ == "__main__":
    main(sys.argv)
