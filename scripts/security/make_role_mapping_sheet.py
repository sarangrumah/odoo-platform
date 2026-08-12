"""Build the fill-in workbook that decides who gets which role.

    docker exec -i odoo19-platform-odoo-mgmt odoo shell \
        -d prd_levis_begbal --no-http < scripts/security/make_role_mapping_sheet.py

Read-only. Writes ``/tmp/pemetaan_role_<db>.xlsx`` — a workbook in Indonesian for
the customer to fill in, not a technical export:

* **Petunjuk** — what to do, in four steps.
* **Pemetaan User** — one row per active user, with their current rights
  summarised in plain language and dropdowns for Role and Operating Unit.
* **Daftar Role** — what each role can and deliberately cannot do.
* **Daftar Operating Unit** — the codes, so nobody types a store name by hand.

The Role and Operating Unit columns are real Excel dropdowns, so a returned file
cannot contain a code that does not exist. Feed the result back through
``xlsx_to_assign_csv.py`` and then ``assign_roles.py`` (dry-run first).

Why a workbook and not a database query: on a tenant that has never had roles,
who *should* be what is not in the system. Every user here looks identical — all
84 hold Accounting, POS, Stock and Purchase Manager alike. Only the customer
knows who actually does what.
"""

import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_logger = logging.getLogger("role_sheet")
logging.basicConfig(level=logging.INFO)

env = env  # noqa: F821 — odoo shell global

# Business-facing Indonesian labels. The role records carry English names for
# the Odoo UI; a sheet going to the customer should not mix languages.
ROLE_ID = {
    "hq_acc_manager": (
        "Finance & Accounting Manager",
        "Hak akuntansi penuh: CoA, lock date, persetujuan dokumen keuangan.",
    ),
    "hq_acc_supervisor": (
        "Accounting Supervisor",
        "Memeriksa dan memposting pekerjaan staff; menjalankan laporan standar.",
    ),
    "hq_acc_staff_ap": (
        "Accounting Staff — AP",
        "Input tagihan vendor dan permintaan pembayaran. Tidak bisa posting ke periode terkunci.",
    ),
    "hq_acc_staff_ar": ("Accounting Staff — AR", "Invoice pelanggan, penerimaan, tindak lanjut piutang."),
    "hq_tax_officer": ("Tax Officer", "e-Faktur, Coretax, PPh/bupot."),
    "hq_treasury": ("Treasury / Kasir HO", "Jurnal bank & kas, eksekusi pembayaran, petty cash."),
    "hq_auditor": (
        "Internal Auditor (baca saja)",
        "Membaca buku besar dan laporan. Tidak membuat atau mengubah apa pun.",
    ),
    "hq_purchase_manager": ("Purchasing Manager", "Menyetujui PO, mengelola harga vendor."),
    "hq_purchase_staff": ("Purchasing Staff", "Membuat PO dan menindaklanjuti vendor."),
    "hq_sales_manager": ("Sales Manager", "Hak penjualan penuh termasuk harga dan diskon."),
    "hq_sales_admin": ("Sales Admin / Merchandising", "Sales order dan pemeliharaan master produk."),
    "hq_inventory_manager": ("Inventory Manager", "Konfigurasi gudang, valuasi, penyesuaian stok."),
    "hq_it_admin": (
        "IT / System Administrator",
        "Administrasi teknis. SENGAJA dipisah — jangan diberikan hanya karena jabatan tinggi.",
    ),
    "store_manager": ("Store Manager", "Menjalankan satu toko: POS, stok, laporan toko, persetujuan lapis pertama."),
    "store_supervisor": ("Store Supervisor", "Buka/tutup sesi POS, mengawasi stock count."),
    "store_cashier": ("Store Staff / Kasir POS", "Mengoperasikan POS saja. Tanpa akuntansi, tanpa konfigurasi stok."),
    "store_stock_keeper": ("Stock Keeper", "Terima dan kirim barang, hitung stok. Tanpa akuntansi."),
    "area_manager": (
        "Area Manager",
        "Mengawasi beberapa toko. Isi kolom Operating Unit dengan unit AREA, bukan satu per satu toko.",
    ),
}

HEAD = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E2F3")
WARN = PatternFill("solid", fgColor="FCE4D6")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

wb = Workbook()

# ---------------------------------------------------------------- Petunjuk ---
ws = wb.active
ws.title = "Petunjuk"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 110
rows = [
    ("Pemetaan Hak Akses Pengguna — %s" % env.cr.dbname, "title"),
    ("", ""),
    ("Tujuan", "h"),
    (
        "Saat ini semua pengguna di sistem ini memegang hak yang sama persis — Accounting Manager, "
        "POS Manager, Stock Manager dan Purchasing Manager sekaligus. Artinya tidak ada yang membedakan "
        "staff dari manajer. Lembar ini untuk menetapkan siapa sebenarnya mengerjakan apa.",
        "p",
    ),
    ("", ""),
    ("Cara mengisi", "h"),
    ("1. Buka sheet 'Pemetaan User'. Setiap baris adalah satu pengguna aktif.", "p"),
    ("2. Kolom ROLE: pilih dari dropdown. Lihat sheet 'Daftar Role' untuk arti tiap peran.", "p"),
    (
        "3. Kolom OPERATING UNIT: isi HANYA untuk orang yang bekerja di satu toko/area tertentu. "
        "Kosongkan untuk orang kantor pusat — mereka tetap melihat semua data.",
        "p",
    ),
    (
        "4. Kirim kembali file ini. Perubahan dijalankan uji-coba dulu, hasilnya kami tunjukkan "
        "per orang sebelum benar-benar diterapkan.",
        "p",
    ),
    ("", ""),
    ("Yang perlu diperhatikan", "h"),
    (
        "• Mengisi OPERATING UNIT membuat orang itu HANYA melihat data unit tersebut. Ini pembatasan "
        "nyata — mulai dari beberapa orang dulu, jangan sekaligus.",
        "warn",
    ),
    (
        "• Peran 'IT / System Administrator' memberi akses Pengaturan sistem. Berikan hanya kepada "
        "orang IT, bukan kepada manajer bisnis.",
        "warn",
    ),
    (
        "• Satu orang boleh memegang lebih dari satu peran. Pisahkan dengan tanda | "
        "(contoh: hq_acc_staff_ap|hq_treasury).",
        "p",
    ),
    ("• Hak yang diberikan manual di luar peran tidak akan dicabut oleh sistem.", "p"),
]
r = 1
for text, kind in rows:
    c = ws.cell(row=r, column=2, value=text)
    c.alignment = WRAP
    if kind == "title":
        c.font = Font(bold=True, size=14, color="1F3864")
    elif kind == "h":
        c.font = BOLD
        c.fill = SUB
    elif kind == "warn":
        c.fill = WARN
    ws.row_dimensions[r].height = 30 if kind in ("p", "warn") else 18
    r += 1

# ------------------------------------------------------------ Daftar Role ---
roles = env["custom.security.role"].search([], order="role_domain,level,name")
ws_r = wb.create_sheet("Daftar Role")
headers = ["Kode", "Peran", "Level", "Untuk", "Apa yang bisa dikerjakan"]
for i, h in enumerate(headers, start=1):
    c = ws_r.cell(row=1, column=i, value=h)
    c.fill, c.font = HEAD, WHITE
for i, w in enumerate([22, 34, 14, 16, 80], start=1):
    ws_r.column_dimensions[get_column_letter(i)].width = w

scope_id = {"head_office": "Kantor Pusat", "retail": "Toko", "both": "Keduanya"}
level_id = {
    "manager": "Manajer",
    "supervisor": "Supervisor",
    "staff": "Staff",
    "operator": "Operator",
    "readonly": "Baca saja",
}
for row, role in enumerate(roles, start=2):
    label, desc = ROLE_ID.get(role.code, (role.name, role.description or ""))
    for col, value in enumerate(
        [role.code, label, level_id.get(role.level, role.level), scope_id.get(role.scope, role.scope), desc], start=1
    ):
        c = ws_r.cell(row=row, column=col, value=value)
        c.alignment = WRAP
        c.border = THIN
    if role.code == "hq_it_admin":
        ws_r.cell(row=row, column=5).fill = WARN
ws_r.freeze_panes = "A2"

# ------------------------------------------------- Daftar Operating Unit ---
units = env["operating.unit"].search([], order="ou_type,complete_name")
ws_u = wb.create_sheet("Daftar Operating Unit")
for i, h in enumerate(["Kode", "Nama", "Jenis"], start=1):
    c = ws_u.cell(row=1, column=i, value=h)
    c.fill, c.font = HEAD, WHITE
for i, w in enumerate([16, 52, 18], start=1):
    ws_u.column_dimensions[get_column_letter(i)].width = w
type_id = {"company": "Kantor Pusat", "area": "Area", "store": "Toko", "other": "Lainnya"}
for row, unit in enumerate(units, start=2):
    for col, value in enumerate([unit.code, unit.name, type_id.get(unit.ou_type, unit.ou_type)], start=1):
        c = ws_u.cell(row=row, column=col, value=value)
        c.border = THIN
ws_u.freeze_panes = "A2"

# ---------------------------------------------------------- Pemetaan User ---
FLAGGED = [
    ("base.group_system", "Pengaturan sistem"),
    ("account.group_account_manager", "Accounting Manager"),
    ("point_of_sale.group_pos_manager", "POS Manager"),
    ("stock.group_stock_manager", "Stock Manager"),
    ("purchase.group_purchase_manager", "Purchasing Manager"),
]
flagged = [(env.ref(x, raise_if_not_found=False), label) for x, label in FLAGGED]
flagged = [(g, label) for g, label in flagged if g]

users = env["res.users"].search([("active", "=", True)], order="login")
ws_m = wb.create_sheet("Pemetaan User", 1)
headers = ["Login", "Nama", "Hak sekarang", "ROLE (isi di sini)", "OPERATING UNIT (isi bila perlu)", "Catatan"]
for i, h in enumerate(headers, start=1):
    c = ws_m.cell(row=1, column=i, value=h)
    c.fill, c.font, c.alignment = HEAD, WHITE, WRAP
    if i in (4, 5):
        c.fill = PatternFill("solid", fgColor="C55A11")
for i, w in enumerate([34, 30, 46, 30, 32, 34], start=1):
    ws_m.column_dimensions[get_column_letter(i)].width = w

for row, user in enumerate(users, start=2):
    held = [label for g, label in flagged if g in user.all_group_ids]
    for col, value in enumerate([user.login, user.partner_id.name or "", ", ".join(held) or "—", "", "", ""], start=1):
        c = ws_m.cell(row=row, column=col, value=value)
        c.alignment = WRAP
        c.border = THIN
ws_m.freeze_panes = "D2"

last = len(users) + 1
dv_role = DataValidation(type="list", formula1="'Daftar Role'!$A$2:$A$%d" % (len(roles) + 1), allow_blank=True)
dv_role.error = "Pilih kode peran dari daftar."
dv_unit = DataValidation(
    type="list", formula1="'Daftar Operating Unit'!$A$2:$A$%d" % (len(units) + 1), allow_blank=True
)
dv_unit.error = "Pilih kode Operating Unit dari daftar."
ws_m.add_data_validation(dv_role)
ws_m.add_data_validation(dv_unit)
dv_role.add("D2:D%d" % last)
dv_unit.add("E2:E%d" % last)

path = "/tmp/pemetaan_role_%s.xlsx" % env.cr.dbname
wb.save(path)
_logger.info("%d user, %d role, %d unit → %s", len(users), len(roles), len(units), path)
