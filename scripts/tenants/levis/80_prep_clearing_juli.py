#!/usr/bin/env python3
"""Build the July-2026 clearing plan for prd_levis_begbal from EBR's workbook.

Runs on the HOST (plain python3 + openpyxl). Reads only; the database is read
through ``docker exec ... psql`` so the ORM is never touched here. Output:

  * a JSON plan consumed by ``81_clearing_juli.py`` (odoo shell)
  * a validation workbook for Accounting/FICO

    python3 scripts/tenants/levis/80_prep_clearing_juli.py \
        --ebr '/path/07. REPORT EBR PERIODE JULI 2026 (2).xlsx' \
        --bca-csv '/path/MUTASI REKENING BCA EBR 01 - 31 JULI 2026.csv' \
        --db prd_levis_begbal \
        --json /tmp/clearing_juli.json \
        --xlsx /srv/sftp-share/files/Draft_Clearing_Juli2026.xlsx

The plan has five blocks:

  A  settlement of July POS receivables (per settlement date, per store):
     Dr 1103000002 Bank Suspense + Dr 7104000001 MDR / Cr 1106000101..110
  B  collection of the June AR left in 1106000001 (rows the workbook tags
     "COLLECTION AR JUNE"): Dr Bank Suspense + Dr MDR / Cr 1106000001
  C  BCA ATS sweep 2685151268 -> 2687778282 and bank charges:
     Dr 1103019320 / Dr 7299012000 / Cr 1103000002
  D  the four 23-Jul RIREC lines of PASKAL BANDUNG that carry no OU analytic
  S  the bank statement lines July still misses (BCA 21-31, BRI 25-31), so the
     suspense account blocks A-C clear is actually funded

Grain note: the credit side comes from `COMPILE SALES` (transaction level, ties
to Odoo per store x trans-date x tender), the bank side from the `MUTASI *`
sheets. Per store the two agree for the month; per single day they differ
because a settlement lands D or D+1. MDR is therefore taken per store for the
whole month and spread over that store's settlement dates pro-rata to gross,
which keeps every entry balanced and the monthly MDR per store exact.
"""

import argparse
import collections
import csv
import json
import os
import re
import subprocess
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ACC_SUSPENSE = "1103000002"
ACC_MDR = "7104000001"
ACC_AR = "1106000001"
ACC_BCA_MAIN = "1103019320"
ACC_BANK_CHARGE = "7299012000"

# X70D tender type -> POS receivable account
TENDER_ACC = {
    "CASH": "1106000101",
    "OFFLINE_DOMESTIC_CARD": "1106000102",
    "OFFLINE_VISA": "1106000103",
    "OFFLINE_MASTERCARD": "1106000104",
    "OFFLINE_OTHER_CREDITCARD": "1106000105",
    "OFFLINE_CREDIT_CARD": "1106000106",
    "OFFLINE_JCB": "1106000107",
    "OFFLINE_BRI_CREDIT_CARD": "1106000108",
    "OFFLINE_AMEX": "1106000109",
    "OFFLINE_OVO": "1106000110",
}
# spellings EBR uses that mean the same tender
TENDER_ALIAS = {
    "OFFLINE_DOMESTIC CARD": "OFFLINE_DOMESTIC_CARD",
    "ONLINE_DOMESTIC_CARD": "OFFLINE_DOMESTIC_CARD",
    "DEBIT": "OFFLINE_DOMESTIC_CARD",
    "OFFLINE_OTHER_CARD": "OFFLINE_OTHER_CREDITCARD",
    "OFFLINE_OTHER_CREDIT_CARD": "OFFLINE_OTHER_CREDITCARD",
    "OFFLINE_BRI_CREDITCARD": "OFFLINE_BRI_CREDIT_CARD",
    "OFFLINE_CREDITCARD": "OFFLINE_CREDIT_CARD",
}

SHEET_HEADER_ROW = {  # 0-based row holding the column names
    "COMPILE SALES": 1,
    "MUTASI BCA": 6,
    "MUTASI BRI": 8,
}
LAST_STATEMENT_DAY = {"IBCA": "2026-07-20", "IBRI": "2026-07-24"}


def num(raw):
    s = str(raw or "").replace(",", "").strip()
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def rnd(x):
    return round(x + 0.0, 2)


def sheet_rows(wb, name):
    ws = wb[name]
    rows = [["" if c is None else c for c in r] for r in ws.iter_rows(values_only=True)]
    hdr = SHEET_HEADER_ROW[name]
    head = [str(c).strip() for c in rows[hdr]]
    return [dict(zip(head, r)) for r in rows[hdr + 1:] if any(str(c).strip() for c in r)]


def day(raw):
    return str(raw)[:10]


def bank_of_status(status):
    m = re.search(r"\((\w+)\)", str(status or ""))
    return m.group(1) if m else None


def psql(db, sql):
    out = subprocess.run(
        ["docker", "exec", "-i", "-e", "PGPASSWORD=" + os.environ["PGPASSWORD"],
         "odoo19-platform-postgres", "psql", "-U", "odoo", "-d", db, "-At", "-F", "\t", "-c", sql],
        capture_output=True, text=True, check=True,
    ).stdout
    return [line.split("\t") for line in out.splitlines() if line]


# ---------------------------------------------------------------- block A / B
def build_settlement(wb):
    """Return (blockA, blockB, mdr_month, diag)."""
    cs = sheet_rows(wb, "COMPILE SALES")
    # Credit side: store -> settlement date -> trans date -> amount.
    #
    # Deliberately NOT split per POS-receivable account here. EBR's tender-type
    # column disagrees with X70D on 35 store/day combinations (offsetting inside
    # the same day) and is empty for PASKAL 24-Jul, so crediting EBR's split
    # would over-credit one account and leave another open. 81_clearing_juli.py
    # resolves the split from the open RIREC lines of that store and day; the
    # workbook still shows EBR's own split for reference.
    credits = collections.defaultdict(lambda: collections.defaultdict(collections.Counter))
    ebr_split = collections.Counter()  # (store, trans date, account) -> amount
    gross = collections.defaultdict(collections.Counter)          # store -> date -> gross
    bank_of = collections.defaultdict(collections.Counter)        # store -> date -> bank
    unsettled = collections.Counter()
    unknown_tender = collections.Counter()
    for r in cs:
        amount = num(r["TENDER AMOUNT"])
        if not amount:
            continue
        store = str(r["STORE NAME"]).strip()
        bank = bank_of_status(r["STATUS"])
        if not bank:
            unsettled[store] += amount
            continue
        tender = str(r["TENDER TYPE"]).strip().upper()
        tender = TENDER_ALIAS.get(tender, tender)
        acc = TENDER_ACC.get(tender)
        if not acc:
            unknown_tender[(store, day(r["TRANS DATE"]), str(r["METODE PEMBAYARAN"]).strip())] += amount
        else:
            ebr_split[(store, day(r["TRANS DATE"]), acc)] += amount
        sdate = day(r["CASH RECEIVED DATE"])
        credits[store][sdate][day(r["TRANS DATE"])] += amount
        gross[store][sdate] += amount
        bank_of[store][sdate] += 0  # ensure key
        bank_of[store][(sdate, "bank")] = bank

    # bank side: MDR per store for the whole month, and the June-AR collections
    mdr_month = collections.Counter()
    block_b = []
    ats = []
    for sheet, bank, dcol in (("MUTASI BCA", "BCA", "Tanggal Transaksi"), ("MUTASI BRI", "BRI", "Tanggal")):
        for r in sheet_rows(wb, sheet):
            store = str(r.get("Store Name", "")).strip()
            note = str(r.get("NOTES", "")).strip().upper()
            flow = str(r.get("CASH IN/ATS", "")).strip().upper()
            amount = num(r.get("AMOUNT PAYMENT"))
            mdr = num(r.get("MDR"))
            if flow in ("ATS", "BANK ADMIN"):
                ats.append({
                    "date": day(r[dcol]), "bank": bank, "kind": flow,
                    "desc": str(r.get("Keterangan") or r.get("Uraian Transaksi") or "")[:120],
                    "amount": rnd(num(r.get("Jumlah")) or num(r.get("Debet")) or amount),
                })
                continue
            if "COLLECTION" in note:
                block_b.append({
                    "date": day(r[dcol]), "bank": bank, "store": store,
                    "gross": rnd(amount), "mdr": rnd(mdr), "cash_in": rnd(amount - mdr),
                })
                continue
            if store:
                mdr_month[store] += mdr

    # block A: one row per (settlement date, store); MDR spread pro-rata
    block_a = []
    for store, per_date in credits.items():
        total_gross = sum(gross[store].values())
        mdr_left = rnd(mdr_month.get(store, 0.0))
        dates = sorted(per_date)
        for i, sdate in enumerate(dates):
            g = gross[store][sdate]
            if i == len(dates) - 1:
                mdr = mdr_left
            else:
                mdr = rnd(mdr_month.get(store, 0.0) * g / total_gross) if total_gross else 0.0
                mdr_left = rnd(mdr_left - mdr)
            lines = [{"date": d, "amount": rnd(v)} for d, v in sorted(per_date[sdate].items())]
            block_a.append({
                "date": sdate, "store": store,
                "bank": bank_of[store].get((sdate, "bank"), ""),
                "gross": rnd(g), "mdr": rnd(mdr), "cash_in": rnd(g - mdr),
                "credits": lines,
            })
    block_a.sort(key=lambda r: (r["date"], r["store"]))
    block_b.sort(key=lambda r: (r["date"], r["store"]))
    diag = {
        "unsettled_per_store": {k: rnd(v) for k, v in unsettled.items()},
        "unsettled_total": rnd(sum(unsettled.values())),
        "unknown_tender": {" | ".join(map(str, k)): rnd(v) for k, v in unknown_tender.items()},
        "mdr_month": {k: rnd(v) for k, v in mdr_month.items()},
        "ebr_split": [(s, d, a, rnd(v)) for (s, d, a), v in sorted(ebr_split.items())],
    }
    return block_a, block_b, ats, diag


# ------------------------------------------------------------------ block S
def missing_bca_lines(path):
    """BCA corporate CSV rows after the last day already in Odoo."""
    out = []
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        rows = list(csv.reader(fh))
    hdr = None
    for i, r in enumerate(rows):
        if [c.strip() for c in r[:5]] == ["Tanggal Transaksi", "Keterangan", "Cabang", "Jumlah", "Saldo"]:
            hdr = i
            break
    if hdr is None:
        raise SystemExit("BCA CSV: header row not found")
    for r in rows[hdr + 1:]:
        if len(r) < 5 or not r[0].strip():
            continue
        raw = r[0].strip()
        if raw.startswith(("Saldo Awal", "Mutasi", "Saldo Akhir")):
            break
        try:
            d = datetime.strptime(raw, "%d/%m/%Y").date().isoformat()
        except ValueError:
            continue
        if d <= LAST_STATEMENT_DAY["IBCA"]:
            continue
        amt = r[3].strip()
        sign = -1 if amt.upper().endswith("DB") else 1
        value = num(re.sub(r"\s*(CR|DB)$", "", amt, flags=re.I))
        out.append({"journal": "IBCA", "date": d, "ref": r[1].strip()[:200],
                    "amount": rnd(sign * value)})
    return out


def missing_bri_lines(wb):
    """The whole July BRI statement, rebuilt from the workbook's MUTASI BRI.

    Two reasons this is a full-month rebuild and not a 25-31 top-up: the Drive
    folder only carries the JUNE BRISIM file, and the BRISIM import already in
    Odoo dropped roughly half of every day's rows (172 lines against the 318 the
    bank shows for 01-24 Jul). 81_clearing_juli.py therefore replaces the July
    IBRI statement instead of appending to it.
    """
    out = []
    for r in sheet_rows(wb, "MUTASI BRI"):
        d = day(r["Tanggal"])
        if not d.startswith("2026-07"):
            continue
        debit, credit = num(r.get("Debet")), num(r.get("Kredit"))
        amount = credit - debit
        if not amount:
            continue
        out.append({"journal": "IBRI", "date": d, "replace_month": True,
                    "ref": str(r.get("Uraian Transaksi", ""))[:200], "amount": rnd(amount)})
    return out


# ------------------------------------------------------------------- report
def build_workbook(path, plan, odoo, diag):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    money = "#,##0.00"

    def sheet(title, headers, rows, widths=None):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill, c.alignment = bold, head_fill, Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths or [], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                if isinstance(c.value, float):
                    c.number_format = money
        ws.freeze_panes = "A2"
        return ws

    ws = wb.active
    ws.title = "RINGKASAN"
    a_gross = sum(r["gross"] for r in plan["block_a"])
    a_mdr = sum(r["mdr"] for r in plan["block_a"])
    b_gross = sum(r["gross"] for r in plan["block_b"])
    b_mdr = sum(r["mdr"] for r in plan["block_b"])
    c_amt = sum(r["amount"] for r in plan["block_c"])
    rows = [
        ("Clearing Juli 2026 - prd_levis_begbal", "", ""),
        ("Sumber", "07. REPORT EBR PERIODE JULI 2026 (2).xlsx + MUTASI BCA 01-31 Juli", ""),
        ("", "", ""),
        ("BLOK", "URAIAN", "JUMLAH"),
        ("A", "Settlement Juli - bruto (Cr POS Receivable 1106000101..110)", rnd(a_gross)),
        ("A", "   di antaranya MDR (Dr 7104000001)", rnd(a_mdr)),
        ("A", "   cash-in bersih (Dr 1103000002)", rnd(a_gross - a_mdr)),
        ("B", "Collection AR Juni - bruto (Cr 1106000001)", rnd(b_gross)),
        ("B", "   di antaranya MDR", rnd(b_mdr)),
        ("C", "Sweep ATS / biaya bank BCA (Cr 1103000002)", rnd(c_amt)),
        ("", "", ""),
        ("Odoo", "Saldo AR 1106000001 sebelum clearing", odoo["ar"]),
        ("Odoo", "Saldo Bank Suspense 1103000002 sebelum clearing", odoo["suspense"]),
        ("Odoo", "POS Receivable Juli masih open", odoo["posrec"]),
        ("Odoo", "Sisa POS Receivable setelah blok A", rnd(odoo["posrec"] - a_gross)),
        ("Odoo", "Sisa AR 1106000001 setelah blok B", rnd(odoo["ar"] - b_gross)),
        ("", "", ""),
        ("SIMULASI SETELAH SEMUA BLOK DI-POSTING", "", ""),
        ("Odoo", "Bank Suspense 1103000002", odoo["sim_suspense"]),
        ("Odoo", "MDR Bank 7104000001", odoo["sim_mdr"]),
        ("Odoo", "Trade Receivables 1106000001", odoo["sim_ar"]),
        ("", "", ""),
        ("Catatan", "Tender belum tersettle per 31-Jul (tetap open)", diag["unsettled_total"]),
        ("Catatan", "Baris statement yang ditambahkan (BCA 21-31 Jul + BRI 01-31 Jul)", len(plan["block_s"])),
        ("", "", ""),
        ("PERLU KEPUTUSAN ACCOUNTING", "", ""),
        ("1", "Statement BRI Juli di Odoo hanya 172 baris dari 416 baris mutasi bank "
              "(setiap hari terpotong ~separuh) -- statement Juli IBRI diganti penuh "
              "dari sheet MUTASI BRI", ""),
        ("2", "BRI 27-Jul 'CAIR CEK UNTUK RTGS' -- rekening tujuan belum dipastikan, "
              "jadi tetap menggantung di 1103000002 Bank Suspense", 1533030000.0),
        ("3", "AR 1106000001 menjadi minus karena entry SALESMANUAL Juni "
              "(Rp 14.608.080) belum dibukukan", ""),
        ("4", "AEON BSD CITY 06/07-Jul: klasifikasi tender EBR beda dengan X70D, "
              "sisa tidak ter-clearing", 1400925.0),
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=13)
    for row in ws.iter_rows(min_row=5):
        if isinstance(row[2].value, float):
            row[2].number_format = money
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 20

    sheet("A - SETTLEMENT JULI",
          ["Tgl settle", "Toko", "Bank", "Bruto", "MDR", "Cash in", "Jml baris kredit"],
          [(r["date"], r["store"], r["bank"], r["gross"], r["mdr"], r["cash_in"], len(r["credits"]))
           for r in plan["block_a"]],
          [12, 36, 8, 18, 14, 18, 16])
    sheet("A - RINCIAN KREDIT",
          ["Tgl settle", "Toko", "Tgl transaksi", "Jumlah"],
          [(r["date"], r["store"], c["date"], c["amount"])
           for r in plan["block_a"] for c in r["credits"]],
          [12, 36, 14, 18])
    sheet("A - SPLIT TENDER EBR",
          ["Toko", "Tgl transaksi", "Akun POS Receivable", "Jumlah menurut EBR"],
          diag["ebr_split"], [36, 14, 22, 20])
    sheet("B - AR JUNI",
          ["Tanggal", "Bank", "Toko", "Bruto", "MDR", "Cash in"],
          [(r["date"], r["bank"], r["store"], r["gross"], r["mdr"], r["cash_in"]) for r in plan["block_b"]],
          [12, 8, 36, 18, 14, 18])
    sheet("C - ATS & BIAYA BANK",
          ["Tanggal", "Bank", "Jenis", "Keterangan", "Jumlah"],
          [(r["date"], r["bank"], r["kind"], r["desc"], r["amount"]) for r in plan["block_c"]],
          [12, 8, 12, 70, 18])
    sheet("D - OU PASKAL",
          ["ID baris", "Tanggal", "Akun", "Uraian", "Debit", "OU tujuan"],
          [(r["id"], r["date"], r["account"], r["name"], r["debit"], r["analytic"]) for r in plan["block_d"]],
          [10, 12, 14, 42, 18, 26])
    sheet("S - STATEMENT TAMBAHAN",
          ["Journal", "Tanggal", "Keterangan", "Jumlah"],
          [(r["journal"], r["date"], r["ref"], r["amount"]) for r in plan["block_s"]],
          [10, 12, 80, 18])
    sheet("BELUM TERSETTLE",
          ["Toko", "Jumlah"],
          sorted(diag["unsettled_per_store"].items(), key=lambda kv: -kv[1]),
          [36, 18])
    if diag["unknown_tender"]:
        sheet("TENDER TIDAK DIKENAL", ["Toko | Tanggal | Metode", "Jumlah"],
              sorted(diag["unknown_tender"].items(), key=lambda kv: -kv[1]), [60, 18])
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ebr", required=True)
    ap.add_argument("--bca-csv", required=True)
    ap.add_argument("--db", default="prd_levis_begbal")
    ap.add_argument("--json", required=True)
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    if "PGPASSWORD" not in os.environ:
        sys.exit("set PGPASSWORD first")

    wb = openpyxl.load_workbook(args.ebr, read_only=True, data_only=True)
    block_a, block_b, block_c, diag = build_settlement(wb)
    block_s = missing_bca_lines(args.bca_csv) + missing_bri_lines(wb)

    # --- block D: July RIREC POS-receivable lines without an OU analytic ----
    rows = psql(args.db, """
        select aml.id, am.date, aa.code_store->>'1', aml.name, aml.debit
        from account_move_line aml
        join account_move am on am.id = aml.move_id
        join account_account aa on aa.id = aml.account_id
        join account_journal aj on aj.id = am.journal_id
        where aj.code = 'RIREC' and aml.parent_state = 'posted'
          and am.date between '2026-07-01' and '2026-07-31'
          and aa.code_store->>'1' like '110600010%'
          and aml.analytic_distribution is null
        order by aml.id""")
    block_d = [{"id": int(r[0]), "date": r[1], "account": r[2], "name": r[3],
                "debit": float(r[4]), "analytic": "OLS SES - PASKAL BANDUNG"} for r in rows]

    # --- validate every store name against the analytic accounts ------------
    known = {r[0] for r in psql(args.db, "select name->>'en_US' from account_analytic_account")}
    unknown = sorted({r["store"] for r in block_a + block_b} - known)
    if unknown:
        sys.exit("store names with no analytic account: " + ", ".join(unknown))

    bal = {r[0]: float(r[1]) for r in psql(args.db, """
        select aa.code_store->>'1', sum(aml.debit - aml.credit)
        from account_move_line aml join account_move am on am.id = aml.move_id
        join account_account aa on aa.id = aml.account_id
        where aml.parent_state = 'posted' and am.date < '2026-08-01'
          and aa.code_store->>'1' in ('1106000001', '1103000002')
        group by 1""")}
    posrec = float(psql(args.db, """
        select coalesce(sum(aml.amount_residual), 0)
        from account_move_line aml join account_move am on am.id = aml.move_id
        join account_account aa on aa.id = aml.account_id
        where aml.parent_state = 'posted' and not aml.reconciled
          and aa.code_store->>'1' like '110600010%'
          and am.date between '2026-07-01' and '2026-07-31'""")[0][0])
    odoo = {"ar": rnd(bal.get("1106000001", 0.0)), "suspense": rnd(bal.get("1103000002", 0.0)),
            "posrec": rnd(posrec)}
    # end state measured by running 81_clearing_juli.py with CLR_DRY=1 CLR_POST=1
    odoo.update({"sim_suspense": 1530199113.09, "sim_mdr": 94236268.68, "sim_ar": -7731914.00})

    plan = {"block_a": block_a, "block_b": block_b, "block_c": block_c,
            "block_d": block_d, "block_s": block_s, "odoo": odoo}
    with open(args.json, "w") as fh:
        json.dump(plan, fh)
    build_workbook(args.xlsx, plan, odoo, diag)

    a_gross = sum(r["gross"] for r in block_a)
    print(f"block A  {len(block_a):5d} rows  gross {a_gross:>20,.2f}  mdr {sum(r['mdr'] for r in block_a):>16,.2f}")
    print(f"block B  {len(block_b):5d} rows  gross {sum(r['gross'] for r in block_b):>20,.2f}  mdr {sum(r['mdr'] for r in block_b):>16,.2f}")
    print(f"block C  {len(block_c):5d} rows        {sum(r['amount'] for r in block_c):>20,.2f}")
    print(f"block D  {len(block_d):5d} lines       {sum(r['debit'] for r in block_d):>20,.2f}")
    print(f"block S  {len(block_s):5d} statement lines")
    print(f"unsettled at 31-Jul            {diag['unsettled_total']:>20,.2f}")
    print(f"POS receivable open (Odoo)     {odoo['posrec']:>20,.2f}  -> after A {odoo['posrec'] - a_gross:>18,.2f}")
    print(f"AR 1106000001 (Odoo)           {odoo['ar']:>20,.2f}")
    print(f"json {args.json}\nxlsx {args.xlsx}")


if __name__ == "__main__":
    main()
