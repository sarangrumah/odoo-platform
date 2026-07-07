"""Extract the canonical Levi's store list to out_stores.csv (run on HOST).

IMPORTANT data reality (verified 2026-06-11): the X-prefix transaction files
currently provided (X20 on-hand, X24 sales, X70D tenders) are SINGLE-STORE SAMPLES
— they only contain store code 14694 ("OLS SCU - TUNJUNGAN PLAZA 3"). The full
24-store list lives ONLY in the Store Master header (1 DC + 23 OLS), and that sheet
carries store NAMES but no STORE CODE / SAP STORE CODE.

So:
- The canonical store list = Store Master header store columns (cols 34..57).
- STORE CODE is known only for stores that appear in a transaction file (today: 14694).
- The remaining stores' codes must be supplied by the customer (full store master with
  SAP/store codes) or will be learned as fuller X-files arrive (likely via the SFTP feed).

Output: out_stores.csv (store_name, store_code, sap_store_code, is_dc, code_known)
"""

import csv
import glob
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
LEVIS_DIR = os.path.join(REPO_ROOT, "docs", "levis")


def store_master_stores():
    """Return ordered list of store names from the Store Master header (cols 34..57)."""
    hits = glob.glob(os.path.join(LEVIS_DIR, "*STORE MASTER DATA*.xlsx"))
    if not hits:
        raise SystemExit("Store Master xlsx not found in docs/levis/")
    wb = openpyxl.load_workbook(hits[0], read_only=True, data_only=True)
    ws = wb["Stock by Store by PC13-update"] if "Stock by Store by PC13-update" in wb.sheetnames else wb.active
    names = []
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        for cell in row[33:57]:  # 0-based 33..56 == 1-based cols 34..57 (DC + 23 OLS)
            if cell and str(cell).strip():
                nm = str(cell).strip()
                if nm.upper() not in ("DISCOUNT %", "GRAND TOTAL", "TOTAL"):
                    names.append(nm)
        break
    wb.close()
    return names


def known_codes():
    """Map store_name(upper) -> (store_code, sap_code) from any transaction file present."""
    out = {}
    # X20 CSV: 0-based 15=code,16=sap,17=name
    for p in glob.glob(os.path.join(LEVIS_DIR, "X20*.csv")):
        with open(p, encoding="utf-8-sig", errors="replace", newline="") as f:
            for i, row in enumerate(csv.reader(f), start=1):
                if i == 1 or len(row) <= 17:
                    continue
                code = str(row[15]).strip()
                if code.isdigit():
                    out[str(row[17]).strip().upper()] = (code, str(row[16]).strip())
    # X24DN / X70D xlsx: 0-based 0=code,1=sap,2=name
    for pat in ("X24DN*.xlsx", "X70D*.xlsx"):
        for p in glob.glob(os.path.join(LEVIS_DIR, pat)):
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or row[0] is None:
                    continue
                code = str(row[0]).strip()
                if code.isdigit():
                    out[str(row[2]).strip().upper()] = (code, str(row[1]).strip())
            wb.close()
    return out


def main():
    stores = store_master_stores()
    codes = known_codes()
    out = os.path.join(HERE, "out_stores.csv")
    n_known = 0
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["store_name", "store_code", "sap_store_code", "is_dc", "code_known"])
        for nm in stores:
            code, sap = codes.get(nm.upper(), ("", ""))
            # fuzzy fallback: contains-match (e.g. "TUNJUNGAN PLAZA" vs "TUNJUNGAN PLAZA 3")
            if not code:
                for kname, (kc, ks) in codes.items():
                    if nm.upper() in kname or kname in nm.upper():
                        code, sap = kc, ks
                        break
            is_dc = "yes" if "WAREHOUSE" in nm.upper() or nm.upper().startswith("DC") else "no"
            if code:
                n_known += 1
            w.writerow([nm, code, sap, is_dc, "yes" if code else "no"])
    print(
        f"Wrote {out}: {len(stores)} stores ({n_known} with known codes, {len(stores) - n_known} pending codes).",
        flush=True,
    )
    print("ACTION: ask the customer for SAP/store codes for the stores marked code_known=no.", flush=True)


if __name__ == "__main__":
    main()
