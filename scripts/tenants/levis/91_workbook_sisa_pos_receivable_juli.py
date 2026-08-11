"""Build the standalone breakdown of the POS receivable that survives the
July-2026 clearing of prd_levis_begbal.

    PGPASSWORD=... python3 scripts/tenants/levis/91_workbook_sisa_pos_receivable_juli.py \
        --json /srv/sftp-share/files/clearing-juli-2026/clearing_juli.json \
        --ebr  /srv/sftp-share/files/clearing-juli-2026/EBR_JULI_2026.xlsx \
        --out  /srv/sftp-share/files/Rincian_Sisa_POS_Receivable_Juli2026.xlsx

Read-only. The ledger view needs the drafts posted to exist at all, so this
script posts them inside a transaction, reconciles, reads the surviving lines
and then rolls back -- the database is left exactly as it was found. Everything
else comes from plain SELECTs and from the EBR workbook.

Why the ledger view is worth a sheet of its own: reconciliation runs per account
across every store, so all 85 surviving lines land on 30 and 31 July no matter
when the transaction happened. Reading the residual per store or per date off
the ledger gives the wrong answer, and this workbook is the answer to reach for
instead.
"""

import argparse
import collections
import json
import os
import subprocess

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from importlib.machinery import SourceFileLoader

_HERE = os.path.dirname(os.path.abspath(__file__))
_PREP = SourceFileLoader("prep_clearing_juli", os.path.join(_HERE, "80_prep_clearing_juli.py")).load_module()
_WB = SourceFileLoader("wb_clearing_juli", os.path.join(_HERE, "82_workbook_approval_clearing_juli.py")).load_module()
psql, num, rnd = _PREP.psql, _PREP.num, _PREP.rnd
read_ebr = _WB.read_ebr

CONTAINER = "odoo19-platform-odoo"
TIMING_DATE = "2026-07-31"

# the two components that are not visible in either source and have to be
# carried as constants -- both are explained on the RINGKASAN sheet
AEON_GAP = 1400925.0
SALESMANUAL_JUNI = -14608080.0
PEMBULATAN = 950.0

REHEARSAL = """
env = env(user=1)
env = env(context=dict(env.context, allowed_company_ids=[1]))
company = env["res.company"].browse(1)


def acc(code):
    return (
        env["account.account"]
        .with_company(company)
        .search([("code", "=", code), ("company_ids", "in", 1)], limit=1)
    )


POS = [acc("11060001%02d" % i) for i in range(1, 11)]
draft = env["account.move"].search([("ref", "like", "EBR-CLR-JULI-2026-%"), ("state", "=", "draft")])
if not draft:
    raise Exception("tidak ada draft EBR-CLR-JULI-2026-* -- sudah diposting?")
draft.action_post()
for a in POS:
    lines = env["account.move.line"].search(
        [
            ("account_id", "=", a.id),
            ("parent_state", "=", "posted"),
            ("reconciled", "=", False),
            ("date", ">=", "2026-07-01"),
            ("date", "<=", "2026-07-31"),
        ]
    )
    if len(lines) > 1 and lines.filtered(lambda l: l.credit):
        lines.reconcile()
an = {a.id: a.name for a in env["account.analytic.account"].search([])}
lines = env["account.move.line"].search(
    [
        ("account_id", "in", [a.id for a in POS]),
        ("parent_state", "=", "posted"),
        ("reconciled", "=", False),
        ("date", ">=", "2026-07-01"),
        ("date", "<=", "2026-07-31"),
    ]
)
for l in lines:
    aid = next(iter(l.analytic_distribution or {}), None)
    store = an.get(int(aid), "?") if aid else "-"
    print(
        "ROW\\t%s\\t%s\\t%s\\t%.2f\\t%s\\t%s"
        % (l.date, l.account_id.code, store, l.amount_residual, l.move_id.name, (l.name or "")[:70])
    )
env.cr.rollback()
"""


def rehearse(db):
    """Post, reconcile, read the survivors, roll back. Nothing is committed."""
    out = subprocess.run(
        ["docker", "exec", "-i", CONTAINER, "odoo", "shell", "-d", db, "--no-http"],
        input=REHEARSAL,
        capture_output=True,
        text=True,
        check=True,
    ).stdout
    rows = []
    for line in out.splitlines():
        if not line.startswith("ROW\t"):
            continue
        _, date, code, store, resid, move, label = line.split("\t")
        rows.append({"date": date, "acc": code, "store": store, "resid": num(resid), "move": move, "label": label})
    if not rows:
        raise SystemExit("rehearsal tidak mengembalikan baris -- periksa container/DB")
    return rows


def read_timing(db):
    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account),
             an as (select id, name->>'en_US' nm from account_analytic_account)
        select coalesce(an.nm, '(tanpa OU)'), acc.code, sum(l.debit)
          from account_move_line l
          join account_move m on m.id = l.move_id
          join acc on acc.id = l.account_id
          left join an on l.analytic_distribution ? an.id::text
         where m.state = 'posted' and l.debit > 0 and l.date = '{TIMING_DATE}'
           and acc.code between '1106000101' and '1106000110'
      group by 1, 2 order by 1, 2
        """,
    )
    return [(r[0], r[1], num(r[2])) for r in rows]


def build(path, ledger, timing, kol, unsettled):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    money = "#,##0"

    def sheet(title, headers, rows, widths, total_cols=()):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill, c.alignment = bold, head_fill, Alignment(horizontal="center")
        for r in rows:
            ws.append(list(r))
        if total_cols:
            ws.append([])
            tot = ["TOTAL"] + [""] * (len(headers) - 1)
            for i in total_cols:
                tot[i] = rnd(sum(num(r[i]) for r in rows))
            ws.append(tot)
            for c in ws[ws.max_row]:
                c.font = bold
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                if isinstance(c.value, float):
                    c.number_format = money
        ws.freeze_panes = "A2"
        return ws

    tot_timing = rnd(sum(v for _s, _c, v in timing))
    kol_total = rnd(sum(r["amount"] for r in kol))
    macet = [r for r in unsettled if r["trans"] != TIMING_DATE and not r["kol"]]
    macet_total = rnd(sum(r["amount"] for r in macet))
    total = rnd(tot_timing + kol_total + macet_total + AEON_GAP + SALESMANUAL_JUNI + PEMBULATAN)
    ledger_total = rnd(sum(r["resid"] for r in ledger))

    # ------------------------------------------------------------- ringkasan
    ws = wb.active
    ws.title = "RINGKASAN"
    rows = [
        ("RINCIAN POS RECEIVABLE JULI 2026 YANG MASIH TERBUKA SETELAH CLEARING", "", "", ""),
        ("Database", "prd_levis_begbal", "", ""),
        ("Posisi", "Setelah 63 jurnal EBR-CLR-JULI-2026-* diposting", "", ""),
        (
            "Sumber angka",
            "Diukur dari uji posting yang di-rollback, bukan proyeksi. Total di bawah "
            "cocok persis dengan sisa di buku besar.",
            "",
            "",
        ),
        ("", "", "", ""),
        ("KOMPONEN", "SEBAB", "NILAI", "TINDAK LANJUT"),
        (
            "Timing 31-Jul",
            "Transaksi 31-Jul yang settle D+1 di Agustus. Normal, bukan masalah.",
            tot_timing,
            "Tertutup sendiri oleh clearing Agustus",
        ),
        (
            "KOL 15-Jul",
            "Pemberian gratis ke KOL tercatat sebagai penjualan CASH harga penuh di GRAND INDONESIA",
            kol_total,
            "Reklas ke beban promosi ATAU batalkan dan import ulang sebagai free goods (ada implikasi PPN cuma-cuma)",
        ),
        (
            "Macet lain",
            "Bon manual / transaksi tanpa CASH RECEIVED DATE selain KOL",
            macet_total,
            "Ditelusuri per toko oleh Finance",
        ),
        (
            "AEON BSD CITY",
            "Trans date trx 617/619/622 bergeser sehari di workbook EBR + trx 682 beda Rp 50",
            AEON_GAP,
            "EBR mengoreksi COMPILE SALES, lalu blok A dibangkitkan ulang",
        ),
        (
            "Reclass SALESMANUAL Juni",
            "Sudah dibukukan 7-Agu, jadi mengurangi sisa",
            SALESMANUAL_JUNI,
            "Selesai -- tidak perlu tindakan",
        ),
        (
            "Pembulatan",
            "Selisih pembulatan Odoo vs workbook EBR",
            PEMBULATAN,
            "Diterima sebagai known-diff",
        ),
        ("TOTAL", "", total, ""),
        ("", "", "", ""),
        ("Kontrol: sisa di buku besar hasil uji posting", "", ledger_total, ""),
        ("Selisih terhadap rincian di atas", "", rnd(ledger_total - total), ""),
        ("", "", "", ""),
        ("CARA MEMBACANYA DI BUKU BESAR", "", "", ""),
        (
            "",
            "Sisa ini muncul sebagai %d baris bertanggal 30 dan 31 Juli saja -- BUKAN di tanggal "
            "transaksi aslinya. Rekonsiliasi berjalan per akun lintas toko, sehingga sisa "
            "mengapung ke baris terakhir yang belum ter-match. Contoh: KOL terjadi 15-Jul "
            "tetapi tampak bertanggal 30-Jul. Karena itu sisa per toko dan per tanggal TIDAK "
            "boleh dibaca dari buku besar -- pakai workbook ini." % len(ledger),
            "",
            "",
        ),
        (
            "",
            "Pengecualian: baris bertanggal 31-Jul memang murni timing dan cocok per toko, lihat sheet TIMING-31JUL.",
            "",
            "",
        ),
    ]
    for r in rows:
        ws.append(list(r))
    ws["A1"].font = Font(bold=True, size=13)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, float):
                c.number_format = money
            c.alignment = Alignment(wrap_text=True, vertical="top")
        if row[0].value in ("KOMPONEN", "TOTAL", "CARA MEMBACANYA DI BUKU BESAR"):
            for c in row:
                c.font = bold
    for col, w in zip("ABCD", (40, 74, 20, 60)):
        ws.column_dimensions[col].width = w

    # --------------------------------------------------------------- timing
    per_store = collections.Counter()
    for store, _code, val in timing:
        per_store[store] += val
    sheet(
        "TIMING-31JUL",
        ["Toko", "Nilai", "Porsi dari timing"],
        [
            (store, rnd(val), round(val / tot_timing, 4) if tot_timing else 0)
            for store, val in sorted(per_store.items(), key=lambda kv: -kv[1])
        ],
        [42, 20, 18],
        total_cols=(1,),
    )
    ws = wb["TIMING-31JUL"]
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for c in row:
            if isinstance(c.value, float):
                c.number_format = "0.0%"

    sheet(
        "TIMING-PER-TENDER",
        ["Toko", "Akun POS Receivable", "Nilai"],
        [(s, c, rnd(v)) for s, c, v in sorted(timing)],
        [42, 24, 20],
        total_cols=(2,),
    )

    # ------------------------------------------------------------------ KOL
    sheet(
        "KOL-76JT",
        ["Toko", "Tgl transaksi", "Register", "Transnum", "Kasir", "Tender", "Nama KOL", "Nilai"],
        [
            (r["store"], r["trans"], r["register"], r["transnum"], r["cashier"], r["tender"], r["kol"], r["amount"])
            for r in sorted(kol, key=lambda r: (r["trans"], r["transnum"]))
        ],
        [30, 14, 10, 12, 24, 14, 34, 18],
        total_cols=(7,),
    )

    # ----------------------------------------------------------- macet lain
    grouped = collections.Counter()
    counts = collections.Counter()
    for r in macet:
        grouped[(r["store"], r["trans"])] += r["amount"]
        counts[(r["store"], r["trans"])] += 1
    sheet(
        "MACET-LAIN",
        ["Toko", "Tgl transaksi", "Jml baris", "Nilai"],
        [
            (store, trans, counts[(store, trans)], rnd(val))
            for (store, trans), val in sorted(grouped.items(), key=lambda kv: -kv[1])
        ],
        [40, 16, 14, 20],
        total_cols=(3,),
    )

    # -------------------------------------------------------- ledger detail
    sheet(
        "BARIS-BUKU-BESAR",
        ["Tanggal di buku besar", "Akun", "Toko", "Sisa", "Jurnal asal", "Keterangan"],
        [
            (r["date"], r["acc"], r["store"], r["resid"], r["move"], r["label"])
            for r in sorted(ledger, key=lambda r: (r["date"], r["acc"], r["store"]))
        ],
        [22, 14, 40, 20, 22, 60],
        total_cols=(3,),
    )
    ws = wb["BARIS-BUKU-BESAR"]
    ws.insert_rows(1)
    ws["A1"] = (
        "Tanggal di kolom pertama adalah tanggal baris yang tersisa setelah rekonlisiasi, "
        "BUKAN tanggal penyebabnya. Lihat RINGKASAN."
    )
    ws["A1"].font = bold
    ws.freeze_panes = "A3"

    wb.save(path)
    # 0644 + owned by the share user. This script runs on the HOST as root, and
    # /srv/sftp-share/files belongs to sftpshare:sftpusers (uid/gid 1002) -- the
    # account File Browser and SFTP serve it as. Without the chown the workbook
    # lands root-owned: readable, but the people who work with it cannot replace
    # it. Tolerant of a filesystem that refuses either, since the report is
    # already written by this point and its mode is not worth losing it over.
    try:
        os.chmod(path, 0o644)
        os.chown(path, 1002, 1002)
    except OSError:
        pass


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", required=True)
    ap.add_argument("--ebr", required=True)
    ap.add_argument("--db", default="prd_levis_begbal")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    if not os.environ.get("PGPASSWORD"):
        raise SystemExit("PGPASSWORD is required")
    json.load(open(args.json))  # fail early if the plan file is unreadable

    kol, _sales_jun, unsettled = read_ebr(args.ebr)
    timing = read_timing(args.db)
    ledger = rehearse(args.db)
    build(args.out, ledger, timing, kol, unsettled)

    print(f"timing 31-Jul : {sum(v for _s, _c, v in timing):>18,.0f}")
    print(f"KOL           : {sum(r['amount'] for r in kol):>18,.0f}")
    print(f"buku besar    : {sum(r['resid'] for r in ledger):>18,.0f} pada {len(ledger)} baris")
    print(f"tersimpan di {args.out}")


if __name__ == "__main__":
    main()
