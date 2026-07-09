# HOST script (needs openpyxl). Extracts the EBR finance workbook into loader CSVs.
#
#   python scripts/tenants/levis/59_export_ebr.py "<path to 2026-06 - EBR - TB and GL ....xlsx>"
#
# Produces, next to this script:
#   ebr_coa.csv   code,name,account_type                    -> feeds 30_fix_coa.py (as /tmp/ebr_coa.csv)
#   tb_ebr.csv    account,index,description,12x(beg,d,c,end),ending_ytd,diff  -> feeds 60_load_tb.py
#   gl_ebr.csv    the raw single-sided GL line extract        -> reference / future 61_load_gl.py
#
# The source GL sheet is single-sided (each row is one leg only, no contra account,
# Document No on <5% of rows) so it CANNOT be posted as balanced journal entries as-is.
# gl_ebr.csv is exported for reference; 61_load_gl.py expects a corrected 2-sided export.
import csv
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
src = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "ebr_workbook.xlsx")
if not os.path.exists(src):
    sys.exit("workbook not found: %s (pass path as argv[1])" % src)

wb = openpyxl.load_workbook(src, read_only=True, data_only=True)


def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def s(x):
    return "" if x is None else str(x).strip()


# Accounts the TB references but the finance CoA sheet omits (verified against the workbook).
# 1117400001 "Tax Deposit" — sibling 1117300001 "GST-IN" is asset_current.
SUPPLEMENT = [
    ("1117400001", "Tax Deposit", "asset_current"),
]

# ---------------------------------------------------------------- CoA
ws = wb["CoA EBR"]
rows = list(ws.iter_rows(values_only=True))
out = os.path.join(HERE, "ebr_coa.csv")
with open(out, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["code", "name", "account_type"])
    seen = set()
    n = 0
    for r in rows[2:]:  # header at row index 1
        if not s(r[0]):
            continue
        w.writerow([s(r[0]), s(r[1]), s(r[2])])  # strip trailing/leading spaces (e.g. ' asset_current')
        seen.add(s(r[0]))
        n += 1
    for code, name, at in SUPPLEMENT:
        if code not in seen:
            w.writerow([code, name, at])
            n += 1
print("[export] ebr_coa.csv     : %d accounts (incl %d supplement)" % (n, len(SUPPLEMENT)))

# ---------------------------------------------------------------- Trial Balance
ws = wb["Trial Balance EBR 2026"]
rows = list(ws.iter_rows(values_only=True))
out = os.path.join(HERE, "tb_ebr.csv")
with open(out, "w", newline="") as f:
    w = csv.writer(f)
    head = ["account", "index", "description"]
    for m in range(1, 13):
        head += ["m%d_beg" % m, "m%d_d" % m, "m%d_c" % m, "m%d_end" % m]
    head += ["ending_ytd", "diff"]
    w.writerow(head)
    n = 0
    for r in rows[6:]:  # header at row index 5, data from 6
        acc = s(r[0])
        if not acc.replace(".", "", 1).isdigit():
            continue
        row = [acc, s(r[1]), s(r[2])]
        for m in range(12):
            base = 6 + 4 * m  # Beginning,D,C,Ending blocks start at col 6
            row += [num(r[base]), num(r[base + 1]), num(r[base + 2]), num(r[base + 3])]
        row += [num(r[54]), num(r[55])]
        w.writerow(row)
        n += 1
print("[export] tb_ebr.csv      : %d accounts" % n)

# ---------------------------------------------------------------- GL (reference)
ws = wb["GL EBR 2026"]
rows = list(ws.iter_rows(values_only=True))
out = os.path.join(HERE, "gl_ebr.csv")
GL_COLS = [
    "no",
    "year",
    "period",
    "txn_type",
    "doc_no",
    "invoice_ref",
    "doc_date",
    "posting_date",
    "account",
    "account_desc",
    "d",
    "c",
    "amount",
    "notes",
    "bank_mapping",
    "business_partner",
    "store",
    "remarks",
]
with open(out, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(GL_COLS)
    n = 0
    for r in rows[6:]:  # header at row index 5
        if r[1] in (None, "") or not str(r[1]).replace(".", "", 1).isdigit():
            continue
        w.writerow([s(x) for x in r[1:19]])
        n += 1
print("[export] gl_ebr.csv      : %d lines (single-sided, reference only)" % n)
print("[export] done -> %s" % HERE)
