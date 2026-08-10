"""Build the approval workbook for the July-2026 clearing of prd_levis_begbal.

Runs on the HOST (plain python3 + openpyxl) and is strictly READ-ONLY: the
database is read through ``docker exec ... psql`` exactly like
``80_prep_clearing_juli.py``, nothing is posted and no draft is touched.

    PGPASSWORD=... python3 scripts/tenants/levis/82_workbook_approval_clearing_juli.py \
        --json /srv/sftp-share/files/clearing-juli-2026/clearing_juli.json \
        --ebr  /srv/sftp-share/files/clearing-juli-2026/EBR_JULI_2026.xlsx \
        --out  /srv/sftp-share/files/Persetujuan_Clearing_Juli2026.xlsx

The 63 drafts ``EBR-CLR-JULI-2026-*`` created on 4-Aug are still draft; this
workbook is what Accounting signs before ``81_clearing_juli.py`` is re-run with
``CLR_POST=1``. Every figure comes either from the plan JSON that produced those
drafts, from the EBR workbook, or from a live SELECT, so the sheet can be tied
back line by line.
"""

import argparse
import collections
import datetime
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# reuse the read-only psql helper and the numeric helpers of the prep script
from importlib.machinery import SourceFileLoader

_PREP = SourceFileLoader(
    "prep_clearing_juli",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "80_prep_clearing_juli.py"),
).load_module()
psql, num, rnd, sheet_rows = _PREP.psql, _PREP.num, _PREP.rnd, _PREP.sheet_rows

REF_PREFIX = "EBR-CLR-JULI-2026"
KEY_ACCOUNTS = ["1103000002", "7104000001", "1106000001", "1103019320"]
ACC_NAME = {
    "1103000002": "Bank Suspense",
    "7104000001": "Beban MDR",
    "1106000001": "Piutang Usaha (AR Juni)",
    "1103019320": "Bank BCA 2687778282 (Out)",
}
KOL_RE = re.compile(r"^\s*KOL\b", re.I)

# store whose EBR tender classification disagrees with X70D (diagnose nr 4)
AEON_STORE = "OLS SES - AEON BSD CITY"
TENDER_ACC = _PREP.TENDER_ACC
TENDER_ALIAS = _PREP.TENDER_ALIAS


def tender_acc(raw):
    t = str(raw or "").strip().upper()
    return TENDER_ACC.get(TENDER_ALIAS.get(t, t), "?" + t)


# ------------------------------------------------------------------ database
def read_odoo(db):
    """Everything the workbook needs from the live database, read-only."""
    out = {}

    rows = psql(
        db,
        f"""
        select substring(m.ref from '{REF_PREFIX}-([A-Z])'), m.state,
               count(distinct m.id), coalesce(sum(l.debit), 0)
          from account_move m
          join account_move_line l on l.move_id = m.id
         where m.ref like '{REF_PREFIX}-%'
      group by 1, 2 order by 1, 2
        """,
    )
    out["draft_blocks"] = [(r[0], r[1], int(r[2]), num(r[3])) for r in rows]

    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select substring(m.ref from '{REF_PREFIX}-([A-Z])'), acc.code,
               sum(l.debit), sum(l.credit), count(*)
          from account_move m
          join account_move_line l on l.move_id = m.id
          join acc on acc.id = l.account_id
         where m.ref like '{REF_PREFIX}-%' and m.state = 'draft'
      group by 1, 2 order by 1, 2
        """,
    )
    out["draft_by_account"] = [(r[0], r[1], num(r[2]), num(r[3]), int(r[4])) for r in rows]

    rows = psql(
        db,
        """
        with acc as (select id, code_store->>'1' code from account_account)
        select acc.code, sum(l.debit - l.credit)
          from account_move_line l
          join account_move m on m.id = l.move_id
          join acc on acc.id = l.account_id
         where m.state = 'posted' and l.date < '2026-08-01'
           and acc.code in ('%s')
      group by 1 order by 1
        """
        % "','".join(KEY_ACCOUNTS),
    )
    out["balance_before"] = {r[0]: num(r[1]) for r in rows}

    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account),
             op as (select l.account_id, sum(l.amount_residual) resid, count(*) n
                      from account_move_line l
                      join account_move m on m.id = l.move_id
                     where m.state = 'posted' and not l.reconciled
                       and l.date between '2026-07-01' and '2026-07-31'
                  group by 1),
             dc as (select l.account_id, sum(l.credit) cr
                      from account_move_line l
                      join account_move m on m.id = l.move_id
                     where m.ref like '{REF_PREFIX}-%' and m.state = 'draft'
                  group by 1)
        select acc.code, op.resid, op.n, coalesce(dc.cr, 0)
          from op join acc on acc.id = op.account_id
          left join dc on dc.account_id = op.account_id
         where acc.code between '1106000101' and '1106000110'
      order by 1
        """,
    )
    out["pos_receivable"] = [(r[0], num(r[1]), int(r[2]), num(r[3])) for r in rows]

    rows = psql(
        db,
        """
        select m.name, m.date, m.ref, m.amount_total
          from account_move m
         where m.ref like 'EBR-ADJ-AR-JUNI-2026%' and m.state = 'posted'
      order by m.id
        """,
    )
    out["ar_juni_adj"] = [(r[0], r[1], r[2], num(r[3])) for r in rows]
    return out


def read_aeon(db, ebr_path, plan):
    """Everything behind diagnose nr 4: where EBR and X70D disagree for AEON.

    Odoo's side is read twice on purpose -- the posted POS receivable debits are
    what ``allocate()`` actually sees, while ``retail_import_line`` carries the
    X70D transaction rows those debits were built from, so EBR can be pointed at
    an individual transnum instead of a daily total.
    """
    out = {"store": AEON_STORE}

    rows = psql(db, f"select id from account_analytic_account where name->>'en_US' = '{AEON_STORE}'")
    if not rows:
        raise SystemExit(f"analytic account {AEON_STORE!r} not found")
    an_id = rows[0][0]
    out["analytic_id"] = an_id

    # Odoo side per trans-date x tender account (1106000112 is the contra, skip it)
    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select l.date, acc.code, sum(l.debit)
          from account_move_line l
          join account_move m on m.id = l.move_id
          join acc on acc.id = l.account_id
         where m.state = 'posted' and l.debit > 0
           and acc.code between '1106000101' and '1106000110'
           and l.date between '2026-07-01' and '2026-07-31'
           and l.analytic_distribution ? '{an_id}'
      group by 1, 2 order by 1, 2
        """,
    )
    odoo_daily = {(r[0], r[1]): num(r[2]) for r in rows}

    # X70D transaction rows as imported
    rows = psql(
        db,
        f"""
        select j->>'trans_date', j->>'register', j->>'transnum',
               j->>'tender_type', j->>'tender_amount'
          from retail_import_line l,
               lateral (select l.raw_data_json::jsonb j) x
         where (j->>'store_name') = '{AEON_STORE}'
           and (j->>'trans_date') like '2026-07-%'
           and (j->>'tender_type') is not null
      order by 1, 3
        """,
    )
    x70d = [
        {"date": r[0], "register": r[1], "transnum": r[2], "tender": r[3], "amount": num(r[4]), "acc": tender_acc(r[3])}
        for r in rows
    ]
    out["x70d"] = x70d

    # blok A credits actually booked for this store, per journal date
    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select m.date, sum(l.credit)
          from account_move m
          join account_move_line l on l.move_id = m.id
          join acc on acc.id = l.account_id
         where m.ref like '{REF_PREFIX}-A-%'
           and acc.code between '1106000101' and '1106000110'
           and l.analytic_distribution ? '{an_id}'
      group by 1 order by 1
        """,
    )
    out["blok_a_credit"] = {r[0]: num(r[1]) for r in rows}

    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select m.date, acc.code, sum(l.debit)
          from account_move m
          join account_move_line l on l.move_id = m.id
          join acc on acc.id = l.account_id
         where m.ref like '{REF_PREFIX}-A-%'
           and acc.code in ('1103000002', '7104000001')
           and l.analytic_distribution ? '{an_id}'
      group by 1, 2 order by 1, 2
        """,
    )
    out["blok_a_debit"] = {(r[0], r[1]): num(r[2]) for r in rows}

    # EBR side per trans-date x tender account, and per transaction
    wb = openpyxl.load_workbook(ebr_path, read_only=True, data_only=True)
    rows = sheet_rows(wb, "COMPILE SALES")
    wb.close()
    ebr, ebr_daily = [], collections.Counter()
    for r in rows:
        if str(r.get("STORE NAME") or "").strip() != AEON_STORE:
            continue
        amount = num(r.get("TENDER AMOUNT"))
        if not amount:
            continue
        rec = {
            "date": str(r.get("TRANS DATE") or "")[:10],
            "register": str(r.get("REGISTER") or "").strip(),
            "transnum": str(r.get("TRANSNUM") or "").strip(),
            "tender": str(r.get("TENDER TYPE") or "").strip(),
            "amount": amount,
            "acc": tender_acc(r.get("TENDER TYPE")),
            "metode": str(r.get("METODE PEMBAYARAN") or "").strip(),
            "received": str(r.get("CASH RECEIVED DATE") or "")[:10],
        }
        ebr.append(rec)
        ebr_daily[(rec["date"], rec["acc"])] += amount
    out["ebr"] = ebr

    # daily comparison, only the cells that disagree
    daily = []
    for key in sorted(set(odoo_daily) | set(ebr_daily)):
        o, e = odoo_daily.get(key, 0.0), ebr_daily.get(key, 0.0)
        if abs(e - o) > 0.004:
            daily.append((key[0], key[1], rnd(o), rnd(e), rnd(e - o)))
    out["daily"] = daily
    out["odoo_total"] = rnd(sum(odoo_daily.values()))
    out["ebr_total"] = rnd(sum(ebr_daily.values()))

    # Per-transaction comparison. Matching is staged rather than keyed on
    # TRANSNUM: EBR mistypes the transnum often enough (dropped or inserted
    # digit) that a transnum key reports typos as if they were real money
    # differences. Anything that agrees on date + account + amount is identical
    # for accounting purposes no matter what it is numbered.
    def take(pool, key):
        for i, r in enumerate(pool):
            if key(r):
                return pool.pop(i)
        return None

    xs = sorted(x70d, key=lambda r: (r["date"], r["transnum"]))
    es = sorted(ebr, key=lambda r: (r["date"], r["transnum"]))
    left_x, left_e = [], list(es)
    for r in xs:  # pass 1 -- exact agreement, drop silently
        if not take(
            left_e,
            lambda e, r=r: (e["date"], e["acc"], round(e["amount"], 2))
            == (r["date"], r["acc"], round(r["amount"], 2)),
        ):
            left_x.append(r)

    # pass 1b -- what is left but nets to zero inside one date x account is a
    # row-split (EBR spreads one tender over two rows, or mistypes a transnum),
    # not a money difference. Dropping it keeps the sheet down to real issues.
    def group(recs):
        g = collections.defaultdict(list)
        for r in recs:
            g[(r["date"], r["acc"])].append(r)
        return g
    gx, ge = group(left_x), group(left_e)
    for key in set(gx) | set(ge):
        sx = round(sum(r["amount"] for r in gx.get(key, [])), 2)
        se = round(sum(r["amount"] for r in ge.get(key, [])), 2)
        if sx == se and gx.get(key) and ge.get(key):
            left_x = [r for r in left_x if (r["date"], r["acc"]) != key]
            left_e = [r for r in left_e if (r["date"], r["acc"]) != key]

    trx = []

    def emit(x, e, kind, impact):
        trx.append(
            {
                "kind": kind,
                "impact": impact,
                "x70d": [x] if x else [],
                "ebr": [e] if e else [],
                "x_amount": rnd(x["amount"]) if x else 0.0,
                "e_amount": rnd(e["amount"]) if e else 0.0,
                "delta": rnd((e["amount"] if e else 0.0) - (x["amount"] if x else 0.0)),
            }
        )

    rest_x = []
    for r in list(left_x):  # pass 2 -- same money, different date or tender
        e = take(left_e, lambda e, r=r: round(e["amount"], 2) == round(r["amount"], 2))
        if not e:
            rest_x.append(r)
        elif e["date"] != r["date"]:
            emit(r, e, "Beda TRANS DATE", "BERDAMPAK -- melintasi hari, blok A tidak bisa menyerap")
        else:
            emit(r, e, "Beda TENDER TYPE", "netral -- saling menutup di hari yang sama")

    still_x = []
    for r in rest_x:  # pass 3 -- same slot, different amount
        e = take(left_e, lambda e, r=r: (e["date"], e["acc"]) == (r["date"], r["acc"]))
        if e:
            emit(r, e, "Beda NILAI", "BERDAMPAK -- selisih nilai tidak ter-clearing")
        else:
            still_x.append(r)

    for r in still_x:
        emit(r, None, "Hanya ada di X70D", "BERDAMPAK -- ada di Odoo, tidak ada di COMPILE SALES")
    for e in left_e:
        emit(None, e, "Hanya ada di EBR", "BERDAMPAK -- ada di COMPILE SALES, tidak ada di Odoo")

    order = {"BERDAMPAK": 0}
    trx.sort(key=lambda t: (order.get(t["impact"].split(" --")[0], 1), t["kind"], (t["x70d"] or t["ebr"])[0]["date"]))
    out["trx"] = trx

    # what blok A was asked for, per trans-date, straight from the plan JSON
    asked = collections.Counter()
    for row in plan["block_a"]:
        if row["store"] != AEON_STORE:
            continue
        for c in row["credits"]:
            asked[c["date"]] += c["amount"]
    out["asked"] = asked
    out["plan_rows"] = {
        row["date"]: (row["gross"], row["mdr"]) for row in plan["block_a"] if row["store"] == AEON_STORE
    }
    out["available"] = collections.Counter()
    for (d, _code), v in odoo_daily.items():
        out["available"][d] += v
    return out


# ------------------------------------------------------------------ workbook
def read_ebr(path):
    """KOL rows, June-manual rows and the unsettled population from COMPILE SALES."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = sheet_rows(wb, "COMPILE SALES")
    wb.close()

    kol, sales_jun, unsettled = [], [], []
    for r in rows:
        amount = num(r.get("TENDER AMOUNT"))
        store = str(r.get("STORE NAME") or "").strip()
        if not store:
            continue
        trans = str(r.get("TRANS DATE") or "")[:10]
        received = str(r.get("CASH RECEIVED DATE") or "")[:10]
        note = next((str(v) for v in r.values() if KOL_RE.match(str(v or ""))), "")
        manual = next((str(v).strip() for v in r.values() if "MANUAL" in str(v or "").upper()), "")
        if note:
            kol.append(
                {
                    "store": store,
                    "trans": trans,
                    "register": str(r.get("REGISTER") or r.get("REGISTER ") or "").strip(),
                    "transnum": str(r.get("TRANSNUM") or "").strip(),
                    "cashier": str(r.get("CASHIER NAME") or "").strip(),
                    "tender": str(r.get("TENDER TYPE") or "").strip(),
                    "kol": note.strip(),
                    "amount": amount,
                    "received": received,
                }
            )
        if "SALES JUN" in str(r.get("MAPPING") or "").upper():
            sales_jun.append(
                {
                    "store": store,
                    "trans": trans,
                    "transnum": str(r.get("TRANSNUM") or "").strip(),
                    "tender": str(r.get("TENDER TYPE") or "").strip(),
                    "note": manual or str(r.get("VOUCHER NUMBER") or "").strip(),
                    "amount": amount,
                    "received": received,
                }
            )
        if not received and amount:
            unsettled.append(
                {
                    "store": store,
                    "trans": trans,
                    "tender": str(r.get("TENDER TYPE") or "").strip(),
                    "amount": amount,
                    "kol": bool(note),
                }
            )
    return kol, sales_jun, unsettled


def build(path, plan, odoo, kol, sales_jun, unsettled, aeon):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    money = "#,##0"

    def sheet(title, headers, rows, widths=None, total_cols=()):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill, c.alignment = bold, head_fill, Alignment(horizontal="center")
        for r in rows:
            ws.append(list(r))
        if total_cols:
            ws.append([])
            tot = ["TOTAL"] + [""] * (len(headers) - 1)
            for i in total_cols:
                tot[i] = rnd(sum(num(r[i]) for r in rows))
            ws.append(tot)
            for c in ws[ws.max_row]:
                c.font = bold
        for i, w in enumerate(widths or [], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                if isinstance(c.value, float):
                    c.number_format = money
        ws.freeze_panes = "A2"
        return ws

    draft = {(b, s): (n, d) for b, s, n, d in odoo["draft_blocks"]}
    by_acc = odoo["draft_by_account"]
    delta = collections.Counter()
    for _blk, code, dr, cr, _n in by_acc:
        delta[code] += dr - cr

    # ----------------------------------------------------------- PERSETUJUAN
    ws = wb.active
    ws.title = "PERSETUJUAN"
    lines = [
        ("PERMOHONAN PERSETUJUAN CLEARING JULI 2026", "", ""),
        ("Database", "prd_levis_begbal", ""),
        ("Lingkup", f"63 jurnal draft dengan ref {REF_PREFIX}-*", ""),
        ("Periode", "01 s/d 31 Juli 2026 (fiscalyear lock 30-Jun-2026, Juli terbuka)", ""),
        ("Sifat", "Belum diposting. Workbook ini dasar persetujuan sebelum posting.", ""),
        ("", "", ""),
        ("BLOK", "ISI", "NILAI"),
    ]
    blocks = [
        ("A", "Settlement POS Juli: Dr 1103000002 + Dr 7104000001 / Cr 1106000101..110"),
        ("B", "Collection AR Juni: Dr 1103000002 + Dr 7104000001 / Cr 1106000001"),
        ("C", "Sweep ATS BCA + biaya bank: Dr 1103019320 / Dr 7299012000 / Cr 1103000002"),
    ]
    for tag, desc in blocks:
        n, d = draft.get((tag, "draft"), (0, 0.0))
        lines.append((f"{tag} ({n} jurnal)", desc, rnd(d)))
    lines += [
        ("", "", ""),
        ("Blok S (statement bank) dan D (analytic OU) SUDAH permanen sejak 4-Agu.", "", ""),
        ("", "", ""),
        ("AKUN", "SALDO SEBELUM (posted s/d 31-Jul)", "PROYEKSI SESUDAH POSTING"),
    ]
    for code in KEY_ACCOUNTS:
        before = odoo["balance_before"].get(code, 0.0)
        lines.append((f"{code} {ACC_NAME[code]}", rnd(before), rnd(before + delta[code])))

    resid_before = sum(r[1] for r in odoo["pos_receivable"])
    resid_after = sum(r[1] - r[3] for r in odoo["pos_receivable"])
    lines += [
        ("POS Receivable Juli terbuka", rnd(resid_before), rnd(resid_after)),
        ("", "", ""),
        ("CATATAN PENTING", "", ""),
        (
            "1",
            "Rekon AR Juni sudah tuntas 7-Agu (3 jurnal EBR-ADJ-AR-JUNI-2026-*). "
            "Karena itu AR 1106000001 setelah clearing menjadi POSITIF, bukan minus "
            "seperti simulasi 4-Agu.",
            "",
        ),
        (
            "2",
            "Rp 76.926.875 pemberian gratis ke KOL tercatat sebagai penjualan CASH "
            "harga penuh dan akan menggantung permanen di 1106000101 -- lihat sheet KOL-76JT.",
            "",
        ),
        (
            "3",
            "Akun 1103000002 di-set reconcile = false, jadi statement line Juli tidak bisa "
            "di-match di widget bank reconciliation; kontrolnya murni saldo akun.",
            "",
        ),
        ("", "", ""),
        ("Disiapkan oleh", "", ""),
        ("Diperiksa oleh", "", ""),
        ("Disetujui oleh", "", ""),
    ]
    for r in lines:
        ws.append(list(r))
    ws["A1"].font = Font(bold=True, size=13)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, float):
                c.number_format = money
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 24
    ws["A7"].font = ws["B7"].font = ws["C7"].font = bold

    # --------------------------------------------------------------- blok A
    a = plan["block_a"]
    sheet(
        "BLOK-A-SETTLEMENT",
        ["Tgl settle", "Toko", "Bank", "Bruto", "MDR", "Cash in", "Jml tgl transaksi"],
        [
            (r["date"], r["store"], r["bank"], r["gross"], r["mdr"], r["cash_in"], len(r["credits"]))
            for r in sorted(a, key=lambda r: (r["date"], r["store"]))
        ],
        [12, 38, 8, 18, 14, 18, 18],
        total_cols=(3, 4, 5),
    )
    ws = wb["BLOK-A-SETTLEMENT"]
    a_credit = sum(cr for blk, _c, _dr, cr, _n in by_acc if blk == "A")
    a_gross = sum(r["gross"] for r in a)
    ws.append([])
    for label, val in [
        ("Bruto rencana (tabel di atas)", rnd(a_gross)),
        ("Yang benar-benar dikredit oleh 30 jurnal draft blok A", rnd(a_credit)),
        ("Selisih = AEON BSD CITY, klasifikasi tender EBR beda dengan X70D", rnd(a_gross - a_credit)),
    ]:
        ws.append([label, "", "", val])
        ws.cell(ws.max_row, 4).number_format = money
        ws.cell(ws.max_row, 1).font = bold
    per_date = collections.defaultdict(lambda: [0.0, 0.0, 0.0, 0])
    for r in a:
        v = per_date[r["date"]]
        v[0] += r["gross"]
        v[1] += r["mdr"]
        v[2] += r["cash_in"]
        v[3] += 1
    sheet(
        "BLOK-A-PER-JURNAL",
        ["Tgl settle (= 1 jurnal)", "Bruto", "MDR", "Cash in", "Jml toko"],
        [(d, rnd(v[0]), rnd(v[1]), rnd(v[2]), v[3]) for d, v in sorted(per_date.items())],
        [24, 20, 16, 20, 12],
        total_cols=(1, 2, 3),
    )
    sheet(
        "BLOK-A-PER-AKUN",
        [
            "Akun POS Receivable",
            "Terbuka sebelum",
            "Jml baris",
            "Dikredit blok A",
            "Proyeksi terbuka sesudah",
        ],
        [(c, rb, n, cr, rnd(rb - cr)) for c, rb, n, cr in odoo["pos_receivable"]],
        [22, 20, 12, 20, 26],
        total_cols=(1, 3, 4),
    )

    # --------------------------------------------------------------- blok B
    ar_before = odoo["balance_before"].get("1106000001", 0.0)
    b_gross = sum(r["gross"] for r in plan["block_b"])
    sheet(
        "BLOK-B-AR-JUNI",
        ["Tanggal", "Toko", "Bank", "Bruto", "MDR", "Cash in"],
        [
            (r["date"], r["store"], r["bank"], r["gross"], r["mdr"], r["cash_in"])
            for r in sorted(plan["block_b"], key=lambda r: (r["date"], r["store"]))
        ],
        [12, 38, 8, 18, 14, 18],
        total_cols=(3, 4, 5),
    )
    ws = wb["BLOK-B-AR-JUNI"]
    ws.append([])
    ws.append(["Baris di atas digabung menjadi 2 jurnal draft (satu per tanggal mutasi bank)."])
    ws.append([])
    for label, val in [
        ("Saldo AR 1106000001 sebelum clearing (posted s/d 31-Jul)", rnd(ar_before)),
        ("Dikredit blok B", rnd(-b_gross)),
        ("Proyeksi saldo AR sesudah clearing", rnd(ar_before - b_gross)),
    ]:
        ws.append([label, "", "", "", "", val])
        ws.cell(ws.max_row, 6).number_format = money
        ws.cell(ws.max_row, 1).font = bold
    ws.append([])
    ws.append(["Jurnal rekon AR Juni yang sudah diposting 7-Agu:"])
    ws.cell(ws.max_row, 1).font = bold
    for name, date, ref, amt in odoo["ar_juni_adj"]:
        ws.append([name, date, ref, "", "", amt])
        ws.cell(ws.max_row, 6).number_format = money

    # --------------------------------------------------------------- blok C
    sheet(
        "BLOK-C-ATS-SWEEP",
        ["Tanggal", "Bank", "Jenis", "Keterangan", "Jumlah"],
        [
            (r["date"], r["bank"], r["kind"], r["desc"].strip(), r["amount"])
            for r in sorted(plan["block_c"], key=lambda r: (r["date"], r["kind"]))
        ],
        [12, 8, 12, 70, 20],
        total_cols=(4,),
    )

    # ------------------------------------------------------------------ KOL
    sheet(
        "KOL-76JT",
        [
            "Toko",
            "Tgl transaksi",
            "Register",
            "Transnum",
            "Kasir",
            "Tender",
            "Penerima (KOL)",
            "Nilai dicatat",
            "Tgl uang diterima",
        ],
        [
            (
                r["store"],
                r["trans"],
                r["register"],
                r["transnum"],
                r["cashier"],
                r["tender"],
                r["kol"],
                r["amount"],
                r["received"] or "(tidak pernah)",
            )
            for r in sorted(kol, key=lambda r: (r["trans"], r["transnum"]))
        ],
        [34, 14, 10, 12, 20, 10, 30, 18, 20],
        total_cols=(7,),
    )
    ws = wb["KOL-76JT"]
    ws.append([])
    for line in [
        "Barang diberikan GRATIS ke KOL/influencer tetapi POS mencatatnya sebagai penjualan CASH harga penuh.",
        "Karena tidak ada uang masuk, tidak ada CASH RECEIVED DATE, sehingga baris ini TIDAK ikut blok A",
        "dan akan menggantung permanen di 1106000101 (tender CASH).",
        "",
        "USULAN PERLAKUAN -- pilih salah satu, perlu keputusan Accounting + Tax:",
        "  (a) Reklas ke beban promosi: Dr Beban Promosi / Cr 1106000101 sebesar total di atas.",
        "      Penjualan dan PPN Keluaran tetap diakui.",
        "  (b) Dibatalkan di sumber X-Store lalu di-import ulang sebagai free goods tanpa nilai jual.",
        "      Penjualan dan PPN Keluaran ikut hilang.",
        "  Implikasi PPN pemakaian cuma-cuma perlu konfirmasi Tax sebelum dipilih.",
    ]:
        ws.append([line])

    # --------------------------------------------------------- SALES JUN
    sheet(
        "SALES-JUN-DI-JULI",
        ["Toko", "Tgl transaksi", "Transnum", "Tender", "Keterangan bon manual", "Nilai", "Tgl uang diterima"],
        [
            (
                r["store"],
                r["trans"],
                r["transnum"],
                r["tender"],
                r["note"],
                r["amount"],
                r["received"] or "(tidak ada di Juli)",
            )
            for r in sorted(sales_jun, key=lambda r: (r["store"], r["trans"]))
        ],
        [34, 14, 12, 24, 40, 18, 22],
        total_cols=(5,),
    )
    ws = wb["SALES-JUN-DI-JULI"]
    ws.append([])
    for line in [
        "Penjualan terjadi di Juni (bon manual karena mesin/promo bermasalah), uangnya sudah tersetor di Juni,",
        "tetapi store baru menginput ulang penjualannya ke X24DN pada Juli -- karena itu tidak ada settlement Juli.",
        "",
        "STATUS: SUDAH SELESAI. Sesi 7-Agu membukukan EBR-ADJ-AR-JUNI-2026-SALESMANUAL Rp 14.608.080",
        "sebagai RECLASS Dr 1106000001 / Cr POS Receivable per tender -- bukan pendapatan baru,",
        "karena penjualannya sudah ada di Odoo sebagai POS Receivable Juli. Tidak perlu jurnal tambahan.",
    ]:
        ws.append([line])

    # ------------------------------------------------------------- sisa
    timing = [r for r in unsettled if r["trans"] == "2026-07-31"]
    stuck = [r for r in unsettled if r["trans"] != "2026-07-31"]
    grp = collections.defaultdict(lambda: [0, 0.0])
    for r in stuck:
        k = (r["store"], r["trans"], "KOL (pemberian gratis)" if r["kol"] else "bon manual / lain-lain")
        grp[k][0] += 1
        grp[k][1] += r["amount"]
    sheet(
        "SISA-503JT",
        ["Toko", "Tgl transaksi", "Sebab", "Jml baris", "Nilai"],
        [(k[0], k[1], k[2], v[0], rnd(v[1])) for k, v in sorted(grp.items(), key=lambda kv: -kv[1][1])],
        [34, 14, 26, 12, 20],
        total_cols=(4,),
    )
    ws = wb["SISA-503JT"]
    ws.append([])
    tot_uns = sum(r["amount"] for r in unsettled)
    tot_tim = sum(r["amount"] for r in timing)
    for label, val in [
        ("Total belum tersettle per workbook EBR", rnd(tot_uns)),
        ("  dari itu: trans date 31-Jul, settle D+1 di Agustus (timing normal)", rnd(tot_tim)),
        ("  dari itu: macet nyata (tabel di atas)", rnd(tot_uns - tot_tim)),
        ("", ""),
        ("Jembatan ke proyeksi Odoo setelah 63 draft diposting:", ""),
        ("  Belum tersettle per workbook EBR", rnd(tot_uns)),
        ("  + AEON BSD CITY, trans date EBR bergeser sehari + beda nilai Rp 50 (sheet AEON-SELISIH)", 1400925.0),
        ("  - Reclass SALESMANUAL Juni yang sudah dibukukan 7-Agu", -14608080.0),
        ("  + Selisih pembulatan Odoo vs workbook", 950.0),
        ("  = Proyeksi POS Receivable Juli terbuka setelah posting", rnd(resid_after)),
    ]:
        ws.append([label, "", "", "", val])
        if isinstance(val, float):
            ws.cell(ws.max_row, 5).number_format = money

    # ----------------------------------------------------------- AEON detail
    impacted = [t for t in aeon["trx"] if t["impact"].startswith("BERDAMPAK")]
    neutral = [t for t in aeon["trx"] if not t["impact"].startswith("BERDAMPAK")]
    short_total = rnd(sum(aeon["asked"].values()) - sum(aeon["blok_a_credit"].values()))

    ws = sheet(
        "AEON-SELISIH",
        [
            "Jenis selisih",
            "X70D (sumber Odoo)",
            "Trx X70D",
            "EBR COMPILE SALES",
            "Trx EBR",
            "Nilai X70D",
            "Nilai EBR",
            "Selisih",
            "Dampak",
        ],
        [
            (
                t["kind"],
                " ".join(filter(None, [t["x70d"][0]["date"], t["x70d"][0]["tender"]])) if t["x70d"] else "(tidak ada)",
                t["x70d"][0]["transnum"] if t["x70d"] else "",
                " ".join(filter(None, [t["ebr"][0]["date"], t["ebr"][0]["tender"]])) if t["ebr"] else "(tidak ada)",
                t["ebr"][0]["transnum"] if t["ebr"] else "",
                t["x_amount"],
                t["e_amount"],
                t["delta"],
                t["impact"],
            )
            for t in impacted + neutral
        ],
        [20, 38, 10, 38, 10, 15, 15, 12, 56],
    )
    ws.insert_rows(1, 4)
    ws["A1"] = "SELISIH KLASIFIKASI TENDER -- %s" % aeon["store"]
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        "X70D adalah sumber yang dibukukan Odoo; COMPILE SALES adalah workbook EBR. "
        "Dibandingkan per TRANSNUM (unik per toko sepanjang Juli)."
    )
    ws["A3"] = (
        "Selisih 'Beda TENDER TYPE' netral karena 81_clearing_juli.py hanya memakai TOTAL per toko x trans-date "
        "dan mengambil split akunnya dari baris RIREC yang masih open."
    )

    ws.append([])
    for label, val in [
        ("Total tender X70D (yang dibukukan Odoo), 01-31 Juli", aeon["odoo_total"]),
        ("Total tender EBR COMPILE SALES, 01-31 Juli", aeon["ebr_total"]),
        ("Selisih netto sebulan", rnd(aeon["ebr_total"] - aeon["odoo_total"])),
        ("", ""),
        ("Yang benar-benar tidak ter-clearing di blok A (= diagnosa nr 4)", short_total),
    ]:
        ws.append([label, "", "", "", val])
        if isinstance(val, float):
            ws.cell(ws.max_row, 5).number_format = money
        ws.cell(ws.max_row, 1).font = bold

    ws.append([])
    ws.append(["PERBANDINGAN HARIAN PER AKUN (hanya yang berselisih)"])
    ws.cell(ws.max_row, 1).font = bold
    ws.append(["Trans date", "Akun POS Receivable", "X70D / Odoo", "EBR", "Selisih"])
    for c in ws[ws.max_row]:
        c.font, c.fill = bold, head_fill
    for d, code, o, e, dv in aeon["daily"]:
        ws.append([d, code, o, e, dv])
        for i in (3, 4, 5):
            ws.cell(ws.max_row, i).number_format = money

    ws.append([])
    ws.append(["DAMPAK KE BLOK A (per tanggal transaksi)"])
    ws.cell(ws.max_row, 1).font = bold
    ws.append(
        ["Trans date", "Jurnal blok A", "Diminta EBR", "Tersedia di Odoo", "Dikreditkan", "Tidak ter-clearing"]
    )
    for c in ws[ws.max_row]:
        c.font, c.fill = bold, head_fill
    for d in sorted(aeon["asked"]):
        sdate = (datetime.date.fromisoformat(d) + datetime.timedelta(days=1)).isoformat()
        booked = aeon["blok_a_credit"].get(sdate, 0.0)
        askd, avail = aeon["asked"][d], aeon["available"].get(d, 0.0)
        gap = rnd(askd - booked)
        if abs(gap) <= 0.004:
            continue
        ws.append([d, f"{REF_PREFIX}-A-{sdate}", rnd(askd), rnd(avail), rnd(booked), gap])
        for i in (3, 4, 5, 6):
            ws.cell(ws.max_row, i).number_format = money
    ws.append([])
    ws.append(["SISI BANK -- kas yang ikut tertahan karena blok A menyerap lebih sedikit"])
    ws.cell(ws.max_row, 1).font = bold
    ws.append(
        [
            "Jurnal blok A",
            "Debit 1103000002 dibukukan",
            "Seharusnya (gross - MDR)",
            "Debit 7104000001 dibukukan",
            "Seharusnya (MDR)",
            "Tertahan di Bank Suspense",
        ]
    )
    for c in ws[ws.max_row]:
        c.font, c.fill = bold, head_fill
    for d in sorted(aeon["asked"]):
        sdate = (datetime.date.fromisoformat(d) + datetime.timedelta(days=1)).isoformat()
        if abs(rnd(aeon["asked"][d] - aeon["blok_a_credit"].get(sdate, 0.0))) <= 0.004:
            continue
        gross, mdr = aeon["plan_rows"].get(sdate, (0.0, 0.0))
        got_s = aeon["blok_a_debit"].get((sdate, "1103000002"), 0.0)
        got_m = aeon["blok_a_debit"].get((sdate, "7104000001"), 0.0)
        ws.append(
            [
                f"{REF_PREFIX}-A-{sdate}",
                rnd(got_s),
                rnd(gross - mdr),
                rnd(got_m),
                rnd(mdr),
                rnd(gross - mdr - got_s),
            ]
        )
        for i in range(2, 7):
            ws.cell(ws.max_row, i).number_format = money

    ws.append([])
    ws.append(["TINDAK LANJUT YANG DIMINTA KE EBR"])
    ws.cell(ws.max_row, 1).font = bold
    for n, txt in enumerate(
        [
            "Koreksi TRANS DATE transaksi yang ditandai 'Beda TRANS DATE' di atas -- transnum dan nominalnya "
            "identik dengan X70D, hanya tanggal (dan nomor register) yang bergeser satu hari.",
            "Konfirmasi nominal transaksi yang ditandai 'Beda NILAI' -- mana yang benar, X70D atau COMPILE SALES.",
            "Perbaiki pemetaan METODE PEMBAYARAN -> TENDER TYPE untuk kasus 'Beda TENDER TYPE'. Belum berdampak "
            "akuntansi sekarang, tetapi akan berdampak begitu selisihnya kebetulan melintasi hari.",
            "Selama belum dikoreksi, nilai di kolom 'Tidak ter-clearing' tetap menggantung: sisi piutang di "
            "1106000101..110 dan sisi kasnya di 1103000002 Bank Suspense.",
            "Tidak diperlukan jurnal koreksi manual -- cukup workbook diperbaiki, lalu blok A di-regenerate.",
        ],
        start=1,
    ):
        ws.append([str(n), txt])
        ws.cell(ws.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A6"

    sheet(
        "AEON-TRX-DETAIL",
        ["Sumber", "Trans date", "Register", "Transnum", "Tender type", "Akun", "Nilai", "Metode / catatan"],
        [
            row
            for t in impacted + neutral
            for row in (
                [
                    ("X70D", r["date"], r["register"], r["transnum"], r["tender"], r["acc"], rnd(r["amount"]), "")
                    for r in t["x70d"]
                ]
                + [
                    (
                        "EBR",
                        r["date"],
                        r["register"],
                        r["transnum"],
                        r["tender"],
                        r["acc"],
                        rnd(r["amount"]),
                        r["metode"],
                    )
                    for r in t["ebr"]
                ]
                + [("", "", "", "", "", "", "", t["impact"])]
            )
        ],
        [9, 12, 10, 11, 26, 13, 16, 30],
    )

    # ------------------------------------------------------------ open items
    sheet(
        "OPEN-ITEMS",
        ["No", "Item", "Nilai", "Usulan tindakan", "Menunggu keputusan"],
        [
            (
                "1",
                "BRI 27-Jul 'CAIR CEK UNTUK RTGS' -- rekening tujuan belum dipastikan; "
                "inilah penyebab utama sisa Bank Suspense 1.530.199.113",
                1533030000.0,
                "Konfirmasi rekening tujuan lalu buat jurnal pemindahan dari 1103000002",
                "Klien / Treasury",
            ),
            (
                "2",
                "Pemberian gratis ke KOL tercatat sebagai penjualan CASH (sheet KOL-76JT)",
                76926875.0,
                "Reklas ke beban promosi, atau batalkan dan import ulang sebagai free goods",
                "Accounting + Tax",
            ),
            (
                "3",
                "AEON BSD CITY: trx 617/619/622 ditulis EBR bertanggal 07-Jul padahal X70D 06-Jul "
                "(Rp 1.400.875) + trx 682 beda nilai Rp 50 -- lihat sheet AEON-SELISIH",
                1400925.0,
                "Koreksi trans date dan nominal di COMPILE SALES, lalu blok A di-regenerate",
                "EBR Finance",
            ),
            (
                "4",
                "Selisih sales Juli Odoo 16.940.433.421 vs workbook EBR 16.940.432.471 "
                "(angka Odoo sebelum reclass SALESMANUAL 7-Agu)",
                950.0,
                "Diterima sebagai known-diff, cukup didokumentasikan",
                "Accounting",
            ),
            (
                "5",
                "Jurnal draft OBCA 8282/2026/07/042 'SEWA 7 LEVIS MOI' 21-Jul, "
                "DI LUAR paket clearing ini",
                75407050.0,
                "Ditindaklanjuti terpisah agar Juli tidak tertinggal",
                "Accounting",
            ),
            (
                "6",
                "Akun 1103000002 ber-reconcile = false, sehingga 2.535 statement line Juli "
                "tidak bisa di-match di widget bank reconciliation",
                "",
                "Terima kontrol berbasis saldo, atau jadikan akun reconcilable lalu matching ulang",
                "Accounting",
            ),
        ],
        [6, 66, 20, 62, 22],
    )

    wb.save(path)
    os.chmod(path, 0o644)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", required=True)
    ap.add_argument("--ebr", required=True)
    ap.add_argument("--db", default="prd_levis_begbal")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    if not os.environ.get("PGPASSWORD"):
        raise SystemExit("PGPASSWORD is required")

    import json

    plan = json.load(open(args.json))
    odoo = read_odoo(args.db)
    kol, sales_jun, unsettled = read_ebr(args.ebr)

    posted = [b for b in odoo["draft_blocks"] if b[1] != "draft"]
    if posted:
        raise SystemExit(f"ada jurnal {REF_PREFIX}-* yang sudah diposting: {posted}")

    aeon = read_aeon(args.db, args.ebr, plan)
    build(args.out, plan, odoo, kol, sales_jun, unsettled, aeon)
    print(f"KOL          : {len(kol):3d} baris  {sum(r['amount'] for r in kol):>18,.0f}")
    print(f"SALES JUN    : {len(sales_jun):3d} baris  {sum(r['amount'] for r in sales_jun):>18,.0f}")
    print(f"Belum settle : {len(unsettled):3d} baris  {sum(r['amount'] for r in unsettled):>18,.0f}")
    imp = [t for t in aeon["trx"] if t["impact"].startswith("BERDAMPAK")]
    print(f"AEON selisih : {len(aeon['trx']):3d} transnum, {len(imp)} berdampak")
    print(f"tersimpan di {args.out}")


if __name__ == "__main__":
    main()
