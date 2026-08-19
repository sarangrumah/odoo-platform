"""Trace, to the rupiah, the POS receivable that survived the July-2026 clearing.

    PGPASSWORD=... python3 scripts/tenants/levis/100_report_selisih_clearing_juli.py \
        --ebr /srv/sftp-share/files/clearing-juli-2026/EBR_JULI_2026.xlsx \
        --out /srv/sftp-share/files/Rincian_Selisih_Clearing_Juli2026.xlsx

Read-only: plain SELECTs against the live database plus the EBR workbook.

This is the successor of 91_workbook_sisa_pos_receivable_juli.py, which had to
post the drafts inside a rolled-back transaction because the clearing was still
draft when it was written. The 63 journals are posted since 11-Aug-2026, so the
survivors can simply be read, and the bridge closes against the ledger instead
of carrying AEON and the rounding as constants.

The bridge is built per Operating Unit rather than per ledger line on purpose:
reconciliation ran per account across every store, so all 85 surviving lines
carry a 30 or 31 July date no matter when the transaction happened. Reading the
residual per date off the ledger gives the wrong answer.
"""

import argparse
import collections
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from importlib.machinery import SourceFileLoader

_HERE = os.path.dirname(os.path.abspath(__file__))
_PREP = SourceFileLoader("prep_clearing_juli", os.path.join(_HERE, "80_prep_clearing_juli.py")).load_module()
_WB = SourceFileLoader("wb_clearing_juli", os.path.join(_HERE, "82_workbook_approval_clearing_juli.py")).load_module()
psql, num, rnd = _PREP.psql, _PREP.num, _PREP.rnd
read_ebr = _WB.read_ebr

MONTH = ("2026-07-01", "2026-07-31")
TIMING_DATE = "2026-07-31"
AEON_STORE = "OLS SES - AEON BSD CITY"
KOL_STORE = "OLS SES - GRAND INDONESIA"

# 1106000112 is the POS suspense contra, debited and credited on the same day by
# every RIREC entry. It sits inside the 11060001xx range but is not a tender
# account -- leaving it in doubles every store's July debit.
TENDER = "acc.code between '1106000101' and '1106000110'"


def read_bridge(db):
    """Per store: what the POS tender accounts were charged, and what cleared."""
    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select coalesce(ou.name, '(tanpa OU)'),
               sum(l.debit), sum(l.credit),
               sum(l.debit) filter (where l.date = '{TIMING_DATE}')
          from account_move_line l
          join account_move m on m.id = l.move_id
          join acc on acc.id = l.account_id
          left join operating_unit ou on ou.id = l.operating_unit_id
         where m.state = 'posted' and l.date between '{MONTH[0]}' and '{MONTH[1]}'
           and {TENDER}
      group by 1 order by 1
        """,
    )
    out = []
    for store, dr, cr, dr31 in rows:
        sel = num(dr) - num(cr)
        if abs(sel) <= 0.004:
            continue
        out.append(
            {
                "store": store,
                "debit": num(dr),
                "credit": num(cr),
                "selisih": sel,
                "timing": num(dr31),
                "sisa": sel - num(dr31),
            }
        )
    return sorted(out, key=lambda r: -r["selisih"])


def read_timing(db):
    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select coalesce(ou.name, '(tanpa OU)'), acc.code, aa.name->>'en_US', sum(l.debit)
          from account_move_line l
          join account_move m on m.id = l.move_id
          join acc on acc.id = l.account_id
          join account_account aa on aa.id = l.account_id
          left join operating_unit ou on ou.id = l.operating_unit_id
         where m.state = 'posted' and l.debit > 0 and l.date = '{TIMING_DATE}'
           and {TENDER}
      group by 1, 2, 3 order by 1, 2
        """,
    )
    return [(r[0], r[1], r[2], num(r[3])) for r in rows]


def read_ledger(db):
    """The 85 lines a user actually sees when opening the accounts."""
    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select l.date, acc.code, coalesce(ou.name, '(tanpa OU)'),
               l.amount_residual, m.name, l.name
          from account_move_line l
          join account_move m on m.id = l.move_id
          join acc on acc.id = l.account_id
          left join operating_unit ou on ou.id = l.operating_unit_id
         where m.state = 'posted' and l.date <= '{MONTH[1]}'
           and l.full_reconcile_id is null and abs(l.amount_residual) > 0.004
           and {TENDER}
      order by l.date, acc.code, 3
        """,
    )
    return [{"date": r[0], "acc": r[1], "store": r[2], "resid": num(r[3]), "move": r[4], "label": r[5]} for r in rows]


def read_tender_split(db, stores):
    """Per tender account for the stores whose residual is not pure timing."""
    quoted = ", ".join("'" + s.replace("'", "''") + "'" for s in stores)
    rows = psql(
        db,
        f"""
        with acc as (select id, code_store->>'1' code from account_account)
        select coalesce(ou.name, '(tanpa OU)'), acc.code, aa.name->>'en_US',
               sum(l.debit), sum(l.credit),
               sum(l.debit) filter (where l.date = '{TIMING_DATE}')
          from account_move_line l
          join account_move m on m.id = l.move_id
          join acc on acc.id = l.account_id
          join account_account aa on aa.id = l.account_id
          left join operating_unit ou on ou.id = l.operating_unit_id
         where m.state = 'posted' and l.date between '{MONTH[0]}' and '{MONTH[1]}'
           and {TENDER} and ou.name in ({quoted})
      group by 1, 2, 3 order by 1, 2
        """,
    )
    out = []
    for store, code, nm, dr, cr, dr31 in rows:
        sel = num(dr) - num(cr)
        if abs(sel) <= 0.004 and num(dr31) <= 0:
            continue
        out.append((store, code, nm, num(dr), num(cr), sel, num(dr31), sel - num(dr31)))
    return out


def read_aeon(db, ebr_path):
    """AEON per trans-date: what X70D imported vs what the EBR workbook claims."""
    rows = psql(
        db,
        f"""
        select j->>'trans_date', j->>'register', j->>'transnum',
               j->>'tender_type', j->>'tender_amount'
          from retail_import_line l,
               lateral (select l.raw_data_json::jsonb j) x
         where (j->>'store_name') = '{AEON_STORE}'
           and (j->>'trans_date') like '2026-07-%'
           and (j->>'tender_type') is not null
        """,
    )
    x70d = collections.Counter()
    x70d_trx = collections.defaultdict(list)
    for date, register, transnum, tender, amount in rows:
        x70d[date[:10]] += num(amount)
        x70d_trx[date[:10]].append((register, transnum, tender, num(amount)))

    wb = openpyxl.load_workbook(ebr_path, read_only=True, data_only=True)
    ebr_rows = _PREP.sheet_rows(wb, "COMPILE SALES")
    wb.close()
    ebr = collections.Counter()
    ebr_trx = collections.defaultdict(list)
    for r in ebr_rows:
        if str(r.get("STORE NAME") or "").strip() != AEON_STORE:
            continue
        date = str(r.get("TRANS DATE") or "")[:10]
        if not date.startswith("2026-07"):
            continue
        amount = num(r.get("TENDER AMOUNT"))
        ebr[date] += amount
        ebr_trx[date].append(
            (
                str(r.get("REGISTER") or r.get("REGISTER ") or "").strip(),
                str(r.get("TRANSNUM") or "").strip(),
                str(r.get("TENDER TYPE") or "").strip(),
                amount,
            )
        )

    days = sorted(set(x70d) | set(ebr))
    daily = [(d, rnd(x70d[d]), rnd(ebr[d]), rnd(x70d[d] - ebr[d])) for d in days]
    detail = []
    for d in days:
        if abs(x70d[d] - ebr[d]) <= 0.004:
            continue
        for src, bucket in (("X70D (Odoo)", x70d_trx), ("EBR (workbook)", ebr_trx)):
            for register, transnum, tender, amount in sorted(bucket[d], key=lambda t: t[1]):
                detail.append((d, src, register, transnum, tender, amount))
    return daily, detail


def build(path, bridge, timing, ledger, kol, aeon_daily, aeon_detail, split, aug):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    money = "#,##0"

    def sheet(title, headers, rows, widths, total_cols=(), note=None):
        ws = wb.create_sheet(title)
        if note:
            ws.append([note])
            ws["A1"].font = bold
            ws.append([])
        ws.append(list(headers))
        for c in ws[ws.max_row]:
            c.font = bold
            c.fill = head_fill
        head_row = ws.max_row
        for r in rows:
            ws.append(list(r))
        if rows and total_cols:
            total = ["TOTAL"] + [""] * (len(headers) - 1)
            for i in total_cols:
                total[i] = rnd(sum(num(r[i]) for r in rows))
            ws.append(total)
            for c in ws[ws.max_row]:
                c.font = bold
        for row in ws.iter_rows(min_row=head_row + 1):
            for c in row:
                if isinstance(c.value, float):
                    c.number_format = money
        for col, w in zip("ABCDEFGHI", widths):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
        return ws

    total_selisih = sum(r["selisih"] for r in bridge)
    tot_timing = sum(r["timing"] for r in bridge)
    tot_kol = sum(r["amount"] for r in kol)
    aeon_sisa = next((r["sisa"] for r in bridge if r["store"] == AEON_STORE), 0.0)
    kol_sisa = next((r["sisa"] for r in bridge if r["store"] == KOL_STORE), 0.0)
    lain = total_selisih - tot_timing - kol_sisa - aeon_sisa

    # ----------------------------------------------------------- RINGKASAN
    ws = wb.active
    ws.title = "RINGKASAN"
    rows = [
        ("RINCIAN SELISIH CLEARING JULI 2026 - prd_levis_begbal", "", "", ""),
        ("Sumber: buku besar (posted) + workbook EBR Juli. Read-only.", "", "", ""),
        ("", "", "", ""),
        ("KOMPONEN", "PENJELASAN", "RUPIAH", "TINDAK LANJUT"),
        (
            "Timing 31-Jul",
            "Penjualan hari terakhir Juli. Settlement kartu masuk bank D+1, jadi uangnya "
            "baru tiba 1-2 Agustus. Bukan selisih, murni beda waktu.",
            rnd(tot_timing),
            "Tutup sendiri saat clearing Agustus dijalankan.",
        ),
        (
            "KOL Grand Indonesia",
            "%d transaksi 15-Jul, tender CASH, kolom approval berisi 'KOL <nama>'. Barang "
            "diberikan gratis tetapi POS mencatatnya sebagai penjualan tunai harga penuh, "
            "sehingga tidak akan pernah ada uang masuk." % len(kol),
            rnd(kol_sisa),
            "Keputusan klien: reklas ke beban promosi, atau batalkan di X-Store dan "
            "impor ulang sebagai free goods (ada implikasi PPN cuma-cuma).",
        ),
        (
            "AEON BSD City",
            "X70D dan workbook EBR tidak sepakat soal tanggal/register untuk beberapa "
            "transaksi, plus satu transaksi beda Rp 50. AEON hanya punya register 1 "
            "sepanjang Juli, jadi workbook EBR yang keliru -- bukan buku.",
            rnd(aeon_sisa),
            "Koreksi workbook EBR di sisi klien. Buku tidak perlu diubah.",
        ),
        (
            "Selisih kecil antar tender",
            "Sisa pembulatan dan clearing yang berlebih/kurang di level akun tender "
            "(Central Park, Pondok Indah Mall 1, Paris Van Java). Total per toko benar.",
            rnd(lain),
            "Boleh dibiarkan, atau direklas antar akun tender dalam toko yang sama (nol dampak ke total).",
        ),
        ("TOTAL", "POS Receivable Juli yang masih terbuka", rnd(total_selisih), ""),
        ("", "", "", ""),
        ("KONTROL", "", "", ""),
        (
            "Sisa di buku besar",
            "%d baris terbuka pada akun 1106000101..110" % len(ledger),
            rnd(sum(r["resid"] for r in ledger)),
            "",
        ),
        ("Selisih terhadap rincian", "", rnd(sum(r["resid"] for r in ledger) - total_selisih), ""),
        ("", "", "", ""),
        ("CARA MEMBACANYA DI BUKU BESAR", "", "", ""),
        (
            "",
            "Ke-%d baris sisa itu semuanya bertanggal 30 dan 31 Juli -- BUKAN tanggal "
            "penyebabnya. Rekonsiliasi berjalan per akun lintas toko, jadi sisa mengapung "
            "ke baris terakhir yang belum ter-match. Contoh: KOL terjadi 15-Jul tetapi "
            "tampak bertanggal 30-Jul. Karena itu sisa per tanggal TIDAK boleh dibaca "
            "dari buku besar; pakai sheet PER-TOKO." % len(ledger),
            "",
            "",
        ),
        (
            "",
            "Sisa per akun tender pun tidak persis benar: alokasi setoran tunai sempat "
            "menyentuh akun kartu. Total per toko benar, dan itulah yang dipakai di sini.",
            "",
            "",
        ),
        ("", "", "", ""),
        ("KENAPA TIMING BELUM TERTUTUP", "", "", ""),
        (
            "",
            "Clearing Agustus belum diposting (run POSCLR/2026/0010 masih berstatus "
            "'computed'), dan impor bank Agustus belum lengkap: %s. Selama jurnal bank "
            "belum lengkap, clearing Agustus akan melaporkan settlement yang banknya "
            "belum masuk sebagai kurang bayar." % aug,
            "",
            "",
        ),
    ]
    for r in rows:
        ws.append(list(r))
    ws["A1"].font = Font(bold=True, size=13)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, float):
                c.number_format = money
            c.alignment = Alignment(wrap_text=True, vertical="top")
        if row[0].value in (
            "KOMPONEN",
            "TOTAL",
            "KONTROL",
            "CARA MEMBACANYA DI BUKU BESAR",
            "KENAPA TIMING BELUM TERTUTUP",
        ):
            for c in row:
                c.font = bold
    for col, w in zip("ABCD", (32, 82, 20, 58)):
        ws.column_dimensions[col].width = w

    # ------------------------------------------------------------- PER-TOKO
    sheet(
        "PER-TOKO",
        [
            "Toko",
            "Didebet ke piutang POS",
            "Sudah di-clearing",
            "Selisih",
            "di antaranya timing 31-Jul",
            "Sisa di luar timing",
        ],
        [
            (r["store"], rnd(r["debit"]), rnd(r["credit"]), rnd(r["selisih"]), rnd(r["timing"]), rnd(r["sisa"]))
            for r in bridge
        ],
        [42, 24, 24, 20, 24, 22],
        total_cols=(1, 2, 3, 4, 5),
        note="Inilah pembacaan yang benar: selisih per toko, bukan per tanggal buku besar.",
    )

    # -------------------------------------------------------- TIMING-31JUL
    per_store = collections.Counter()
    for store, _code, _nm, val in timing:
        per_store[store] += val
    sheet(
        "TIMING-31JUL",
        ["Toko", "Nilai", "Porsi"],
        [
            (store, rnd(val), round(val / tot_timing, 4) if tot_timing else 0)
            for store, val in sorted(per_store.items(), key=lambda kv: -kv[1])
        ],
        [42, 20, 12],
        total_cols=(1,),
        note="Penjualan 31-Jul yang settle D+1. Cocok per toko, tidak perlu alokasi.",
    )
    ws = wb["TIMING-31JUL"]
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=3):
        for c in row:
            if isinstance(c.value, float):
                c.number_format = "0.0%"

    sheet(
        "TIMING-PER-TENDER",
        ["Toko", "Akun", "Nama akun", "Nilai"],
        [(s, c, nm, rnd(v)) for s, c, nm, v in timing],
        [42, 14, 40, 20],
        total_cols=(3,),
    )

    # ------------------------------------------------------------------ KOL
    sheet(
        "KOL-GRAND-INDONESIA",
        ["Toko", "Tgl transaksi", "Register", "Transnum", "Kasir", "Tender", "Nama KOL", "Nilai"],
        [
            (r["store"], r["trans"], r["register"], r["transnum"], r["cashier"], r["tender"], r["kol"], r["amount"])
            for r in sorted(kol, key=lambda r: (r["trans"], r["transnum"]))
        ],
        [30, 14, 10, 12, 24, 14, 34, 18],
        total_cols=(7,),
        note="Tidak ada 'CASH RECEIVED DATE' pada baris-baris ini: uangnya memang tidak pernah masuk.",
    )

    # ----------------------------------------------------------------- AEON
    sheet(
        "AEON-HARIAN",
        ["Tgl transaksi", "X70D (Odoo)", "EBR (workbook)", "Selisih"],
        [r for r in aeon_daily],
        [16, 20, 20, 18],
        total_cols=(1, 2, 3),
        note="Netto sebulan mendekati nol: yang berbeda adalah tanggal/register, bukan nilainya.",
    )
    sheet(
        "AEON-TRX-DETAIL",
        ["Tgl transaksi", "Sumber", "Register", "Transnum", "Tender", "Nilai"],
        aeon_detail,
        [16, 18, 12, 14, 22, 18],
        note="Hanya hari yang timpang. Jangan cocokkan per TRANSNUM -- EBR sering salah "
        "ketik nomornya; cocokkan tanggal + nilai.",
    )

    # ------------------------------------------------------- selisih kecil
    sheet(
        "SELISIH-KECIL",
        ["Toko", "Akun", "Nama akun", "Debet", "Kredit", "Selisih", "Timing 31-Jul", "Di luar timing"],
        [(s, c, nm, rnd(dr), rnd(cr), rnd(sel), rnd(t), rnd(x)) for s, c, nm, dr, cr, sel, t, x in split],
        [40, 14, 40, 20, 20, 18, 18, 18],
        total_cols=(3, 4, 5, 6, 7),
        note="Toko yang sisanya tidak persis sama dengan penjualan 31-Jul, dibedah per akun tender.",
    )

    # -------------------------------------------------------- ledger detail
    sheet(
        "BARIS-BUKU-BESAR",
        ["Tanggal di buku besar", "Akun", "Toko", "Sisa", "Jurnal asal", "Keterangan"],
        [(r["date"], r["acc"], r["store"], rnd(r["resid"]), r["move"], r["label"]) for r in ledger],
        [22, 14, 40, 20, 24, 60],
        total_cols=(3,),
        note="Tanggal di kolom pertama adalah tanggal baris yang tersisa setelah "
        "rekonsiliasi, BUKAN tanggal penyebabnya. Lihat RINGKASAN.",
    )

    wb.save(path)
    # 0644 + owned by the share user (uid/gid 1002), the account File Browser and
    # SFTP serve /srv/sftp-share/files as. Without it the workbook lands
    # root-owned: readable, but the people who work with it cannot replace it.
    try:
        os.chmod(path, 0o644)
        os.chown(path, 1002, 1002)
    except OSError:
        pass


def august_state(db):
    rows = psql(
        db,
        """
        select j.code, min(m.date), max(m.date), count(*)
          from account_bank_statement_line sl
          join account_move m on m.id = sl.move_id
          join account_journal j on j.id = m.journal_id
         where m.date >= '2026-08-01'
      group by 1 order by 1
        """,
    )
    if not rows:
        return "belum ada satu pun baris rekening koran Agustus"
    return "; ".join(f"{r[0]} {r[3]} baris {r[1]} s/d {r[2]}" for r in rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ebr", required=True)
    ap.add_argument("--db", default="prd_levis_begbal")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    if not os.environ.get("PGPASSWORD"):
        raise SystemExit("PGPASSWORD is required")

    kol, _sales_jun, _unsettled = read_ebr(args.ebr)
    bridge = read_bridge(args.db)
    timing = read_timing(args.db)
    ledger = read_ledger(args.db)
    aeon_daily, aeon_detail = read_aeon(args.db, args.ebr)
    odd = [r["store"] for r in bridge if abs(r["sisa"]) > 0.004 and r["store"] != KOL_STORE]
    split = read_tender_split(args.db, odd) if odd else []
    build(args.out, bridge, timing, ledger, kol, aeon_daily, aeon_detail, split, august_state(args.db))

    total = sum(r["selisih"] for r in bridge)
    print(f"selisih per toko : {total:>18,.0f} pada {len(bridge)} toko")
    print(f"timing 31-Jul    : {sum(r['timing'] for r in bridge):>18,.0f}")
    print(f"KOL              : {sum(r['amount'] for r in kol):>18,.0f} pada {len(kol)} transaksi")
    print(f"buku besar       : {sum(r['resid'] for r in ledger):>18,.0f} pada {len(ledger)} baris")
    print(f"tersimpan di {args.out}")


if __name__ == "__main__":
    main()
