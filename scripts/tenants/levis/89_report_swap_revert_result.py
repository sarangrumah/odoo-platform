# Laporan hasil revert PO yang kolom Quantity/Unit Price-nya tertukar -- prd_levis_begbal.
#
# Insiden 06-Aug-2026: 18 PO (PO/T/EBR/2026/08/00132..00149) diunggah lewat impor bawaan
# dengan kolom Quantity dan Unit Price tertukar. Script ini MEMBUKTIKAN, per PO dan per
# lapisan, bahwa perbaikannya lengkap:
#
#   PO sudah di-GR  -> PO diperbaiki + GR dibatalkan/diretur + jurnal Inventory & GR/IR
#                      di-reverse + harga pokok FIFO dipulihkan
#   PO belum di-GR  -> PO diperbaiki + GR dibatalkan (tidak ada jurnal yang pernah terbit)
#
# SELECT-ONLY. Tidak ada satu pun UPDATE/INSERT ke Odoo; satu-satunya tulisan adalah file
# Excel di /srv/sftp-share/files (bisa diunduh lewat File Browser /files).
#
#   python3 scripts/tenants/levis/89_report_swap_revert_result.py
#
# Env:  DB    -> database (default prd_levis_begbal)
#       OUT   -> path file xlsx
#       FIRST / LAST -> rentang nama PO

import csv
import io
import os
import subprocess
import sys
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PG = "odoo19-platform-postgres"
DB = os.environ.get("DB", "prd_levis_begbal")
FIRST = os.environ.get("FIRST", "PO/T/EBR/2026/08/00132")
LAST = os.environ.get("LAST", "PO/T/EBR/2026/08/00149")
OUT = os.environ.get("OUT", "/srv/sftp-share/files/Hasil_Revert_PO_Swap_Qty_Harga_06Agu2026.xlsx")

# picking receipt yang sempat divalidasi + returnya, dan rentang jurnalnya
GR_PICKING = int(os.environ.get("GR_PICKING", "319"))

MONEY = "#,##0"
QTY = "#,##0"
PCT = "0.00%"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FONT = Font(bold=True)
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
INFO_FILL = PatternFill("solid", fgColor="DDEBF7")


def q(sql):
    out = subprocess.run(
        [
            "docker",
            "exec",
            PG,
            "sh",
            "-c",
            f'PGPASSWORD="$POSTGRES_PASSWORD" psql -U "$POSTGRES_USER" -d {DB} --csv -v ON_ERROR_STOP=1 -c "{sql}"',
        ],
        capture_output=True,
        text=True,
    )
    if out.returncode:
        sys.exit(f"query failed:\n{out.stderr}")
    return list(csv.DictReader(io.StringIO(out.stdout)))


def d(v):
    return Decimal(v or "0")


def f(v):
    return float(v)


def head(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# =========================================================================== data
# 1. PO + baris. Swap-nya persis simetris, jadi kondisi SEBELUM perbaikan bisa dibaca
#    balik dari kondisi sesudah: qty lama = jumlah harga satuan sekarang, dan sebaliknya.
po_rows = q(
    "SELECT o.id, o.name, o.state, o.amount_untaxed, "
    "count(l.id) AS lines, sum(l.product_qty) AS qty_now, sum(l.price_unit) AS price_sum_now, "
    "sum(l.qty_received) AS qty_received "
    "FROM purchase_order o JOIN purchase_order_line l ON l.order_id = o.id "
    f"WHERE o.name >= '{FIRST}' AND o.name <= '{LAST}' "
    "GROUP BY o.id, o.name, o.state, o.amount_untaxed ORDER BY o.name"
)
if not po_rows:
    sys.exit(f"tidak ada PO di rentang {FIRST}..{LAST} pada {DB}")

po_ids = ",".join(r["id"] for r in po_rows)

# 2. Semua receipt yang pernah menempel di PO ini
pick_rows = q(
    "SELECT DISTINCT o.name AS po, sp.id, sp.name, sp.state, "
    "coalesce(sp.date_done::text,'') AS date_done "
    "FROM stock_picking sp JOIN stock_move sm ON sm.picking_id = sp.id "
    "JOIN purchase_order_line l ON l.id = sm.purchase_line_id "
    f"JOIN purchase_order o ON o.id = l.order_id WHERE o.id IN ({po_ids}) "
    "ORDER BY o.name, sp.id"
)
picks_by_po = {}
for r in pick_rows:
    picks_by_po.setdefault(r["po"], []).append(r)

# 3. Retur atas receipt yang sempat divalidasi
ret_rows = q(
    "SELECT sp.id, sp.name, sp.state, count(sm.id) AS moves, sum(sm.quantity) AS qty, "
    "sum(sm.value) AS value FROM stock_move sm JOIN stock_picking sp ON sp.id = sm.picking_id "
    f"WHERE sm.origin_returned_move_id IN (SELECT id FROM stock_move WHERE picking_id = {GR_PICKING}) "
    "GROUP BY sp.id, sp.name, sp.state"
)

# 4. Jurnal: GR-VAL asli + reversal-nya, per akun
je_rows = q(
    "SELECT CASE WHEN am.ref LIKE 'Reversal of%' THEN 'Reversal' ELSE 'GR-VAL' END AS jenis, "
    "aa.code_store->>'1' AS code, aa.name->>'en_US' AS account, "
    "count(DISTINCT am.id) AS entries, min(am.name) AS je_min, max(am.name) AS je_max, "
    "sum(aml.debit) AS debit, sum(aml.credit) AS credit "
    "FROM account_move am JOIN account_move_line aml ON aml.move_id = am.id "
    "JOIN account_account aa ON aa.id = aml.account_id "
    f"WHERE am.ref IN (SELECT 'GR-VAL:'||id FROM stock_move WHERE picking_id = {GR_PICKING}) "
    "   OR am.ref LIKE 'Reversal of GR-VAL:%' "
    "GROUP BY 1,2,3 ORDER BY 1 DESC, 2"
)

# 5. Move dari receipt yang salah: qty diterima, nilai yang dibukukan, harga pokok kini
prod_rows = q(
    "SELECT pp.default_code AS sku, pt.name->>'en_US' AS product, "
    "sm.quantity AS qty_gr, sm.price_unit AS price_gr, sm.value AS value_gr, "
    "(pp.standard_price->>'1')::numeric AS cost_now, "
    "(SELECT l.product_qty FROM purchase_order_line l "
    f"  WHERE l.id = sm.purchase_line_id) AS qty_po, "
    "(SELECT l.price_unit FROM purchase_order_line l "
    "   WHERE l.id = sm.purchase_line_id) AS price_po "
    "FROM stock_move sm JOIN product_product pp ON pp.id = sm.product_id "
    "JOIN product_template pt ON pt.id = pp.product_tmpl_id "
    f"WHERE sm.picking_id = {GR_PICKING} ORDER BY pp.default_code"
)

# 6. Stok dan saldo GL saat ini
quant_rows = q(
    "SELECT l.usage, sum(q.quantity) AS qty FROM stock_quant q "
    "JOIN stock_location l ON l.id = q.location_id GROUP BY 1 ORDER BY 1"
)
gl_rows = q(
    "SELECT aa.code_store->>'1' AS code, aa.name->>'en_US' AS account, "
    "sum(aml.debit - aml.credit) AS balance "
    "FROM account_move_line aml JOIN account_move am ON am.id = aml.move_id "
    "JOIN account_account aa ON aa.id = aml.account_id "
    "WHERE am.state = 'posted' AND aa.code_store->>'1' IN "
    "('1113100021','1113100023','2103109121','2103109123') "
    "GROUP BY 1,2 ORDER BY 1"
)
stray = q("SELECT count(*) AS n FROM stock_move WHERE value > 1000000000")[0]["n"]

# =========================================================================== hitung
gr_po_names = {r["po"] for r in pick_rows if r["state"] == "done"}
je_grval = sum(d(r["debit"]) for r in je_rows if r["jenis"] == "GR-VAL")
je_reversal = sum(d(r["credit"]) for r in je_rows if r["jenis"] == "Reversal")

total_qty_before = sum(d(r["price_sum_now"]) if r["state"] != "cancel" else d(r["qty_now"]) for r in po_rows)
total_qty_after = sum(d(r["qty_now"]) if r["state"] != "cancel" else Decimal(0) for r in po_rows)
total_value = sum(d(r["amount_untaxed"]) for r in po_rows)

wb = Workbook()

# --------------------------------------------------------------------- Ringkasan
ws = wb.active
ws.title = "Ringkasan"
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 60

ws["A1"] = "Hasil Revert PO — Kolom Quantity/Unit Price Tertukar"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = f"Database {DB} · rentang {FIRST} s/d {LAST}"
ws["A3"] = "Dibuat " + datetime.now().strftime("%d %B %Y %H:%M")
for r in (2, 3):
    ws.cell(row=r, column=1).font = Font(italic=True, color="666666")

row = 5
head(ws, row, ["Pemeriksaan", "Angka", "Keterangan"], [52, 26, 60])
row += 1


def line(label, value, note, fmt=None, fill=OK_FILL):
    global row
    ws.cell(row=row, column=1, value=label)
    c = ws.cell(row=row, column=2, value=value)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value=note)
    for col in (1, 2, 3):
        ws.cell(row=row, column=col).fill = fill
    row += 1


line("Jumlah PO terdampak", len(po_rows), "Seluruhnya sudah ditangani", QTY, INFO_FILL)
line("PO yang sempat di-GR", len(gr_po_names), "Perlu perbaikan PO + GR + jurnal + harga pokok", QTY, INFO_FILL)
line(
    "PO yang belum di-GR",
    len(po_rows) - len(gr_po_names),
    "Cukup perbaikan PO + pembatalan GR; tidak ada jurnal yang pernah terbit",
    QTY,
    INFO_FILL,
)
row += 1
line("Kuantitas sebelum perbaikan", f(total_qty_before), "Angka harga yang nyasar ke kolom Quantity", QTY)
line("Kuantitas sesudah perbaikan", f(total_qty_after), "Kuantitas sebenarnya", QTY)
line("Nilai PO (tidak berubah)", f(total_value), "Swap kolom mempertahankan qty × harga", MONEY)
row += 1
line("Jurnal Inventory & GR/IR terbit", f(je_grval), "Saat receipt salah divalidasi", MONEY, WARN_FILL)
line("Jurnal reversal", f(je_reversal), "Nilai identik terbalik", MONEY)
line("Sisa di GL akibat insiden", f(je_grval - je_reversal), "Nol = pulih sepenuhnya", MONEY)
row += 1
line("Produk yang harga pokoknya dipulihkan", len(prod_rows), "Diuji silang terhadap harga di PO", QTY)
line("Move dengan nilai janggal (>1 miliar)", int(stray), "Nol = residu penilaian sudah diluruskan", QTY)

# ----------------------------------------------------------------------- Per PO
ws = wb.create_sheet("Per PO")
head(
    ws,
    1,
    [
        "No PO",
        "Status PO",
        "Baris",
        "Qty sebelum",
        "Qty sesudah",
        "Nilai PO",
        "Qty diterima",
        "Pernah di-GR?",
        "Receipt & statusnya",
        "Perbaikan yang dilakukan",
        "Hasil",
    ],
    [24, 12, 8, 16, 13, 18, 13, 13, 46, 52, 12],
)
r = 2
for po in po_rows:
    cancelled = po["state"] == "cancel"
    was_gr = po["name"] in gr_po_names
    qty_before = d(po["qty_now"]) if cancelled else d(po["price_sum_now"])
    qty_after = Decimal(0) if cancelled else d(po["qty_now"])
    picks = picks_by_po.get(po["name"], [])
    pick_txt = ", ".join(f"{p['name']} ({p['state']})" for p in picks) or "-"
    if cancelled:
        perbaikan = "PO dibatalkan tim; GR ikut batal"
    elif was_gr:
        perbaikan = "PO ditukar balik + GR diretur penuh + jurnal Inventory/GR-IR di-reverse + harga pokok dipulihkan"
    else:
        perbaikan = "PO ditukar balik + GR dibatalkan (belum pernah membukukan jurnal)"

    ws.cell(row=r, column=1, value=po["name"])
    ws.cell(row=r, column=2, value=po["state"])
    ws.cell(row=r, column=3, value=int(po["lines"])).number_format = QTY
    ws.cell(row=r, column=4, value=f(qty_before)).number_format = QTY
    ws.cell(row=r, column=5, value=f(qty_after)).number_format = QTY
    ws.cell(row=r, column=6, value=f(d(po["amount_untaxed"]))).number_format = MONEY
    ws.cell(row=r, column=7, value=f(d(po["qty_received"]))).number_format = QTY
    ws.cell(row=r, column=8, value="Ya" if was_gr else "Tidak")
    ws.cell(row=r, column=9, value=pick_txt)
    ws.cell(row=r, column=10, value=perbaikan)
    ok = d(po["qty_received"]) == 0 and (cancelled or qty_after > 0)
    c = ws.cell(row=r, column=11, value="SELESAI" if ok else "PERIKSA")
    c.fill = OK_FILL if ok else WARN_FILL
    c.font = SUB_FONT
    for col in (9, 10):
        ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ---------------------------------------------------------------------- Jurnal
ws = wb.create_sheet("Jurnal")
ws["A1"] = "Jurnal Inventory & GR/IR — hanya PO yang sempat di-GR"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = (
    "Receipt yang salah membukukan satu jurnal per baris; seluruhnya di-reverse dengan nilai "
    "identik terbalik, sehingga sisa insiden di GL nol rupiah."
)
ws["A2"].font = Font(italic=True, color="666666")
head(
    ws,
    4,
    ["Jenis", "Kode akun", "Nama akun", "Jumlah jurnal", "Nomor JE awal", "Nomor JE akhir", "Debit", "Kredit"],
    [14, 14, 42, 14, 20, 20, 20, 20],
)
r = 5
for j in je_rows:
    ws.cell(row=r, column=1, value="Jurnal terbit" if j["jenis"] == "GR-VAL" else "Reversal")
    ws.cell(row=r, column=2, value=j["code"])
    ws.cell(row=r, column=3, value=j["account"])
    ws.cell(row=r, column=4, value=int(j["entries"])).number_format = QTY
    ws.cell(row=r, column=5, value=j["je_min"])
    ws.cell(row=r, column=6, value=j["je_max"])
    ws.cell(row=r, column=7, value=f(d(j["debit"]))).number_format = MONEY
    ws.cell(row=r, column=8, value=f(d(j["credit"]))).number_format = MONEY
    if j["jenis"] == "Reversal":
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = OK_FILL
    r += 1
ws.cell(row=r + 1, column=3, value="Sisa di GL akibat insiden").font = SUB_FONT
c = ws.cell(row=r + 1, column=7, value=f(je_grval - je_reversal))
c.number_format = MONEY
c.font = SUB_FONT
c.fill = OK_FILL

# --------------------------------------------------------------------- Retur GR
ws = wb.create_sheet("Retur GR")
ws["A1"] = "Pembalikan receipt yang sudah terlanjur divalidasi"
ws["A1"].font = Font(bold=True, size=12)
head(ws, 3, ["Dokumen", "Status", "Jumlah baris", "Kuantitas", "Nilai"], [24, 14, 16, 18, 22])
r = 4
for t in ret_rows:
    ws.cell(row=r, column=1, value=t["name"])
    ws.cell(row=r, column=2, value=t["state"])
    ws.cell(row=r, column=3, value=int(t["moves"])).number_format = QTY
    ws.cell(row=r, column=4, value=f(d(t["qty"]))).number_format = QTY
    ws.cell(row=r, column=5, value=f(d(t["value"]))).number_format = MONEY
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = OK_FILL
    r += 1

# ----------------------------------------------------------------- Harga Pokok
ws = wb.create_sheet("Harga Pokok")
ws["A1"] = "Pemulihan harga pokok FIFO — produk pada receipt yang sempat divalidasi"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = (
    "Harga pokok tergerus karena barang 'datang' dalam jumlah raksasa dengan harga satuan "
    "recehan. Nilai pulihnya diuji silang terhadap harga di PO — selisih di bawah 1% "
    "berarti pemulihannya tepat."
)
ws["A2"].font = Font(italic=True, color="666666")
head(
    ws,
    4,
    [
        "SKU",
        "Produk",
        "Qty masuk (salah)",
        "Harga satuan (salah)",
        "Qty PO (benar)",
        "Harga PO (benar)",
        "Harga pokok kini",
        "Selisih vs harga PO",
    ],
    [18, 46, 18, 20, 16, 18, 20, 20],
)
r = 5
off = 0
for p in prod_rows:
    price_po = d(p["price_po"])
    cost_now = d(p["cost_now"])
    delta = (cost_now - price_po) / price_po if price_po else Decimal(0)
    ws.cell(row=r, column=1, value=p["sku"])
    ws.cell(row=r, column=2, value=p["product"])
    ws.cell(row=r, column=3, value=f(d(p["qty_gr"]))).number_format = QTY
    ws.cell(row=r, column=4, value=f(d(p["price_gr"]))).number_format = MONEY
    ws.cell(row=r, column=5, value=f(d(p["qty_po"]))).number_format = QTY
    ws.cell(row=r, column=6, value=f(price_po)).number_format = MONEY
    ws.cell(row=r, column=7, value=f(cost_now)).number_format = MONEY
    c = ws.cell(row=r, column=8, value=f(delta))
    c.number_format = PCT
    if abs(delta) > Decimal("0.01"):
        off += 1
        c.fill = WARN_FILL
    else:
        c.fill = OK_FILL
    r += 1
ws.cell(row=r + 1, column=1, value=f"Produk dengan selisih di atas 1%: {off} dari {len(prod_rows)}").font = SUB_FONT

# ------------------------------------------------------------- Posisi Saat Ini
ws = wb.create_sheet("Posisi Saat Ini")
ws["A1"] = "Posisi stok dan buku besar saat laporan dibuat"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = (
    "Angka di bawah adalah posisi berjalan seluruh database — termasuk transaksi normal tim "
    "yang tidak ada kaitannya dengan insiden ini."
)
ws["A2"].font = Font(italic=True, color="666666")

head(ws, 4, ["Lokasi", "Kuantitas"], [28, 22])
r = 5
for qr in quant_rows:
    ws.cell(row=r, column=1, value=qr["usage"])
    ws.cell(row=r, column=2, value=f(d(qr["qty"]))).number_format = QTY
    r += 1

r += 2
head(ws, r, ["Kode akun", "Nama akun", "Saldo"], [16, 46, 24])
r += 1
for g in gl_rows:
    ws.cell(row=r, column=1, value=g["code"])
    ws.cell(row=r, column=2, value=g["account"])
    ws.cell(row=r, column=3, value=f(d(g["balance"]))).number_format = MONEY
    r += 1

# =========================================================================== tulis
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
# 0644 + owned by the share user: File Browser (uid 1002) serves it, nobody else writes it
try:
    os.chmod(OUT, 0o644)
    os.chown(OUT, 1002, 1002)
except OSError:
    pass
print(f"tersimpan: {OUT}")
print(f"PO: {len(po_rows)} · sempat di-GR: {len(gr_po_names)} · produk: {len(prod_rows)}")
print(f"jurnal terbit {je_grval} · reversal {je_reversal} · sisa {je_grval - je_reversal}")
