# Skenario detail 4 insiden pembayaran Juli 2026 -- prd_levis_begbal.
#
# SELECT-ONLY. Tidak ada satu pun UPDATE/INSERT ke Odoo; satu-satunya tulisan adalah
# file Excel di /srv/sftp-share/files (bisa diunduh lewat File Browser /files).
#
#   python3 scripts/tenants/levis/92_report_payment_incident_scenarios.py
#
# Env:  DB   -> database (default prd_levis_begbal)
#       OUT  -> path file xlsx
#
# --------------------------------------------------------------------------
# Apa yang dibuktikan workbook ini
# --------------------------------------------------------------------------
# Empat jurnal yang dilaporkan Accounting sebagai "AP minus" / "selisih TB vs Aging"
# ternyata satu akar masalah, bukan empat kesalahan terpisah:
#
#   Di Odoo 19, RESET TO DRAFT tidak lagi membatalkan rekonsiliasi.
#   `account.move.button_draft()` hanya menghapus analytic line, mengubah state, dan
#   melepas attachment -- panggilan `remove_move_reconcile()` yang ada di versi lama
#   sudah hilang. Penjaganya, `account.move.line._check_reconciliation()`, masih ada
#   di source tapi TIDAK PERNAH DIPANGGIL oleh apa pun, dan hanya mencakup baris
#   posted. Jadi begitu jurnal ditarik ke draft, akun pada baris yang masih
#   ter-reconcile bisa diganti bebas tanpa peringatan, dan matching-nya hancur.
#
# Keempatnya dibuat lewat Register Payment dari tagihan yang BENAR -- terbukti dari
# kolom `ref` saat pembuatan, yang diisi otomatis Odoo dengan nomor tagihan vendor.
# Yang terjadi sesudahnya (reset to draft, ganti akun, posting ulang) yang merusak.
#
# Kronologi di sheet "Kronologi" diambil dari mail_message + mail_tracking_value, jadi
# ini catatan Odoo sendiri, bukan rekonstruksi. Catatan penting: Odoo melacak perubahan
# `state`, `ref`, `account_id`, tetapi TIDAK melacak rekonsiliasi -- aksi unreconcile
# manual tidak meninggalkan jejak sama sekali.

import csv
import io
import os
import subprocess
import sys
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PG = "odoo19-platform-postgres"
DB = os.environ.get("DB", "prd_levis_begbal")
OUT = os.environ.get(
    "OUT", "/srv/sftp-share/files/Skenario_Insiden_Pembayaran_Juli_2026.xlsx"
)

MOVES = ("8282/2026/07/009", "8282/2026/07/016", "8282/2026/07/017",
         "8282/2026/07/042", "8282/2026/07/045")

MONEY = "#,##0.00"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FONT = Font(bold=True)
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
INFO_FILL = PatternFill("solid", fgColor="DDEBF7")
NOTE_FONT = Font(italic=True, color="595959")
WRAP = Alignment(wrap_text=True, vertical="top")


def q(sql):
    out = subprocess.run(
        [
            "docker", "exec", PG, "sh", "-c",
            'PGPASSWORD="$POSTGRES_PASSWORD" psql -U "$POSTGRES_USER" '
            f"-d {DB} --csv -v ON_ERROR_STOP=1 -c \"{sql}\"",
        ],
        capture_output=True, text=True,
    )
    if out.returncode:
        sys.exit(f"query failed:\n{out.stderr}")
    return list(csv.DictReader(io.StringIO(out.stdout)))


def d(v):
    return Decimal(v or "0")


def head(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def note(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font, c.alignment = NOTE_FONT, WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 15 * (1 + len(text) // (span * 17))


IN_MOVES = ", ".join("'%s'" % m for m in MOVES)

# --------------------------------------------------------------------------- data
kepala = q(
    "select am.name as mv, am.state, am.date, am.ref, aj.name->>'en_US' as jurnal, "
    "       rp.name as partner, cu.login as dibuat_oleh, am.create_date, "
    "       ap.id as payment_id, ap.amount, ap.is_reconciled "
    "from account_move am "
    "join account_journal aj on aj.id = am.journal_id "
    "left join res_partner rp on rp.id = am.partner_id "
    "left join res_users cu on cu.id = am.create_uid "
    "left join account_payment ap on ap.id = am.origin_payment_id "
    f"where am.name in ({IN_MOVES}) order by am.create_date"
)

jejak = q(
    "select mm.date, u.login, "
    "       coalesce(am.name, 'payment#' || mm.res_id::text) as mv, "
    "       mm.model, coalesce(imf.name, '(dibuat)') as field, "
    "       coalesce(mtv.old_value_char, mtv.old_value_integer::text, '') as dari, "
    "       coalesce(mtv.new_value_char, mtv.new_value_integer::text, '') as jadi "
    "from mail_message mm "
    "left join res_users u on u.id = mm.create_uid "
    "left join mail_tracking_value mtv on mtv.mail_message_id = mm.id "
    "left join ir_model_fields imf on imf.id = mtv.field_id "
    "left join account_move am on (mm.model = 'account.move' and am.id = mm.res_id) "
    "   or (mm.model = 'account.payment' and am.origin_payment_id = mm.res_id) "
    f"where (mm.model = 'account.move' and am.name in ({IN_MOVES})) "
    "   or (mm.model = 'account.payment' and mm.res_id in ("
    f"        select origin_payment_id from account_move where name in ({IN_MOVES}))) "
    "order by mm.date, mm.id"
)

dampak = q(
    "select am.name as mv, aa.code_store->>'1' as akun, aa.name->>'en_US' as nama_akun, "
    "       aml.debit, aml.credit, aml.amount_residual as resid, aml.reconciled "
    "from account_move am "
    "join account_move_line aml on aml.move_id = am.id "
    "join account_account aa on aa.id = aml.account_id "
    f"where am.name in ({IN_MOVES}) "
    "order by am.name, aa.code_store->>'1'"
)

# Tagihan yang SEHARUSNYA dilunasi -- terbaca dari ref asli di jejak audit.
tagihan = q(
    "select am.name as bill, am.ref, rp.name as partner, aml.credit, "
    "       aml.amount_residual as resid, aml.reconciled "
    "from account_move am "
    "join account_move_line aml on aml.move_id = am.id "
    "join account_account aa on aa.id = aml.account_id "
    "left join res_partner rp on rp.id = am.partner_id "
    "where aa.account_type = 'liability_payable' and am.name in ("
    "  'BILL/NT/EBR/2026/07/00086','BILL/NT/EBR/2026/07/00087','BILL/NT/EBR/2026/07/00088',"
    "  'BILL/NT/EBR/2026/07/00005','BILL/NT/EBR/2026/07/00006',"
    "  'BILL/NT/EBR/2026/07/00014','BILL/NT/EBR/2026/07/00040') "
    "order by am.name"
)

kepala_map = {r["mv"]: r for r in kepala}

# --------------------------------------------------------------------------- narasi
# Nilai dan tanggal dibaca dari database; sebabnya dari jejak audit di sheet Kronologi.
KASUS = [
    {
        "mv": "8282/2026/07/009",
        "tagihan": "BILL/NT/EBR/2026/07/00086, 00087, 00088",
        "ref_asli": "DM2026070044, DM2026070478, DM2026070479",
        "sebab": (
            "Dibuat lewat Register Payment dari 3 tagihan, tetapi LAHIR DI JURNAL PETTY "
            "CASH sebagai PCSH1/2026/00001 dengan akun kas 1102000001 Cash on hand. "
            "Operator lalu reset to draft dan mengganti akun kas ke BCA-2687778282, "
            "mengganti akun hutang 2103300001 -> 2103100001 lalu balik lagi, dengan 4 kali "
            "siklus draft/posting. Rekonsiliasinya hancur di tengah jalan dan tidak pernah pulih."
        ),
        "akibat": (
            "Pembayaran posted tanpa menerapkan tagihan apa pun. Tagihan 00086/87/88 "
            "kembali terbuka, pembayarannya menggantung sebagai AP minus. Aged payable "
            "menampilkan keduanya sebagai baris terpisah yang tidak saling meniadakan."
        ),
        "status": "SUDAH DIPERBAIKI oleh skrip 90 (direkonsiliasi ke tagihan aslinya)",
        "keputusan": "-",
    },
    {
        "mv": "8282/2026/07/017",
        "tagihan": "BILL/NT/EBR/2026/07/00014",
        "ref_asli": "N260702567",
        "sebab": (
            "Dibuat lewat Register Payment dari tagihan 00014. Ref diubah jadi "
            "'SEWA 2 SC PL JUL 26 LEVIS PVJ BDG', lalu satu siklus reset to draft dan "
            "posting ulang. Satu siklus saja sudah cukup menghapus matching-nya."
        ),
        "akibat": (
            "Sama seperti 009: tagihan kembali terbuka, pembayaran menggantung sebagai "
            "AP minus di aging."
        ),
        "status": "SUDAH DIPERBAIKI oleh skrip 90 (direkonsiliasi ke tagihan aslinya)",
        "keputusan": "-",
    },
    {
        "mv": "8282/2026/07/016",
        "tagihan": "BILL/NT/EBR/2026/07/00005, 00006",
        "ref_asli": "00821/MTLA/MMB/INV-DRT/VII/26, 00822/MTLA/MMB/INV-DSC/VII/26",
        "sebab": (
            "Dibuat lewat Register Payment dari tagihan 00005 + 00006 (111.492.000 + "
            "31.464.000 = 142.956.000). Ref/memo lalu diubah jadi 'DEPOSIT SEWA DAN "
            "DEPOSIT SC LEVIS MM BEKASI' dan ditambah rekening vendor. Jejak audit TIDAK "
            "menunjukkan reset to draft -- jadi matching-nya kemungkinan dilepas manual "
            "(aksi unreconcile tidak terlacak Odoo). Karena tidak menerapkan tagihan, "
            "status kedua bill tetap 'Not Paid'."
        ),
        "akibat": (
            "Keesokan harinya pembayaran yang sama diulang sebagai 8282/2026/07/045 lewat "
            "Register Payment, dan ITU yang ter-reconcile. Di pembukuan kas keluar DUA "
            "KALI untuk satu kewajiban. Yang pertama (016) menggantung sebagai AP minus."
        ),
        "status": "BELUM DIPERBAIKI -- menunggu konfirmasi",
        "keputusan": (
            "Butuh rekening koran BCA 2687778282 tanggal 14/07/2026. Kalau bank hanya "
            "terdebet SEKALI: 016 dijurnal balik. Kalau terdebet DUA KALI: ada kelebihan "
            "bayar ke vendor, 016 direklas ke piutang/uang muka, bukan dijurnal balik."
        ),
    },
    {
        "mv": "8282/2026/07/042",
        "tagihan": "BILL/NT/EBR/2026/07/00040",
        "ref_asli": "RNT/MJS/0726/006",
        "sebab": (
            "Dibuat lewat Register Payment dari tagihan 00040 pukul 10:05:10, langsung "
            "posted dan reconciled. Enam detik kemudian (10:05:16) di-reset to draft, "
            "lalu ref diubah, dan TIDAK PERNAH diposting ulang. Karena Odoo 19 tidak "
            "membatalkan rekonsiliasi saat reset to draft, partial reconcile-nya tetap hidup."
        ),
        "akibat": (
            "Tagihan 00040 tetap terbaca LUNAS (residual 0) melawan jurnal yang berstatus "
            "draft -- artinya di luar Trial Balance yang hanya membaca posted. TB mencatat "
            "hutangnya, aging tidak. Selisih persis sebesar nilai ini, tanpa peringatan apa pun."
        ),
        "status": "BELUM DIPERBAIKI -- menunggu konfirmasi",
        "keputusan": (
            "Butuh rekening koran BCA 2687778282 tanggal 21/07/2026. Kalau uang benar "
            "keluar: posting jurnalnya. Kalau tidak: batalkan rekonsiliasinya lalu hapus, "
            "dan tagihan sewa MOI kembali muncul sebagai hutang terbuka."
        ),
    },
]

PENCEGAHAN = [
    (
        "Reset to draft membatalkan rekonsiliasi",
        "custom_account_reconcile / account_move.py -> button_draft()",
        "Sebelum jurnal turun ke draft, rekonsiliasinya dibatalkan lebih dulu (perilaku "
        "Odoo versi lama), dan dokumen yang dilepas dicatat di chatter.",
        "042 (tagihan tidak lagi bisa terbaca lunas melawan jurnal draft), dan mencegah "
        "pengulangan 009/017.",
    ),
    (
        "Akun / partner baris ter-reconcile tidak bisa diganti",
        "custom_account_reconcile / account_move_line.py -> write()",
        "Mengganti account_id atau partner_id pada baris yang masih memegang partial "
        "ditolak, di state apa pun. Nominal sengaja tidak dijaga supaya selisih kurs dan "
        "write-off tetap jalan.",
        "009 (pergantian akun kas dan akun hutang saat draft yang menghancurkan matching).",
    ),
    (
        "Pembayaran kembar ditolak saat posting",
        "custom_account_reconcile / account_payment.py -> action_post()",
        "Partner + nominal + tanggal + tipe yang sama dan sudah ada yang posted -> posting "
        "ditolak dengan menyebut nomor kembarannya. Ada centang 'Duplicate Checked' untuk "
        "pembayaran kedua yang memang sah.",
        "016 vs 045 (dua cash-out Rp 142.957.500 di tanggal yang sama untuk satu kewajiban).",
    ),
    (
        "Daftar pantau Unapplied Payments",
        "Accounting > Reconciliation > Unapplied Payments",
        "Semua pembayaran posted yang tidak menerapkan tagihan apa pun, bisa difilter dan "
        "digrup per vendor/jurnal/tanggal. Field is_unapplied tersimpan sehingga bisa dicari.",
        "Ketiga-tiganya -- ketahuan harian, bukan menunggu rekonsiliasi aging akhir bulan.",
    ),
]

# --------------------------------------------------------------------------- workbook
wb = Workbook()

# ---- Sheet 1: Ringkasan
ws = wb.active
ws.title = "Ringkasan"
ws["A1"] = "Insiden pembayaran Juli 2026 — 4 kasus, 1 akar masalah"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = f"{DB} — disusun dari jejak audit Odoo (mail_message + mail_tracking_value)"
ws["A2"].font = NOTE_FONT

ws["A4"] = "AKAR MASALAH"
ws["A4"].font = SUB_FONT
akar = (
    "Di Odoo 19, RESET TO DRAFT tidak lagi membatalkan rekonsiliasi. "
    "account.move.button_draft() hanya menghapus analytic line, mengubah state, dan melepas "
    "attachment — panggilan remove_move_reconcile() yang ada di versi lama sudah hilang. "
    "Penjaganya, account.move.line._check_reconciliation(), masih ada di source Odoo tetapi "
    "TIDAK PERNAH DIPANGGIL oleh apa pun, dan hanya mencakup baris posted. Akibatnya begitu "
    "jurnal ditarik ke draft, akun pada baris yang masih ter-reconcile bisa diganti bebas "
    "tanpa peringatan. Keempat pembayaran di bawah dibuat lewat Register Payment dari tagihan "
    "yang BENAR — terbukti dari kolom ref saat pembuatan, yang diisi otomatis Odoo dengan "
    "nomor tagihan vendor. Yang terjadi sesudahnya yang merusak, bukan salah pilih tagihan."
)
c = ws.cell(row=5, column=1, value=akar)
c.alignment, c.fill = WRAP, INFO_FILL
ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
ws.row_dimensions[5].height = 100

head(ws, 7, ["Jurnal", "Tanggal", "Partner", "Nilai", "Dibuat oleh", "Tagihan terkait",
             "Status"], [20, 12, 34, 18, 30, 42, 46])
r = 8
for k in KASUS:
    h = kepala_map.get(k["mv"], {})
    ws.cell(row=r, column=1, value=k["mv"])
    ws.cell(row=r, column=2, value=h.get("date"))
    ws.cell(row=r, column=3, value=h.get("partner"))
    ws.cell(row=r, column=4, value=float(d(h.get("amount")))).number_format = MONEY
    ws.cell(row=r, column=5, value=h.get("dibuat_oleh"))
    ws.cell(row=r, column=6, value=k["tagihan"]).alignment = WRAP
    c = ws.cell(row=r, column=7, value=k["status"])
    c.alignment = WRAP
    c.fill = OK_FILL if "SUDAH" in k["status"] else WARN_FILL
    r += 1

r += 1
head(ws, r, ["Jurnal", "Apa yang terjadi", "Akibatnya", "Keputusan yang ditunggu"],
     [20, 78, 60, 60])
r += 1
for k in KASUS:
    ws.cell(row=r, column=1, value=k["mv"])
    for col, key in ((2, "sebab"), (3, "akibat"), (4, "keputusan")):
        cell = ws.cell(row=r, column=col, value=k[key])
        cell.alignment = WRAP
    if "Butuh" in k["keputusan"]:
        ws.cell(row=r, column=4).fill = WARN_FILL
    ws.row_dimensions[r].height = 92
    r += 1

# ---- Sheet 2: Kronologi
ws2 = wb.create_sheet("Kronologi")
ws2["A1"] = "Jejak audit Odoo — apa adanya, urut waktu"
ws2["A1"].font = Font(bold=True, size=12)
head(ws2, 3, ["Waktu", "User", "Jurnal", "Model", "Field", "Dari", "Jadi"],
     [21, 30, 20, 18, 17, 46, 46])
r = 4
warna = {}
for x in jejak:
    ws2.cell(row=r, column=1, value=(x["date"] or "")[:19])
    ws2.cell(row=r, column=2, value=x["login"])
    ws2.cell(row=r, column=3, value=x["mv"])
    ws2.cell(row=r, column=4, value=x["model"])
    c = ws2.cell(row=r, column=5, value=x["field"])
    ws2.cell(row=r, column=6, value=x["dari"])
    ws2.cell(row=r, column=7, value=x["jadi"])
    # Sorot dua jenis perubahan yang merusak rekonsiliasi.
    if x["field"] == "state" and (x["jadi"] or "").lower() == "draft":
        for col in range(1, 8):
            ws2.cell(row=r, column=col).fill = WARN_FILL
    elif x["field"] == "account_id":
        for col in range(1, 8):
            ws2.cell(row=r, column=col).fill = INFO_FILL
    r += 1
r += 1
note(
    ws2, r,
    "Baris oranye = reset to draft (di Odoo 19 rekonsiliasinya tetap hidup). Baris biru = "
    "pergantian akun pada baris jurnal (saat draft tidak ada penjagaan sama sekali). "
    "Odoo melacak perubahan state, ref, dan account_id, tetapi TIDAK melacak rekonsiliasi — "
    "aksi unreconcile manual tidak meninggalkan jejak apa pun, itulah kenapa 016 tidak "
    "menunjukkan penyebab yang eksplisit.", 7,
)

# ---- Sheet 3: Dampak GL
ws3 = wb.create_sheet("Dampak GL")
ws3["A1"] = "Isi jurnal keempat kasus (plus 045 sebagai pembanding)"
ws3["A1"].font = Font(bold=True, size=12)
head(ws3, 3, ["Jurnal", "State", "Akun", "Nama Akun", "Debit", "Credit", "Residual",
              "Reconciled"], [20, 10, 14, 34, 18, 18, 18, 12])
r = 4
for x in dampak:
    h = kepala_map.get(x["mv"], {})
    ws3.cell(row=r, column=1, value=x["mv"])
    c = ws3.cell(row=r, column=2, value=h.get("state"))
    if h.get("state") != "posted":
        c.fill = WARN_FILL
    ws3.cell(row=r, column=3, value=x["akun"])
    ws3.cell(row=r, column=4, value=x["nama_akun"])
    for col, key in ((5, "debit"), (6, "credit"), (7, "resid")):
        ws3.cell(row=r, column=col, value=float(d(x[key]))).number_format = MONEY
    ws3.cell(row=r, column=8, value="ya" if x["reconciled"] == "t" else "tidak")
    if d(x["resid"]):
        ws3.cell(row=r, column=7).fill = WARN_FILL
    r += 1

r += 2
ws3.cell(row=r, column=1, value="Tagihan yang terkait").font = SUB_FONT
r += 1
head(ws3, r, ["Bill", "Reference vendor", "Partner", "Nilai", "Residual", "Reconciled"],
     [30, 34, 34, 18, 18, 12])
r += 1
for x in tagihan:
    ws3.cell(row=r, column=1, value=x["bill"])
    ws3.cell(row=r, column=2, value=x["ref"])
    ws3.cell(row=r, column=3, value=x["partner"])
    ws3.cell(row=r, column=4, value=float(d(x["credit"]))).number_format = MONEY
    c = ws3.cell(row=r, column=5, value=float(d(x["resid"])))
    c.number_format = MONEY
    c.fill = WARN_FILL if d(x["resid"]) else OK_FILL
    ws3.cell(row=r, column=6, value="ya" if x["reconciled"] == "t" else "tidak")
    r += 1

# ---- Sheet 4: Pencegahan
ws4 = wb.create_sheet("Pencegahan")
ws4["A1"] = "Pencegahan yang dibangun — modul custom_account_reconcile 19.0.3.0.0"
ws4["A1"].font = Font(bold=True, size=12)
ws4["A2"] = "Kode sudah selesai dan lulus tes. BELUM di-deploy; menunggu review."
ws4["A2"].font = NOTE_FONT
head(ws4, 4, ["Pencegahan", "Letak", "Cara kerja", "Menutup kasus"],
     [46, 56, 74, 62])
r = 5
for nama, letak, cara, menutup in PENCEGAHAN:
    ws4.cell(row=r, column=1, value=nama).alignment = WRAP
    ws4.cell(row=r, column=2, value=letak).alignment = WRAP
    ws4.cell(row=r, column=3, value=cara).alignment = WRAP
    ws4.cell(row=r, column=4, value=menutup).alignment = WRAP
    ws4.row_dimensions[r].height = 74
    r += 1
r += 1
note(
    ws4, r,
    "Catatan: penjagaan 'tagihan yang sudah lunas tidak bisa dipilih lagi' sudah ada di "
    "Odoo sejak awal — wizard Register Payment melewati baris dengan residual nol dan "
    "menolak kalau tidak ada sisa (account/wizard/account_payment_register.py:969-975). "
    "Penjagaan itu tidak akan menolong di sini, karena keempat pembayaran memang sudah "
    "memilih tagihan yang benar; matching-nya yang hilang sesudahnya.", 4,
)

wb.save(OUT)
os.chmod(OUT, 0o644)
# File Browser serves /srv/sftp-share/files as sftpshare:sftpusers; a root-owned
# drop lands there readable but out of place. Best effort -- ignore if the ids
# do not exist on this host.
try:
    import grp
    import pwd

    os.chown(OUT, pwd.getpwnam("sftpshare").pw_uid, grp.getgrnam("sftpusers").gr_gid)
except (KeyError, PermissionError, OSError):
    pass

print(f"Database   : {DB}")
print(f"Kasus      : {len(KASUS)} + 1 pembanding (045)")
print(f"Jejak audit: {len(jejak)} baris")
print(f"Baris GL   : {len(dampak)}   tagihan terkait: {len(tagihan)}")
print(f"Tersimpan  : {OUT}")
