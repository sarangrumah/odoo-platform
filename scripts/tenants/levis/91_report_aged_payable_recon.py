# Laporan rekonsiliasi Aged Payable vs Trial Balance -- prd_levis_begbal.
#
# SELECT-ONLY. Tidak ada satu pun UPDATE/INSERT ke Odoo; satu-satunya tulisan adalah
# file Excel di /srv/sftp-share/files (bisa diunduh lewat File Browser /files).
#
#   python3 scripts/tenants/levis/91_report_aged_payable_recon.py
#
# Env:  DB       -> database (default prd_levis_begbal)
#       DATE_TO  -> cut-off (default 2026-07-31)
#       OUT      -> path file xlsx
#
# --------------------------------------------------------------------------
# Menjawab tiga pertanyaan Accounting atas Aged Payable per 31/07/2026
# --------------------------------------------------------------------------
# 1. Kenapa total Aging tidak sama dengan Trial Balance.
#    Dua sebab, dan penjumlahannya cocok persis:
#      C1  Report memakai nilai residual HIDUP, bukan posisi per tanggal cut-off.
#          Tagihan yang masih terbuka per 31 Juli tapi dibayar di Agustus HILANG
#          total dari aging Juli -- karena itu angkanya mengecil terus tiap kali
#          report dibuka ulang. Sudah diperbaiki di custom_accounting_reports
#          (_residual_as_of), sheet ini membuktikan selisihnya.
#      C2  Satu jurnal DRAFT (8282/2026/07/042) masih memegang rekonsiliasi ke bill
#          posted BILL/NT/EBR/2026/07/00040. Bill hilang dari aging, TB masih
#          mencatatnya. Butuh keputusan Finance.
# 2. Kenapa nomor bill dan nomor pembayarannya tidak saling meniadakan.
#    Karena keduanya memang tidak pernah direkonsiliasi -- lihat sheet
#    "Payment Menggantung". Diperbaiki oleh skrip 90.
# 3. Baris EBR-GL. Upload beginning balance memuat sisi tagihan dan sisi bank
#    dua-duanya ke akun payable tanpa saling direkonsiliasi; net-nya persis nol.
#    Sheet "EBR-GL" membuktikannya. Diperbaiki oleh skrip 90 (bukan disembunyikan
#    dari report, supaya aging tetap bisa tie-out ke TB).

import csv
import io
import os
import subprocess
import sys
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PG = "odoo19-platform-postgres"
DB = os.environ.get("DB", "prd_levis_begbal")
DATE_TO = os.environ.get("DATE_TO", "2026-07-31")
OUT = os.environ.get(
    "OUT", "/srv/sftp-share/files/Rekonsiliasi_Aged_Payable_%s.xlsx" % DATE_TO
)

MONEY = "#,##0.00"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FONT = Font(bold=True)
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
NOTE_FONT = Font(italic=True, color="595959")


def q(sql):
    """Run a read-only query and return a list of dicts."""
    out = subprocess.run(
        [
            "docker", "exec", PG, "sh", "-c",
            'PGPASSWORD="$POSTGRES_PASSWORD" psql -U "$POSTGRES_USER" '
            f"-d {DB} --csv -v ON_ERROR_STOP=1 -c \"{sql}\"",
        ],
        capture_output=True, text=True,
    )
    if out.returncode:
        sys.exit(f"query failed:\n{out.stderr}")
    return list(csv.DictReader(io.StringIO(out.stdout)))


def d(v):
    return Decimal(v or "0")


def head(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def note(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


# --------------------------------------------------------------------------- data
# Basis payable yang sama dipakai di semua query.
PAYABLE = (
    "join account_account aa on aa.id = aml.account_id "
    "join account_move am on am.id = aml.move_id "
    "where aa.account_type = 'liability_payable' and am.state = 'posted' "
    f"and aml.date <= '{DATE_TO}'"
)

tb = q(
    "select aa.code_store->>'1' as acct, max(aa.name->>'en_US') as nm, "
    "sum(aml.balance) as v from account_move_line aml " + PAYABLE + " group by 1 order by 1"
)

# Apa yang report LAMA hitung: residual hidup atas baris yang belum reconciled.
lama = q(
    "select aa.code_store->>'1' as acct, sum(aml.amount_residual) as v "
    "from account_move_line aml " + PAYABLE + " and aml.reconciled = false group by 1 order by 1"
)

# Apa yang report BARU hitung: balance dikoreksi partial yang match <= cut-off.
baru = q(
    "select aa.code_store->>'1' as acct, sum(aml.balance "
    "  + coalesce((select sum(p.amount) from account_partial_reconcile p "
    f"      where p.credit_move_id = aml.id and p.max_date <= '{DATE_TO}'), 0) "
    "  - coalesce((select sum(p.amount) from account_partial_reconcile p "
    f"      where p.debit_move_id = aml.id and p.max_date <= '{DATE_TO}'), 0)) as v "
    "from account_move_line aml " + PAYABLE + " group by 1 order by 1"
)

tb_m = {r["acct"]: d(r["v"]) for r in tb}
lama_m = {r["acct"]: d(r["v"]) for r in lama}
baru_m = {r["acct"]: d(r["v"]) for r in baru}
nama_m = {r["acct"]: r["nm"] for r in tb}

# C2 -- jurnal belum posted yang masih memegang rekonsiliasi.
draft = q(
    "select am.name as mv, am.state, am.date, aml.id as lid, aml.debit, "
    "       am2.name as lawan, am2.state as lawan_state, p.amount "
    "from account_move_line aml "
    "join account_move am on am.id = aml.move_id "
    "join account_partial_reconcile p "
    "  on p.debit_move_id = aml.id or p.credit_move_id = aml.id "
    "join account_move_line aml2 on aml2.id = case when p.debit_move_id = aml.id "
    "     then p.credit_move_id else p.debit_move_id end "
    "join account_move am2 on am2.id = aml2.move_id "
    "where am.state <> 'posted' order by am.name"
)
c2 = sum(d(r["amount"]) for r in draft)

# Payment menggantung: jurnal manual yang men-debit payable, belum ter-match.
gantung = q(
    "select am.name as mv, am.ref, rp.name as partner, aml.debit, aml.date, "
    "       aml.amount_residual as resid "
    "from account_move_line aml "
    "join account_account aa on aa.id = aml.account_id "
    "join account_move am on am.id = aml.move_id "
    "left join res_partner rp on rp.id = aml.partner_id "
    "where aa.account_type = 'liability_payable' and am.state = 'posted' "
    f"and am.move_type = 'entry' and aml.date <= '{DATE_TO}' "
    "and aml.reconciled = false and aml.amount_residual > 0 "
    "and coalesce(am.ref, '') not like 'EBR-GL%' order by aml.debit desc"
)

# Kembar 016 vs 045.
kembar = q(
    "select am.name as mv, am.date, am.ref, rp.name as partner, aml.debit, "
    "       aml.amount_residual as resid, aml.reconciled "
    "from account_move_line aml "
    "join account_account aa on aa.id = aml.account_id "
    "join account_move am on am.id = aml.move_id "
    "left join res_partner rp on rp.id = aml.partner_id "
    "where aa.account_type = 'liability_payable' "
    "and am.name in ('8282/2026/07/016', '8282/2026/07/045') order by am.name"
)

ebrgl = q(
    "select rp.name as partner, am.name as mv, am.ref, aml.date, "
    "       aml.debit, aml.credit, aml.amount_residual as resid "
    "from account_move_line aml "
    "join account_account aa on aa.id = aml.account_id "
    "join account_move am on am.id = aml.move_id "
    "left join res_partner rp on rp.id = aml.partner_id "
    "where aa.account_type = 'liability_payable' and am.state = 'posted' "
    f"and aml.date <= '{DATE_TO}' and aml.reconciled = false "
    "and coalesce(am.ref, '') like 'EBR-GL%' order by rp.name, am.name"
)

# --------------------------------------------------------------------------- workbook
wb = Workbook()

# ---- Sheet 1: tie-out
ws = wb.active
ws.title = "TB vs Aging"
ws["A1"] = f"Rekonsiliasi Trial Balance vs Aged Payable per {DATE_TO}"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = DB
ws["A2"].font = NOTE_FONT

head(
    ws, 4,
    ["Akun", "Nama Akun", "Trial Balance", "Aging (logika lama)",
     "Aging (residual as-of)", "Selisih C1 (dibayar setelah cut-off)"],
    [14, 38, 20, 20, 20, 24],
)
r = 5
t_tb = t_lama = t_baru = Decimal(0)
for acct in sorted(tb_m):
    ws.cell(row=r, column=1, value=acct)
    ws.cell(row=r, column=2, value=nama_m.get(acct))
    vals = (tb_m[acct], lama_m.get(acct, Decimal(0)), baru_m.get(acct, Decimal(0)))
    for col, v in zip((3, 4, 5), vals):
        c = ws.cell(row=r, column=col, value=float(v))
        c.number_format = MONEY
    c1 = vals[2] - vals[1]
    c = ws.cell(row=r, column=6, value=float(c1))
    c.number_format = MONEY
    if c1:
        c.fill = WARN_FILL
    t_tb += vals[0]
    t_lama += vals[1]
    t_baru += vals[2]
    r += 1

ws.cell(row=r, column=2, value="TOTAL").font = SUB_FONT
for col, v in ((3, t_tb), (4, t_lama), (5, t_baru), (6, t_baru - t_lama)):
    c = ws.cell(row=r, column=col, value=float(v))
    c.number_format, c.font = MONEY, SUB_FONT
r += 2

ws.cell(row=r, column=2, value="Sisa selisih TB vs aging as-of").font = SUB_FONT
c = ws.cell(row=r, column=3, value=float(t_tb - t_baru))
c.number_format, c.font = MONEY, SUB_FONT
c.fill = OK_FILL if abs(t_tb - t_baru) == c2 else WARN_FILL
r += 1
ws.cell(row=r, column=2, value="Dijelaskan oleh jurnal draft (C2)").font = SUB_FONT
c = ws.cell(row=r, column=3, value=float(-c2))
c.number_format, c.font = MONEY, SUB_FONT
r += 1
ws.cell(row=r, column=2, value="Selisih tak terjelaskan").font = SUB_FONT
sisa = (t_tb - t_baru) + c2
c = ws.cell(row=r, column=3, value=float(sisa))
c.number_format, c.font = MONEY, SUB_FONT
c.fill = OK_FILL if sisa == 0 else WARN_FILL
r += 2

note(
    ws, r,
    "C1 -- report lama memakai residual HIDUP dan membuang baris yang sudah reconciled, "
    "sehingga tagihan yang masih terbuka per cut-off tapi dibayar sesudahnya hilang dari "
    "aging. Itu sebabnya angka aging periode yang sama mengecil terus tiap kali dibuka. "
    "Kolom 'Aging (residual as-of)' adalah hasil logika baru, yang memakai posisi "
    "rekonsiliasi per tanggal cut-off.", 6,
)
r += 2
note(
    ws, r,
    "C2 -- jurnal yang belum posted tetapi masih memegang rekonsiliasi. Bill lawannya "
    "hilang dari aging (sudah dianggap lunas) padahal TB yang posted-only masih "
    "mencatatnya. Rinciannya di sheet 'Draft Reconciled'.", 6,
)

# ---- Sheet 2: draft reconciled
ws2 = wb.create_sheet("Draft Reconciled")
ws2["A1"] = "Jurnal belum posted yang masih memegang rekonsiliasi"
ws2["A1"].font = Font(bold=True, size=12)
head(ws2, 3, ["Jurnal", "State", "Tanggal", "Debit", "Ter-match ke", "State lawan",
              "Nilai match"], [22, 10, 12, 18, 30, 12, 18])
r = 4
for x in draft:
    ws2.cell(row=r, column=1, value=x["mv"])
    ws2.cell(row=r, column=2, value=x["state"]).fill = WARN_FILL
    ws2.cell(row=r, column=3, value=x["date"])
    ws2.cell(row=r, column=4, value=float(d(x["debit"]))).number_format = MONEY
    ws2.cell(row=r, column=5, value=x["lawan"])
    ws2.cell(row=r, column=6, value=x["lawan_state"])
    ws2.cell(row=r, column=7, value=float(d(x["amount"]))).number_format = MONEY
    r += 1
if not draft:
    ws2.cell(row=r, column=1, value="(tidak ada)")
    r += 1
r += 1
note(
    ws2, r,
    "Butuh keputusan Finance: posting jurnalnya (TB naik, aging tetap), atau batalkan "
    "rekonsiliasinya (TB tetap, bill kembali muncul di aging). Skrip perbaikan "
    "(90_fix_aged_payable_recon.py) sengaja tidak menyentuhnya karena kedua pilihan "
    "mengubah angka yang sudah dilaporkan.", 7,
)

# ---- Sheet 3: payment menggantung
ws3 = wb.create_sheet("Payment Menggantung")
ws3["A1"] = "Pembayaran yang men-debit akun hutang tapi tidak ter-match ke bill mana pun"
ws3["A1"].font = Font(bold=True, size=12)
head(ws3, 3, ["Jurnal", "Tanggal", "Partner", "Keterangan", "Debit", "Residual", "Status"],
     [22, 12, 36, 46, 18, 18, 34])
r = 4
STATUS = {
    "8282/2026/07/009": "diperbaiki skrip 90 -> match ke bill 00086/00087/00088",
    "8282/2026/07/017": "diperbaiki skrip 90 -> match ke bill 00014",
    "8282/2026/07/016": "BUTUH KEPUTUSAN -- lihat sheet '016 vs 045'",
}
for x in gantung:
    ws3.cell(row=r, column=1, value=x["mv"])
    ws3.cell(row=r, column=2, value=x["date"])
    ws3.cell(row=r, column=3, value=x["partner"])
    ws3.cell(row=r, column=4, value=x["ref"])
    ws3.cell(row=r, column=5, value=float(d(x["debit"]))).number_format = MONEY
    ws3.cell(row=r, column=6, value=float(d(x["resid"]))).number_format = MONEY
    st = STATUS.get(x["mv"], "belum ditinjau")
    c = ws3.cell(row=r, column=7, value=st)
    c.fill = WARN_FILL if "KEPUTUSAN" in st else OK_FILL
    r += 1
r += 1
note(
    ws3, r,
    "Ketiganya adalah jurnal manual (move_type = 'entry'), bukan vendor payment. Di "
    "account_partial_reconcile tidak ada satu pun baris untuknya -- jadi memang belum "
    "pernah dipasangkan ke tagihan. Aged Payable menampilkan setiap baris hutang yang "
    "masih terbuka, sehingga bill dan pembayarannya berdiri sebagai dua open item yang "
    "saling plus-minus, bukan saling meniadakan.", 7,
)

# ---- Sheet 4: 016 vs 045
ws4 = wb.create_sheet("016 vs 045")
ws4["A1"] = "Dua jurnal bernominal sama untuk PT Metropolitan Land Tbk."
ws4["A1"].font = Font(bold=True, size=12)
head(ws4, 3, ["Jurnal", "Tanggal", "Keterangan", "Partner", "Debit", "Residual",
              "Reconciled"], [22, 12, 46, 36, 18, 18, 12])
r = 4
for x in kembar:
    ws4.cell(row=r, column=1, value=x["mv"])
    ws4.cell(row=r, column=2, value=x["date"])
    ws4.cell(row=r, column=3, value=x["ref"])
    ws4.cell(row=r, column=4, value=x["partner"])
    ws4.cell(row=r, column=5, value=float(d(x["debit"]))).number_format = MONEY
    c = ws4.cell(row=r, column=6, value=float(d(x["resid"])))
    c.number_format = MONEY
    c.fill = WARN_FILL if d(x["resid"]) else OK_FILL
    ws4.cell(row=r, column=7, value="ya" if x["reconciled"] == "t" else "tidak")
    r += 1
r += 1
note(
    ws4, r,
    "Nominal, tanggal dan partner keduanya sama persis; 045 sudah ter-match ke tagihannya "
    "sedangkan 016 menggantung penuh. Kemungkinan dobel-catat, atau 016 memang deposit "
    "yang belum ada tagihannya. Mohon dikonfirmasi ke Finance sebelum dijurnal balik atau "
    "dipasangkan.", 7,
)

# ---- Sheet 5: EBR-GL
ws5 = wb.create_sheet("EBR-GL")
ws5["A1"] = "Baris beginning balance (Bill Reference berawalan EBR-GL) di akun hutang"
ws5["A1"].font = Font(bold=True, size=12)
head(ws5, 3, ["Partner", "Jurnal", "Reference", "Tanggal", "Debit", "Credit", "Residual"],
     [36, 22, 24, 12, 18, 18, 18])
r = 4
per_partner = {}
t_res = Decimal(0)
for x in ebrgl:
    ws5.cell(row=r, column=1, value=x["partner"])
    ws5.cell(row=r, column=2, value=x["mv"])
    ws5.cell(row=r, column=3, value=x["ref"])
    ws5.cell(row=r, column=4, value=x["date"])
    for col, key in ((5, "debit"), (6, "credit"), (7, "resid")):
        ws5.cell(row=r, column=col, value=float(d(x[key]))).number_format = MONEY
    per_partner.setdefault(x["partner"], Decimal(0))
    per_partner[x["partner"]] += d(x["resid"])
    t_res += d(x["resid"])
    r += 1
ws5.cell(row=r, column=4, value=f"TOTAL ({len(ebrgl)} baris)").font = SUB_FONT
c = ws5.cell(row=r, column=7, value=float(t_res))
c.number_format, c.font = MONEY, SUB_FONT
c.fill = OK_FILL if t_res == 0 else WARN_FILL
r += 2
tidak_nol = {k: v for k, v in per_partner.items() if v != 0}
note(
    ws5, r,
    "Net seluruh baris = %s. Sisi tagihan (BILL/2026/06/xxxx) dan sisi bank "
    "(BNK1/2026/xxxxx) dari upload beginning balance dua-duanya masuk ke akun hutang "
    "tanpa pernah saling direkonsiliasi, jadi keduanya tampil sebagai open item. "
    "Skrip 90 merekonsiliasinya per partner -- baris hilang dari aging karena memang "
    "sudah nol, bukan karena disembunyikan lewat filter reference. Partner dengan net "
    "bukan nol: %s."
    % (
        f"{t_res:,.2f}",
        ", ".join(f"{k} ({v:,.2f})" for k, v in tidak_nol.items()) or "tidak ada",
    ),
    7,
)

wb.save(OUT)
os.chmod(OUT, 0o644)

print(f"Database        : {DB}   cut-off {DATE_TO}")
print(f"Trial Balance   : {t_tb:>20,.2f}")
print(f"Aging lama      : {t_lama:>20,.2f}   selisih C1 {t_baru - t_lama:>20,.2f}")
print(f"Aging as-of     : {t_baru:>20,.2f}   selisih C2 {-c2:>20,.2f}")
print(f"Tak terjelaskan : {sisa:>20,.2f}")
print(f"Payment gantung : {len(gantung)}    baris EBR-GL: {len(ebrgl)} (net {t_res:,.2f})")
print(f"Tersimpan       : {OUT}")
