"""Re-verify the June-2026 AR reconciliation against FICO's final workbook and
build the approval workbook for the adjustment journals.

Runs on the HOST (plain python3 + openpyxl); the database is read through
`docker exec ... psql`, so nothing here touches the Odoo ORM and nothing is
written to the database.

    python3 scripts/tenants/levis/67_rekon_ar_juni_final.py \
        --fico /path/to/'EBR - Check Selisih Outstanding AR Finance vs Accounting Juni 2026.xlsx' \
        --db prd_levis_begbal \
        --out /srv/sftp-share/files/Draft_Adjustment_AR_Juni2026_v2.xlsx

FICO workbook layout (sheet "Summary", rows 4..26):
    I/J  Trade Receivables - Third Parties (Accounting)   per store
    L/M  Deposit from customer trade (Accounting)         per store
    O/P  Net off = J + M
    R/S  Outstanding AR (Finance)
    U    Difference = P - S
    V    MDR not yet journalised by Accounting in June
    W    Bank-in without store info (re-assignment)
    X    Manual sales per Finance's reconciliation
    Y    Difference = U - V + W + X
    Z    Plus-minus (settlement booked against the wrong store)
    AA   Difference = Y - Z   (residual, left unadjusted)
"""

import argparse
import os
import subprocess
import sys
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# --- accounts / constants --------------------------------------------------
COMPANY_ID = 1
ACC_AR = "1106000001"  # Trade Receivables - Third Parties
ACC_DEPOSIT = "2103100003"  # Deposit from customer trade
ACC_MDR = "7104000001"  # MDR Bank
ACC_SUSPENSE = "1103000002"  # Bank Suspense Account
ACC_BCA = "1103019310"  # BCA-IDR-2685151268-MB
CUTOFF = "2026-06-30"

# FICO writes "OLS SES - HEAD QUARTER"; the analytic account is named differently.
OU_ALIAS = {"OLS SES - HEAD QUARTER": "EBR - HEAD OFFICE"}

# Bank-in rows Finance could not attribute to a store by the closing deadline
# ("BANK IN TIDAK ADA INFO PER 09 JULI 2026", sheet "Selisih MDR Mandiri" G34:I38).
# They sit on the HEAD OFFICE OU in Odoo and belong to these stores.
NO_STORE_SPLIT = OrderedDict(
    [
        ("OLS SES - GANDARIA CITY", 5140545),
        ("OLS SES - TRANS STUDIO CIBUBUR", 1798898),
        ("OLS SES - GRAND INDONESIA", 1),
    ]
)
NO_STORE_OU = "EBR - HEAD OFFICE"

# Settlements booked against the wrong store (FICO column Z): (from_ou, to_ou, amount)
PLUS_MINUS = [
    ("OLS SES - METROPOLITAN MALL BEKASI", "OLS SES - GRAND METROPOLITAN BEKASI", 1397725),
    ("OLS SES - TRANS STUDIO CIBUBUR", "OLS SES - TRANS STUDIO MALL BANDUNG", 1275795),
]

TITLE = Font(bold=True, size=13)
HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="4F6228")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="F8CBAD")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
MONEY = "#,##0"


def q(db, sql):
    """Run a read-only query and return a list of tuples of strings."""
    pw = subprocess.check_output(
        ["docker", "exec", "odoo19-platform-odoo", "printenv", "POSTGRES_PASSWORD"], text=True
    ).strip()
    env = dict(os.environ, PGPASSWORD=pw)
    out = subprocess.check_output(
        [
            "docker",
            "exec",
            "-e",
            "PGPASSWORD=" + pw,
            "odoo19-platform-postgres",
            "psql",
            "-U",
            "odoo",
            "-d",
            db,
            "-At",
            "-F",
            "\t",
            "-c",
            sql,
        ],
        text=True,
        env=env,
    )
    return [tuple(line.split("\t")) for line in out.splitlines() if line]


def db_snapshot(db):
    """Balances and per-OU deposit split straight from the ledger."""
    codes = "','".join([ACC_AR, ACC_DEPOSIT, ACC_MDR, ACC_SUSPENSE, ACC_BCA])
    rows = q(
        db,
        """select (a.code_store->>'%s'), round(sum(l.debit-l.credit)::numeric,2), count(*)
             from account_move_line l
             join account_account a on a.id=l.account_id
             join account_move m on m.id=l.move_id
            where (a.code_store->>'%s') in ('%s')
              and m.state='posted' and m.date <= '%s'
            group by 1"""
        % (COMPANY_ID, COMPANY_ID, codes, CUTOFF),
    )
    bal = {c: (float(b), int(n)) for c, b, n in rows}

    rows = q(
        db,
        """select coalesce(aa.name->>'en_US','(tanpa OU)'),
                  round(sum(l.debit-l.credit)::numeric,2), count(*)
             from account_move_line l
             join account_account a on a.id=l.account_id
             join account_move m on m.id=l.move_id
             left join lateral (select (jsonb_object_keys(l.analytic_distribution))::int aid) k on true
             left join account_analytic_account aa on aa.id=k.aid
            where (a.code_store->>'%s')='%s' and m.state='posted' and m.date <= '%s'
            group by 1"""
        % (COMPANY_ID, ACC_DEPOSIT, CUTOFF),
    )
    deposit_by_ou = {n: (float(b), int(c)) for n, b, c in rows}

    rows = q(
        db,
        "select st.id, st.name, st.date, count(l.id) from account_bank_statement st "
        "join account_bank_statement_line l on l.statement_id=st.id group by 1,2,3 order by st.id",
    )
    statements = [(int(i), n, d, int(c)) for i, n, d, c in rows]
    return bal, deposit_by_ou, statements


def read_fico(path):
    ws = openpyxl.load_workbook(path, data_only=True)["Summary"]
    stores = []
    for r in range(4, 27):
        name = ws.cell(r, 18).value  # R = Outstanding AR (Finance) row label
        if not name:
            continue
        rec = dict(
            ou=OU_ALIAS.get(str(name).strip(), str(name).strip()),
            fico_label=str(name).strip(),
            ar=float(ws.cell(r, 10).value or 0),  # J
            deposit=float(ws.cell(r, 13).value or 0),  # M
            finance=float(ws.cell(r, 19).value or 0),  # S
            mdr=float(ws.cell(r, 22).value or 0),  # V
            no_store=float(ws.cell(r, 23).value or 0),  # W
            sales_manual=float(ws.cell(r, 24).value or 0),  # X
            plus_minus=float(ws.cell(r, 26).value or 0),  # Z
            residual=float(ws.cell(r, 27).value or 0),  # AA
            note=ws.cell(r, 29).value or "",  # AC
            is_total=str(name).strip().upper() == "GRAND TOTAL",
        )
        stores.append(rec)
    total = [s for s in stores if s["is_total"]][0]
    stores = [s for s in stores if not s["is_total"]]
    return stores, total


def build_plan(stores):
    """Corrected TBFU + clearing per store, following the FICO decomposition."""
    by_ou = {s["ou"]: s for s in stores}
    tbfu = {s["ou"]: -s["deposit"] for s in stores}  # positive = money received

    # 1. spread the "no store info" bank-in off HEAD OFFICE
    for ou, amt in NO_STORE_SPLIT.items():
        tbfu[ou] = tbfu.get(ou, 0.0) + amt
        tbfu[NO_STORE_OU] = tbfu.get(NO_STORE_OU, 0.0) - amt
    # 2. move settlements booked against the wrong store
    for src, dst, amt in PLUS_MINUS:
        tbfu[src] -= amt
        tbfu[dst] = tbfu.get(dst, 0.0) + amt

    plan = []
    for ou in sorted(tbfu, key=lambda o: -tbfu[o]):
        s = by_ou.get(ou, {})
        ar_adj = s.get("ar", 0.0) + s.get("sales_manual", 0.0) - s.get("mdr", 0.0)
        clearing = min(tbfu[ou], ar_adj) if tbfu[ou] > 0 else 0.0
        plan.append(
            dict(
                ou=ou,
                tbfu_raw=-s.get("deposit", 0.0),
                tbfu=tbfu[ou],
                ar=s.get("ar", 0.0),
                mdr=s.get("mdr", 0.0),
                sales_manual=s.get("sales_manual", 0.0),
                ar_adj=ar_adj,
                clearing=clearing,
                sisa=ar_adj - clearing,
                finance=s.get("finance", 0.0),
                selisih=(ar_adj - clearing) - s.get("finance", 0.0),
                note=s.get("note", ""),
            )
        )
    return plan


def sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    return ws


def head(ws, row, labels):
    for c, lbl in enumerate(labels, start=1):
        cell = ws.cell(row, c, lbl)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def money(ws, row, col, value, fill=None):
    cell = ws.cell(row, col, round(value, 2))
    cell.number_format = MONEY
    if fill:
        cell.fill = fill
    return cell


def build_workbook(out, db, bal, deposit_by_ou, statements, stores, total, plan):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tot_clearing = sum(p["clearing"] for p in plan)
    tot_mdr = sum(p["mdr"] for p in plan)
    tot_sm = sum(p["sales_manual"] for p in plan)

    # --- Ringkasan --------------------------------------------------------
    ws = sheet(wb, "Ringkasan & Usulan", [62, 20, 60])
    ws["A1"] = "DRAFT ADJUSTMENT PIUTANG JUNI 2026 (v2 — atas rekon final FICO)"
    ws["A1"].font = TITLE
    rows = [
        ("Database", db, ""),
        ("Sumber", "EBR - Check Selisih Outstanding AR Finance vs Accounting Juni 2026.xlsx", ""),
        ("", "", ""),
        ("POSISI SEKARANG (buku Odoo, posted s/d 30-Jun)", "", ""),
        ("  1106000001 Trade Receivables", bal[ACC_AR][0], "cocok dengan kolom Accounting FICO"),
        ("  2103100003 Deposit from customer trade", bal[ACC_DEPOSIT][0], "cocok per toko dengan FICO"),
        ("  7104000001 MDR Bank", bal[ACC_MDR][0], "cocok dengan MDR Accounting FICO"),
        ("", "", ""),
        ("SELISIH ACCOUNTING vs FINANCE (kolom U FICO)", total["ar"] + total["deposit"] - total["finance"], ""),
        ("  dijelaskan oleh: MDR belum dijurnal (V)", -tot_mdr, "jurnal #3"),
        ("  dijelaskan oleh: bank-in tanpa info toko (W)", 0, "jurnal #2 — realokasi OU saja"),
        ("  dijelaskan oleh: sales manual rekon Finance (X)", tot_sm, "jurnal #4"),
        ("  dijelaskan oleh: salah toko / plus-minus (Z)", 0, "jurnal #2 — realokasi OU saja"),
        ("  SISA TIDAK DIJURNAL (rounding + adj bank)", total["residual"], "lihat sheet '6. Sisa selisih'"),
        ("", "", ""),
        ("USULAN JURNAL (tanggal 30-Jun-2026, dibuat DRAFT)", "", ""),
        ("  1. Clearing AR: Dr 2103100003 / Cr 1106000001", tot_clearing, "per toko, dengan OU"),
        ("  2. Realokasi OU deposit (tanpa efek saldo akun)", 0, "HQ→3 toko + 2 pasang salah toko"),
        ("  3. MDR belum dijurnal: Dr 7104000001 / Cr 1106000001", tot_mdr, "16 toko"),
        ("  4. Sales manual: Dr 1106000001 / Cr pendapatan", tot_sm, "PERLU KONFIRMASI akun + PPN"),
        ("", "", ""),
        ("HASIL AKHIR PIUTANG 1106000001", bal[ACC_AR][0] + tot_sm - tot_mdr - tot_clearing, ""),
        ("  Outstanding AR versi Finance", total["finance"], ""),
        ("  Selisih tersisa", bal[ACC_AR][0] + tot_sm - tot_mdr - tot_clearing - total["finance"], "= sisa rounding"),
        (
            "  Sisa saldo 2103100003 setelah clearing",
            -(bal[ACC_DEPOSIT][0] + tot_clearing),
            "Grand Indonesia Rp 1 — piutang tokonya nol, tidak bisa di-clear (FICO juga mencatat -1)",
        ),
        ("", "", ""),
        ("PEMBERSIHAN DUPLIKASI (sudah/akan dieksekusi)", "", "lihat sheet '5. Pembersihan duplikasi'"),
    ]
    r = 3
    for label, val, note in rows:
        ws.cell(r, 1, label)
        if isinstance(val, (int, float)):
            money(ws, r, 2, val)
        else:
            ws.cell(r, 2, val)
        ws.cell(r, 3, note)
        if label.startswith(("POSISI", "SELISIH", "USULAN", "HASIL", "PEMBERSIHAN")):
            ws.cell(r, 1).font = Font(bold=True)
        r += 1

    # --- 1. Clearing ------------------------------------------------------
    ws = sheet(wb, "1. Clearing AR", [38, 16, 14, 16, 16, 14, 14, 16, 16, 16, 14, 46])
    head(
        ws,
        1,
        [
            "Operating Unit",
            "TBFU per buku (deposit)",
            "Koreksi toko",
            "TBFU terkoreksi",
            "Piutang per Accounting",
            "MDR (jurnal #3)",
            "Sales manual (jurnal #4)",
            "Piutang setelah #3/#4",
            "CLEARING (Dr 2103100003 / Cr 1106000001)",
            "Sisa piutang",
            "Outstanding Finance",
            "Selisih vs Finance / catatan",
        ],
    )
    r = 2
    for p in plan:
        ws.cell(r, 1, p["ou"])
        money(ws, r, 2, p["tbfu_raw"])
        money(ws, r, 3, p["tbfu"] - p["tbfu_raw"], WARN_FILL if p["tbfu"] != p["tbfu_raw"] else None)
        money(ws, r, 4, p["tbfu"])
        money(ws, r, 5, p["ar"])
        money(ws, r, 6, -p["mdr"])
        money(ws, r, 7, p["sales_manual"])
        money(ws, r, 8, p["ar_adj"])
        money(ws, r, 9, p["clearing"], OK_FILL)
        money(ws, r, 10, p["sisa"])
        money(ws, r, 11, p["finance"])
        money(ws, r, 12, p["selisih"], BAD_FILL if abs(p["selisih"]) > 1 else None)
        r += 1
    ws.cell(r, 1, "TOTAL").font = Font(bold=True)
    for col, key in (
        (2, "tbfu_raw"),
        (4, "tbfu"),
        (5, "ar"),
        (7, "sales_manual"),
        (8, "ar_adj"),
        (9, "clearing"),
        (10, "sisa"),
        (11, "finance"),
        (12, "selisih"),
    ):
        money(ws, r, col, sum(p[key] for p in plan)).font = Font(bold=True)
    money(ws, r, 6, -sum(p["mdr"] for p in plan)).font = Font(bold=True)

    # --- 2. Realokasi -----------------------------------------------------
    ws = sheet(wb, "2. Realokasi OU", [40, 40, 18, 62])
    ws["A1"] = "Realokasi analytic Operating Unit pada akun 2103100003 (saldo akun tidak berubah)"
    ws["A1"].font = TITLE
    head(ws, 3, ["Dari OU (Dr 2103100003)", "Ke OU (Cr 2103100003)", "Jumlah", "Dasar"])
    r = 4
    for ou, amt in NO_STORE_SPLIT.items():
        ws.cell(r, 1, NO_STORE_OU)
        ws.cell(r, 2, ou)
        money(ws, r, 3, amt)
        ws.cell(r, 4, "Bank-in tanpa info toko per 9-Jul (sheet 'Selisih MDR Mandiri')")
        r += 1
    for src, dst, amt in PLUS_MINUS:
        ws.cell(r, 1, src)
        ws.cell(r, 2, dst)
        money(ws, r, 3, amt)
        ws.cell(r, 4, "Settlement tercatat di toko yang salah (kolom Z 'Plus-Minus')")
        r += 1
    ws.cell(r, 2, "TOTAL").font = Font(bold=True)
    money(ws, r, 3, sum(NO_STORE_SPLIT.values()) + sum(a for _, _, a in PLUS_MINUS)).font = Font(bold=True)

    # --- 3. MDR -----------------------------------------------------------
    ws = sheet(wb, "3. MDR belum jurnal", [40, 18, 52])
    ws["A1"] = "MDR Juni yang belum dijurnal Accounting — Dr 7104000001 / Cr 1106000001"
    ws["A1"].font = TITLE
    head(ws, 3, ["Operating Unit", "Jumlah", "Catatan"])
    r = 4
    for p in sorted(plan, key=lambda x: -x["mdr"]):
        if not p["mdr"]:
            continue
        ws.cell(r, 1, p["ou"])
        money(ws, r, 2, p["mdr"])
        ws.cell(r, 3, "kolom V FICO")
        r += 1
    ws.cell(r, 1, "TOTAL").font = Font(bold=True)
    money(ws, r, 2, tot_mdr).font = Font(bold=True)

    # --- 4. Sales manual --------------------------------------------------
    ws = sheet(wb, "4. Sales manual", [40, 18, 70])
    ws["A1"] = "Penjualan manual versi rekon Finance yang belum ada di Odoo"
    ws["A1"].font = TITLE
    ws["A2"] = "PERLU KONFIRMASI ACCOUNTING: akun pendapatan lawan + perlakuan PPN sebelum dijurnal"
    ws["A2"].fill = WARN_FILL
    head(ws, 4, ["Operating Unit", "Jumlah", "Catatan"])
    r = 5
    for p in sorted(plan, key=lambda x: -x["sales_manual"]):
        if not p["sales_manual"]:
            continue
        ws.cell(r, 1, p["ou"])
        money(ws, r, 2, p["sales_manual"])
        ws.cell(r, 3, p["note"] or "kolom X FICO (Rekon Finance)")
        r += 1
    ws.cell(r, 1, "TOTAL").font = Font(bold=True)
    money(ws, r, 2, tot_sm).font = Font(bold=True)

    # --- 5. Duplikasi -----------------------------------------------------
    ws = sheet(wb, "5. Pembersihan duplikasi", [46, 22, 22, 60])
    ws["A1"] = "Pembersihan duplikasi Bank Suspense BCA Juni (Opsi A — hapus statement import)"
    ws["A1"].font = TITLE
    head(ws, 3, ["Item", "Sebelum", "Sesudah (target)", "Catatan"])
    dup = [s for s in statements if s[0] == 12]
    n_dup = dup[0][3] if dup else 0
    data = [
        (
            "Bank statement id 12 (Imported, 30-Jun)",
            "%d baris" % n_dup,
            "dihapus",
            "duplikat 1:1 dari EBR GL load di jurnal BNK1",
        ),
        (
            "Saldo 1103000002 Bank Suspense (s/d 30-Jun)",
            bal.get(ACC_SUSPENSE, (0, 0))[0],
            0,
            "seluruh baris berasal dari statement 12",
        ),
        (
            "Saldo 1103019310 BCA (s/d 30-Jun)",
            bal.get(ACC_BCA, (0, 0))[0],
            bal.get(ACC_BCA, (0, 0))[0] - bal.get(ACC_SUSPENSE, (0, 0))[0] * -1,
            "berhenti double-count arus kas Juni",
        ),
        (
            "Statement Juli (id 17-20)",
            "tidak disentuh",
            "tidak disentuh",
            "EBR GL berhenti 30-Jun, jadi bukan duplikat",
        ),
    ]
    r = 4
    for label, before, after, note in data:
        ws.cell(r, 1, label)
        for col, val in ((2, before), (3, after)):
            if isinstance(val, (int, float)):
                money(ws, r, col, val)
            else:
                ws.cell(r, col, val)
        ws.cell(r, 4, note)
        r += 1

    # --- 6. Sisa selisih --------------------------------------------------
    ws = sheet(wb, "6. Sisa selisih", [40, 18, 70])
    ws["A1"] = "Sisa selisih yang TIDAK dijurnal (kolom AA FICO)"
    ws["A1"].font = TITLE
    head(ws, 3, ["Operating Unit", "Sisa", "Catatan"])
    r = 4
    for s in sorted(stores, key=lambda x: x["residual"]):
        if abs(s["residual"]) < 0.5:
            continue
        ws.cell(r, 1, s["ou"])
        money(ws, r, 2, s["residual"])
        ws.cell(r, 3, s["note"] or "rounding")
        r += 1
    ws.cell(r, 1, "TOTAL").font = Font(bold=True)
    money(ws, r, 2, total["residual"]).font = Font(bold=True)

    # --- Verifikasi -------------------------------------------------------
    ws = sheet(wb, "Verifikasi DB vs FICO", [40, 20, 20, 16, 14])
    ws["A1"] = "Deposit 2103100003 per Operating Unit: buku Odoo vs kolom FICO"
    ws["A1"].font = TITLE
    head(ws, 3, ["Operating Unit", "Odoo (posted s/d 30-Jun)", "FICO kolom M", "Selisih", "Status"])
    r = 4
    for s in sorted(stores, key=lambda x: x["deposit"]):
        odoo = deposit_by_ou.get(s["ou"], (0.0, 0))[0]
        diff = odoo - s["deposit"]
        ws.cell(r, 1, s["ou"])
        money(ws, r, 2, odoo)
        money(ws, r, 3, s["deposit"])
        money(ws, r, 4, diff)
        c = ws.cell(r, 5, "OK" if abs(diff) < 0.5 else "BEDA")
        c.fill = OK_FILL if abs(diff) < 0.5 else BAD_FILL
        r += 1
    r += 1
    for label, odoo, fico in (
        ("1106000001 Trade Receivables", bal[ACC_AR][0], total["ar"]),
        ("2103100003 Deposit", bal[ACC_DEPOSIT][0], total["deposit"]),
        ("7104000001 MDR Bank", bal[ACC_MDR][0], 25971461.73),
    ):
        ws.cell(r, 1, label).font = Font(bold=True)
        money(ws, r, 2, odoo)
        money(ws, r, 3, fico)
        money(ws, r, 4, odoo - fico)
        c = ws.cell(r, 5, "OK" if abs(odoo - fico) < 0.5 else "BEDA")
        c.fill = OK_FILL if abs(odoo - fico) < 0.5 else BAD_FILL
        r += 1

    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fico", required=True)
    ap.add_argument("--db", default="prd_levis_begbal")
    ap.add_argument("--out")
    args = ap.parse_args()

    stores, total = read_fico(args.fico)
    bal, deposit_by_ou, statements = db_snapshot(args.db)
    plan = build_plan(stores)

    print("== %s: saldo buku vs FICO ==" % args.db)
    checks = [
        ("1106000001", bal[ACC_AR][0], total["ar"]),
        ("2103100003", bal[ACC_DEPOSIT][0], total["deposit"]),
        ("7104000001", bal[ACC_MDR][0], 25971461.73),
    ]
    bad = 0
    for code, odoo, fico in checks:
        ok = abs(odoo - fico) < 0.5
        bad += 0 if ok else 1
        print("  %s odoo=%15.2f fico=%15.2f %s" % (code, odoo, fico, "OK" if ok else "<-- BEDA"))

    print("== deposit per OU ==")
    for s in stores:
        odoo = deposit_by_ou.get(s["ou"], (0.0, 0))[0]
        if abs(odoo - s["deposit"]) >= 0.5:
            bad += 1
            print("  BEDA %-38s odoo=%14.2f fico=%14.2f" % (s["ou"], odoo, s["deposit"]))
    print("  %d OU dibandingkan, %d beda" % (len(stores), bad))

    print("== rekonstruksi rumus FICO (Y = U - V + W + X, sisa = Y - Z) ==")
    for s in stores:
        u = s["ar"] + s["deposit"] - s["finance"]
        y = u - s["mdr"] + s["no_store"] + s["sales_manual"]
        if abs(y - s["plus_minus"] - s["residual"]) >= 0.5:
            bad += 1
            print("  TIDAK COCOK %-38s hitung=%12.2f sheet=%12.2f" % (s["ou"], y - s["plus_minus"], s["residual"]))

    tot_clearing = sum(p["clearing"] for p in plan)
    tot_mdr = sum(p["mdr"] for p in plan)
    tot_sm = sum(p["sales_manual"] for p in plan)
    akhir = bal[ACC_AR][0] + tot_sm - tot_mdr - tot_clearing
    print("== rencana ==")
    print("  clearing        %15.2f" % tot_clearing)
    print("  MDR             %15.2f" % tot_mdr)
    print("  sales manual    %15.2f" % tot_sm)
    print(
        "  AR akhir        %15.2f  (Finance %15.2f, selisih %10.2f)"
        % (akhir, total["finance"], akhir - total["finance"])
    )
    for p in plan:
        if abs(p["selisih"]) > 1:
            print("  sisa != Finance: %-38s %12.2f" % (p["ou"], p["selisih"]))

    if args.out:
        build_workbook(args.out, args.db, bal, deposit_by_ou, statements, stores, total, plan)
        print("workbook -> %s" % args.out)
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
